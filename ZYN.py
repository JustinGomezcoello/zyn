import decimal
import re
import tkinter as tk
from decimal import Decimal, InvalidOperation
import pyodbc
import datetime
from tkcalendar import DateEntry
from tkinter import messagebox
import os
import sys
from tkinter import ttk, messagebox, simpledialog
import uuid
import hashlib
import json
import pandas as pd
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from tkinter import filedialog
from datetime import date





# 🔑 Clave secreta para cifrar y validar la licencia
SECRET_KEY = "secretcode"

# 📁 Archivo donde se guarda la licencia
LICENSE_FILE = "license.key"

def get_mac_address():
    """Obtiene la dirección MAC de la PC."""
    mac = uuid.getnode()
    mac_str = ':'.join(("%012X" % mac)[i:i+2] for i in range(0, 12, 2))
    return mac_str

def generate_license_for_mac(mac, expiration_date):
    """Genera un código de licencia cifrado basado en la MAC y la fecha de expiración."""
    data = f"{mac}|{expiration_date}|{SECRET_KEY}"
    return hashlib.sha256(data.encode('utf-8')).hexdigest().upper()

def save_license(expiration_date):
    """Guarda la licencia cifrada en el archivo license.key."""
    mac = get_mac_address()
    license_code = generate_license_for_mac(mac, expiration_date)
    
    data = {
        "mac": mac,
        "expiration_date": expiration_date,
        "license_code": license_code
    }

    with open(LICENSE_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f)

def check_license():
    """Verifica la licencia y la validez de la fecha."""
    if not os.path.exists(LICENSE_FILE):
        return request_license()

    try:
        with open(LICENSE_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)

        mac = get_mac_address()
        expiration_date = data.get("expiration_date")
        stored_license = data.get("license_code")

        # 🔍 Verificar que la licencia no haya sido alterada
        expected_license = generate_license_for_mac(mac, expiration_date)
        if stored_license != expected_license:
            messagebox.showerror("Error de Seguridad", "¡Se detectó manipulación en la licencia! El software se cerrará.")
            os.remove(LICENSE_FILE)
            sys.exit(0)

        # 📅 Comprobar si la licencia ha expirado
        today = datetime.date.today()
        expiration = datetime.datetime.strptime(expiration_date, "%Y-%m-%d").date()
        
        days_left = (expiration - today).days

        if days_left < 0:
            messagebox.showerror("Licencia Expirada", "Su licencia ha expirado. Contacte al administrador para renovarla.")
            os.remove(LICENSE_FILE)
            sys.exit(0)
        elif days_left <= 5:
            messagebox.showwarning("Pago Pendiente", f"Su licencia expira en {days_left} días. ¡Realice el pago pronto!")

        return True

    except Exception as e:
        messagebox.showerror("Error", f"Error al verificar la licencia: {e}")
        sys.exit(0)

def request_license():
    """Solicita al usuario ingresar una licencia válida."""
    mac = get_mac_address()
    today = datetime.date.today()

    # 📝 Pedir días al usuario
    dias_validez = simpledialog.askinteger("Días de Licencia", "¿Cuántos días de validez tendrá la licencia?", minvalue=1, maxvalue=999)

    if dias_validez is None:
        messagebox.showerror("Cancelado", "Debe ingresar un número de días para la licencia.")
        sys.exit(0)

    expiration_date = (today + datetime.timedelta(days=dias_validez)).strftime("%Y-%m-%d")

    
    license_code = simpledialog.askstring("Licencia", "Ingrese su código de licencia:")
    
    if not license_code:
        messagebox.showerror("Licencia Requerida", "Debe ingresar una licencia para continuar.")
        sys.exit(0)

    expected_license = generate_license_for_mac(mac, expiration_date)
    
    if license_code.strip().upper() == expected_license:
        save_license(expiration_date)
        messagebox.showinfo("Licencia Aceptada", "¡Licencia válida! Puede usar el programa.")
        return True
    else:
        messagebox.showerror("Licencia Inválida", "El código ingresado no es válido para esta computadora.")
        sys.exit(0)

# ---------------------- INICIO DE APLICACIÓN ----------------------

root = tk.Tk()
root.withdraw()  # Ocultar ventana mientras se valida

# Validación de licencia
if not check_license():
    root.destroy()
    sys.exit(0)

root.deiconify()  # Mostrar ventana principal

# ---------------------- CONEXIÓN SQL ----------------------
conn = pyodbc.connect(
    "DRIVER={ODBC Driver 17 for SQL Server};"
    "SERVER=localhost\\SQLEXPRESS;"
    "DATABASE=ZYN;"
    "Trusted_Connection=yes;"
)
cursor = conn.cursor()

# ---------------------- VENTANA PRINCIPAL ----------------------
root.title("ZYN")
root.geometry("1920x1080")
root.configure(bg="#E8F0FE")  # Cambiado a un azul claro suave

# Aplicar un tema a la interfaz
style = ttk.Style()
style.theme_use("clam")
style.configure("TFrame", background="#E8F0FE")
style.configure("TNotebook", background="#E8F0FE")
style.configure("TNotebook.Tab", background="#D0DFF9", padding=[10, 2])
style.configure("TLabel", background="#E8F0FE")
style.map('TNotebook.Tab', background=[("selected", "#E8F0FE")])

# ================== CONTENEDOR PRINCIPAL ===================
frame_principal = ttk.Frame(root, padding=10)
frame_principal.pack(fill="both", expand=True)

# ======= NOTEBOOK PARA PESTAÑAS =======
tab_control = ttk.Notebook(frame_principal)

# Pestañas principales
tab_compras = ttk.Frame(tab_control, padding=10)
tab_orden_compra = ttk.Frame(tab_control, padding=10)
tab_cuentas_cobrar = ttk.Frame(tab_control, padding=10)
tab_cuentas_pagar = ttk.Frame(tab_control, padding=10)
tab_prestamos = ttk.Frame(tab_control, padding=10)
tab_reportes = ttk.Frame(tab_control, padding=10)

# Agregar pestañas al control
tab_control.add(tab_compras, text='📦 Compras')
tab_control.add(tab_orden_compra, text='🧾 Orden de Compra')
tab_control.add(tab_cuentas_cobrar, text='💰 Cuentas por Cobrar')
tab_control.add(tab_cuentas_pagar, text='💸 Cuentas por Pagar')
tab_control.add(tab_prestamos, text='📤 Préstamos / Devoluciones')
tab_control.add(tab_reportes, text='📊 Reportes')

tab_control.pack(expand=1, fill="both")



# ---------------------- PESTAÑA COMPRAS ----------------------


def agregar_compra():
    try:
        fecha_compra = entry_fecha_compra.get_date().strftime("%Y-%m-%d")
        codigo = entry_codigo_compra.get().strip()
        cantidad = entry_cantidad_compra.get().strip()
        proveedor = entry_proveedor.get().strip()
        # Validar que los campos no estén vacíos
        if not fecha_compra or not codigo or not cantidad or not proveedor:
            messagebox.showwarning("Advertencia", "Debe ingresar todos los campos requeridos.")
            return

        # Validar formato de fecha (YYYY-MM-DD)
        if not re.match(r"^\d{4}-\d{2}-\d{2}$", fecha_compra):
            messagebox.showerror("Error", "El formato de la Fecha de Compra es incorrecto. Use YYYY-MM-DD.")
            return

        # Convertir cantidad a entero
        cantidad = int(cantidad)
        if cantidad <= 0:
            messagebox.showerror("Error", "La cantidad debe ser un número entero positivo.")
            return

        # Obtener datos del producto (NombreProducto, CostoConIVA, IVA)
        cursor.execute("SELECT NombreProducto, CostoConIVA, IVA, CantidadVendida, CantidadInicial FROM Productos WHERE CodigoProducto = ?", (codigo,))
        producto = cursor.fetchone()

        if not producto:
            messagebox.showerror("Error", f"El código '{codigo}' no existe en la tabla Productos.")
            return

        nombre_producto, costo_con_iva, iva, cantidad_vendida, cantidad_inicial = producto

        # Calcular costos e IVA basados en la cantidad
        costo_con_iva_total = costo_con_iva * cantidad
        valor_iva = costo_con_iva_total - (costo_con_iva_total / (1 + iva))
        costo_sin_iva = costo_con_iva_total - valor_iva

        # Insertar la compra en la BD, incluyendo Proveedor
        cursor.execute("""
            INSERT INTO Compras (FechaCompra, CodigoProducto, NombreProducto, CantidadComprada, CostoSinIVA, PorcentajeIVA, IVA, CostoConIVA, Proveedor)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (fecha_compra, codigo, nombre_producto, cantidad, costo_sin_iva, iva, valor_iva, costo_con_iva_total, proveedor))

        # Actualizar la cantidad inicial y el inventario en Productos
        nueva_cantidad_inicial = cantidad_inicial + cantidad
        nueva_cantidad_inventario = nueva_cantidad_inicial - cantidad_vendida

        cursor.execute("""
            UPDATE Productos
            SET CantidadInicial = ?, CantidadInventario = ?
            WHERE CodigoProducto = ?
        """, (nueva_cantidad_inicial, nueva_cantidad_inventario, codigo))

        conn.commit()

        messagebox.showinfo("Éxito", "Compra registrada y cantidad actualizada correctamente.")

        # Limpiar los campos después de registrar la compra
        from datetime import date
        entry_fecha_compra.set_date(date.today())

        entry_codigo_compra.delete(0, tk.END)
        entry_cantidad_compra.delete(0, tk.END)
        entry_nombre_producto_compra.delete(0, tk.END)
        entry_proveedor.delete(0, tk.END)

    except Exception as e:
        messagebox.showerror("Error", f"Error al registrar la compra: {str(e)}")


# Aplicar estilos
style = ttk.Style()
style.configure("TButton", font=("Arial", 10, "bold"),  background="#4CAF50", foreground="white")
style.configure("DateEntry", font=("Arial", 10))
# 🔴 Botón Rojo (Eliminar)
style.configure("Red.TButton", foreground="white", background="red", font=("Arial", 10, "bold"))
style.map("Red.TButton", background=[("active", "darkred")])

# 🔵 Botón Celeste (Confirmar Venta)
style.configure("Blue.TButton", foreground="white", background="#00BFFF", font=("Arial", 10, "bold"))
style.map("Blue.TButton", background=[("active", "#008CBA")])  # Azul más oscuro al pasar el mouse

# 🟡 Botón Amarillo (Modificar Venta)
style.configure("Yellow.TButton", foreground="black", background="yellow", font=("Arial", 10, "bold"))
style.map("Yellow.TButton", background=[("active", "#FFD700")])  # Amarillo dorado al pasar el mouse

# 🟢 Botón Verde Claro (Mostrar Tablas)
style.configure("LightGreen.TButton", foreground="black", background="#66CDAA", font=("Arial", 10, "bold"))
style.map("LightGreen.TButton", background=[("active", "#5CAD89")])  # Verde más oscuro al pasar el mouse




# Elementos de compras
ttk.Label(tab_compras, text="📌 Código del Producto:").grid(row=0, column=0, padx=10, pady=5, sticky="w")
entry_codigo_compra = ttk.Entry(tab_compras, width=30)
entry_codigo_compra.grid(row=0, column=1, padx=10, pady=5, sticky="w")

ttk.Label(tab_compras, text="🏷️ Nombre del Producto:").grid(row=0, column=2, padx=10, pady=5, sticky="w")
entry_nombre_producto_compra = ttk.Entry(tab_compras, width=100)
entry_nombre_producto_compra.grid(row=0, column=3, padx=10, pady=5, sticky="w")

ttk.Label(tab_compras, text="📦 Cantidad Comprada:").grid(row=1, column=0, padx=10, pady=5, sticky="w")
entry_cantidad_compra = ttk.Entry(tab_compras, width=30)
entry_cantidad_compra.grid(row=1, column=1, padx=10, pady=5, sticky="w")

ttk.Label(tab_compras, text="📅 Fecha compra (YYYY-MM-DD):").grid(row=2, column=0, padx=10, pady=5, sticky="w")
entry_fecha_compra = DateEntry(tab_compras, width=30, date_pattern='yyyy-mm-dd')
entry_fecha_compra.grid(row=2, column=1, padx=10, pady=5, sticky="w")

ttk.Label(tab_compras, text="🏢 Proveedor:").grid(row=3, column=0, padx=10, pady=5, sticky="w")
entry_proveedor = ttk.Entry(tab_compras, width=30)
entry_proveedor.grid(row=3, column=1, padx=10, pady=5, sticky="w")



# Función para buscar el nombre del producto en la pestaña de compras
def buscar_nombre_producto_compra(event=None):
    try:
        codigo_producto = entry_codigo_compra.get().strip()
        if not codigo_producto:
            entry_nombre_producto_compra.delete(0, tk.END)
            return

        cursor.execute("SELECT NombreProducto FROM Productos WHERE CodigoProducto = ?", (codigo_producto,))
        resultado = cursor.fetchone()

        if resultado:
            nombre_producto = resultado[0]
            entry_nombre_producto_compra.delete(0, tk.END)
            entry_nombre_producto_compra.insert(0, nombre_producto)
        else:
            entry_nombre_producto_compra.delete(0, tk.END)
            messagebox.showwarning("Advertencia", f"El producto con código '{codigo_producto}' no existe en la base de datos.")
    except Exception as e:
        messagebox.showerror("Error", f"Error al buscar el producto: {str(e)}")

entry_codigo_compra.bind("<FocusOut>", buscar_nombre_producto_compra)



# Botón para agregar compra
tk.Button(tab_compras, text="Agregar Compra", bg="lightgreen", command=agregar_compra).grid(row=4, column=0, columnspan=2, pady=10)


def eliminar_compra():
    """
    Elimina una compra y, de forma asociada, elimina las ordencompra correspondientes solo si es necesario.
    Si la compra eliminada deja ordenescompra sin cubrir, se solicita al usuario ingresar las IdOrdenCompra a eliminar.
    """
    try:
        id_compra_str = entry_id_compra.get().strip()
        if not id_compra_str:
            messagebox.showerror("Error", "Debe ingresar un IdCompra válido.")
            return

        id_compra_int = int(id_compra_str)

        # Consultar datos de la compra
        cursor.execute("""
            SELECT FechaCompra, CodigoProducto, NombreProducto, CantidadComprada, 
                   CostoSinIVA, PorcentajeIVA, IVA, CostoConIVA, Proveedor
            FROM Compras 
            WHERE IdCompra = ?
        """, (id_compra_int,))
        compra = cursor.fetchone()

        if not compra:
            messagebox.showerror("Error", "No se encontró ninguna compra con el IdCompra ingresado.")
            return

        (fecha_compra, codigo_producto, nombre_producto, cantidad_comprada, 
         costo_sin_iva, iva, valor_iva, costo_con_iva, proveedor) = compra

        cantidad_comprada = int(cantidad_comprada)

        # Consultar datos del producto
        cursor.execute("""
            SELECT CantidadInicial, CantidadInventario, CantidadVendida 
            FROM Productos 
            WHERE CodigoProducto = ?
        """, (codigo_producto,))
        producto = cursor.fetchone()

        if not producto:
            messagebox.showerror("Error", "El código de producto no existe.")
            return

        cantidad_inicial, cantidad_inventario, cantidad_vendida = map(int, producto)

        # Obtener todas las ordens compra asociadas a este producto
        cursor.execute("""
            SELECT IdOrdenCompra 
            FROM OrdenCompra
            WHERE CodigoProducto = ?
        """, (codigo_producto,))
        ordencompras = [str(v[0]) for v in cursor.fetchall()]

        # Mostrar datos de la compra para confirmar
        confirm_window = tk.Toplevel()
        confirm_window.title("Confirmar Eliminación")

        detalles = (
            f"📅 Fecha de Compra: {fecha_compra}\n"
            f"🔢 Código Producto: {codigo_producto}\n"
            f"🏷️ Nombre Producto: {nombre_producto}\n"
            f"📦 Cantidad Comprada: {cantidad_comprada}\n"
            f"💰 Costo Sin IVA: {costo_sin_iva}\n"
            f"💲 IVA: {iva}\n"
            f"💲 Valor IVA: {valor_iva}\n"
            f"💲 Costo Con IVA: {costo_con_iva}\n"
            f"🏭 Proveedor: {proveedor}\n\n"
            f"Esta compra restará {cantidad_comprada} unidades al inventario.\n"
        )

        tk.Label(confirm_window, text=detalles, justify=tk.LEFT).pack(pady=5, anchor="w")

        # Cálculo de cuántas unidades quedarían en inventario
        inventario_resultante = cantidad_inventario - cantidad_comprada

        # Si después de eliminar esta compra, el inventario es negativo y hay ordenes de compra, pedir IdOrdenCOmpra
        requiere_eliminar_ordencompra = inventario_resultante < 0 and cantidad_vendida > 0

        idordencompra_entries = []

        if requiere_eliminar_ordencompra:
            tk.Label(confirm_window, text="🔔 El inventario no es suficiente para cubrir todas las ordenes de compra.\n"
                                           "Debe indicar cuáles ventas eliminar (IdOrdenCompra):").pack(pady=5, anchor="w")

            # Solo pedir las IdOrdenCompras necesarias (inventario negativo implica ventas a cubrir)
            ordencompra_a_eliminar = abs(inventario_resultante)

            for i in range(ordencompra_a_eliminar):
                frame = tk.Frame(confirm_window)
                frame.pack(padx=5, pady=2, anchor="w")
                tk.Label(frame, text=f"IdOrdenCompra #{i+1}:").pack(side=tk.LEFT)
                entry = tk.Entry(frame, width=20)
                entry.pack(side=tk.LEFT)
                idordencompra_entries.append(entry)

            tk.Label(confirm_window, text=f"📋 Ordenes de Compras registradas ID: {', '.join(ordencompras)}").pack(pady=5, anchor="w")

        def realizar_eliminacion():
            idordencompra_a_eliminar = []

            if requiere_eliminar_ordencompra:
                for entry in idordencompra_entries:
                    id_ordencompra = entry.get().strip()
                    if not id_ordencompra:
                        messagebox.showerror("Error", "Debe completar todos los campos de IdOrdenCompra.")
                        return

                    if id_ordencompra not in ordencompras:
                        messagebox.showerror("Error", f"El IdOrdenCompra {id_ordencompra} no es válido.")
                        return

                    idordencompra_a_eliminar.append(id_ordencompra)

            # Eliminar las ventas seleccionadas
            for idordencompra in idordencompra_a_eliminar:
                cursor.execute("DELETE FROM OrdenCompra WHERE IdOrdenCompra = ?", (idordencompra,))

            # Eliminar la compra
            cursor.execute("DELETE FROM Compras WHERE IdCompra = ?", (id_compra_int,))

            # Ajustar las cantidades de productos
            nueva_cantidad_inicial = max(0, cantidad_inicial - cantidad_comprada)
            nueva_cantidad_inventario = max(0, cantidad_inventario - cantidad_comprada)
            nueva_cantidad_vendida = max(0, cantidad_vendida - len(idordencompra_a_eliminar))

            cursor.execute("""
                UPDATE Productos 
                SET CantidadInicial = ?, CantidadInventario = ?, CantidadVendida = ? 
                WHERE CodigoProducto = ?
            """, (nueva_cantidad_inicial, nueva_cantidad_inventario, nueva_cantidad_vendida, codigo_producto))

            conn.commit()
            messagebox.showinfo("Éxito", "Compra eliminada correctamente y stock actualizado.")
            confirm_window.destroy()

            # Limpiar entradas
            entry_id_compra.delete(0, tk.END)
            entry_codigo_compra.delete(0, tk.END)
            entry_nombre_producto_compra.delete(0, tk.END)
            entry_cantidad_compra.delete(0, tk.END)
            entry_proveedor.delete(0, tk.END)

        tk.Button(confirm_window, text="Confirmar Eliminación", command=realizar_eliminacion).pack(pady=10)

    except Exception as e:
        messagebox.showerror("Error", f"Error al eliminar la compra: {str(e)}")




def cargar_datos():
    """Carga los datos de una compra específica en los campos del formulario a partir de su IdCompra."""
    try:
        id_compra = entry_id_compra.get().strip()

        if not id_compra:
            messagebox.showerror("Error", "Debe ingresar un IdCompra válido.")
            return

        # Buscar la compra en la base de datos
        cursor.execute("""
            SELECT FechaCompra, CodigoProducto, NombreProducto, CantidadComprada, CostoSinIVA, PorcentajeIVA, IVA, CostoConIVA, Proveedor
            FROM Compras WHERE IdCompra = ?
        """, (id_compra,))
        compra = cursor.fetchone()

        if not compra:
            messagebox.showerror("Error", "No se encontró ninguna compra con el IdCompra ingresado.")
            return

        # Rellenar los campos con los datos de la compra encontrada
        (fecha_compra, codigo_producto, nombre_producto, cantidad_comprada, 
         costo_sin_iva, iva, valor_iva, costo_con_iva, proveedor) = compra
        
        entry_fecha_compra.set_date(fecha_compra)
        entry_codigo_compra.delete(0, tk.END)
        entry_codigo_compra.insert(0, codigo_producto)
        entry_nombre_producto_compra.delete(0, tk.END)
        entry_nombre_producto_compra.insert(0, nombre_producto)
        entry_cantidad_compra.delete(0, tk.END)
        cantidad_mostrar = int(cantidad_comprada) if cantidad_comprada == int(cantidad_comprada) else float(cantidad_comprada)
        entry_cantidad_compra.insert(0, str(cantidad_mostrar))
        entry_proveedor.delete(0, tk.END)
        entry_proveedor.insert(0, proveedor)

        messagebox.showinfo("Éxito", "Datos de la compra cargados correctamente.")

    except Exception as e:
        messagebox.showerror("Error", f"Error al cargar los datos: {str(e)}")

def modificar_compra():
    """Solo permite modificar FechaCompra y Proveedor de una compra existente."""
    try:
        id_compra = entry_id_compra.get().strip()
        if not id_compra:
            messagebox.showerror("Error", "Debe ingresar un IdCompra válido para modificar.")
            return

        # Obtener los nuevos valores del formulario
        nueva_fecha = entry_fecha_compra.get_date().strftime("%Y-%m-%d")
        nuevo_proveedor = entry_proveedor.get().strip()

        # Validar que los campos requeridos no estén vacíos
        if not nueva_fecha or not nuevo_proveedor:
            messagebox.showerror("Error", "Debe ingresar la Fecha de Compra y el Proveedor.")
            return

        # Obtener los valores originales desde la base de datos
        cursor.execute("""
            SELECT FechaCompra, CodigoProducto, NombreProducto, CantidadComprada, Proveedor
            FROM Compras WHERE IdCompra = ?
        """, (id_compra,))
        datos_originales = cursor.fetchone()

        if not datos_originales:
            messagebox.showerror("Error", "No se encontró la compra con ese IdCompra.")
            return

        fecha_original, codigo_original, nombre_original, cantidad_original, proveedor_original = datos_originales

        # Verificar si el usuario intentó modificar campos no permitidos
        codigo_actual = entry_codigo_compra.get().strip()
        nombre_actual = entry_nombre_producto_compra.get().strip()
        cantidad_actual = entry_cantidad_compra.get().strip()

        campos_modificados_no_permitidos = []

        if codigo_actual != codigo_original:
            campos_modificados_no_permitidos.append("Código del Producto")
        if nombre_actual != nombre_original:
            campos_modificados_no_permitidos.append("Nombre del Producto")
        if cantidad_actual != str(cantidad_original):
            campos_modificados_no_permitidos.append("Cantidad Comprada")

        if campos_modificados_no_permitidos:
            campos_str = ", ".join(campos_modificados_no_permitidos)
            messagebox.showwarning("Modificación no permitida", f"No se permite modificar los siguientes campos: {campos_str}.\n"
                                                                 f"Para modificar estos datos, debe eliminar la compra y volverla a ingresar.")
            return

        # Realizar solo la actualización de los campos permitidos
        cursor.execute("""
            UPDATE Compras 
            SET FechaCompra = ?, Proveedor = ?
            WHERE IdCompra = ?
        """, (nueva_fecha, nuevo_proveedor, id_compra))

        conn.commit()
        messagebox.showinfo("✅ Éxito", "Compra actualizada correctamente.")

    except Exception as e:
        messagebox.showerror("Error", f"Ocurrió un error al modificar la compra:\n{str(e)}")




# Campo IdCompra
ttk.Label(tab_compras, text="🆔 IdCompra:").grid(row=5, column=0, padx=5, pady=5, sticky="w")
entry_id_compra = ttk.Entry(tab_compras, width=20)  # Reducimos el ancho para mayor compactación
entry_id_compra.grid(row=5, column=1, padx=5, pady=5, sticky="w")

# Marco para agrupar los botones en la misma fila
frame_botones = ttk.Frame(tab_compras)
frame_botones.grid(row=5, column=2, columnspan=3, padx=5, pady=5, sticky="w")

# Botones en la misma fila dentro del frame
tk.Button(frame_botones, text="🔄 Cargar Datos", bg="lightgreen", command=cargar_datos).pack(side=tk.LEFT, padx=5)
tk.Button(frame_botones, text="✏️ Modificar Compra",bg="skyblue", command=modificar_compra).pack(side=tk.LEFT, padx=5)
tk.Button(frame_botones, text="🗑️ Eliminar Compra", bg="tomato",command=eliminar_compra).pack(side=tk.LEFT, padx=5)


# 🔄 Código de Producto Nuevo (para Cambio)
ttk.Label(tab_compras, text="🔄 Código Producto a Cambiar:").grid(row=6, column=0,  padx=5, pady=5, sticky="w")
entry_codigo_nuevo = ttk.Entry(tab_compras, width=30)
entry_codigo_nuevo.grid(row=6, column=1, padx=5, pady=5, sticky="w")
import difflib


def buscar_productos_similares():
    try:
        codigo_actual = entry_codigo_nuevo.get().strip().upper()

        if not codigo_actual:
            messagebox.showwarning("Atención", "Debe ingresar primero un Código de Producto a Cambiar.")
            return

        entry_codigo_nuevo.delete(0, tk.END)

        cursor.execute("SELECT NombreProducto FROM Productos WHERE CodigoProducto = ?", (codigo_actual,))
        resultado = cursor.fetchone()
        if not resultado:
            messagebox.showerror("Error", f"No se encontró un producto con el código '{codigo_actual}'.")
            return

        nombre_producto_actual = resultado[0]

        cursor.execute("SELECT CodigoProducto, NombreProducto, CantidadInventario FROM Productos")
        productos = cursor.fetchall()

        if not productos:
            messagebox.showinfo("Sin productos", "No se encontraron productos en la base de datos.")
            return

        ventana = tk.Toplevel(root)
        ventana.title(f"Seleccionar nuevo producto similar a: {nombre_producto_actual}")
        ventana.geometry("950x550")

        frame_busqueda = ttk.Frame(ventana)
        frame_busqueda.pack(fill=tk.X, padx=10, pady=(10, 0))

        ttk.Label(frame_busqueda, text="Buscar por código o nombre:").pack(side=tk.LEFT)
        entry_busqueda = ttk.Entry(frame_busqueda, width=50)
        entry_busqueda.pack(side=tk.LEFT, padx=10)

        frame_lista = ttk.Frame(ventana)
        frame_lista.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        scrollbar_y = ttk.Scrollbar(frame_lista, orient=tk.VERTICAL)
        scrollbar_x = ttk.Scrollbar(frame_lista, orient=tk.HORIZONTAL)
        listbox = tk.Listbox(
            frame_lista, width=120, height=20,
            yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set
        )
        scrollbar_y.config(command=listbox.yview)
        scrollbar_x.config(command=listbox.xview)

        scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)
        scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)
        listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Productos distintos al actual
        opciones_codigos = [
            (codigo, nombre, stock)
            for codigo, nombre, stock in productos if codigo != codigo_actual
        ]
        productos_visibles = []

        def actualizar_lista_busqueda(filtro=""):
            nonlocal productos_visibles
            listbox.delete(0, tk.END)
            productos_visibles = []

            for codigo, nombre, stock in opciones_codigos:
                if filtro.lower() in codigo.lower() or filtro.lower() in nombre.lower():
                    texto = f"🔢 Código: {codigo} | 📦 Nombre: {nombre} | 🏷️ Stock: {stock}"
                    listbox.insert(tk.END, texto)
                    productos_visibles.append((codigo, nombre, stock))  # Actualizar visibilidad

        def on_buscar(event):
            texto = entry_busqueda.get().strip()
            actualizar_lista_busqueda(texto)

        entry_busqueda.bind("<KeyRelease>", on_buscar)

        # Mostrar todos inicialmente
        actualizar_lista_busqueda()

        def seleccionar_producto():
            seleccion = listbox.curselection()
            if not seleccion:
                messagebox.showwarning("Advertencia", "Debe seleccionar un producto de la lista.")
                return

            indice = seleccion[0]
            codigo_seleccionado = productos_visibles[indice][0]  # Usamos la lista actual filtrada

            confirmar_cambio_producto(codigo_actual, codigo_seleccionado)
            ventana.destroy()

        ttk.Button(ventana, text="✅ Confirmar Cambio", command=seleccionar_producto).pack(pady=10)

    except Exception as e:
        messagebox.showerror("Error", f"Ocurrió un error: {str(e)}")



def confirmar_cambio_producto(codigo_actual, codigo_nuevo):
    try:
        if codigo_actual == codigo_nuevo:
            messagebox.showerror("Error", "No puede cambiar al mismo producto. Seleccione uno diferente.")
            return

        # Verificar existencia de productos
        cursor.execute("SELECT CantidadInicial, CantidadInventario FROM Productos WHERE CodigoProducto = ?", (codigo_actual,))
        prod_actual = cursor.fetchone()

        cursor.execute("SELECT CantidadInicial, CantidadInventario FROM Productos WHERE CodigoProducto = ?", (codigo_nuevo,))
        prod_nuevo = cursor.fetchone()

        if not prod_actual or not prod_nuevo:
            messagebox.showerror("Error", "Uno de los productos no existe.")
            return

        cantidad_inicial_actual, cantidad_inventario_actual = prod_actual
        cantidad_inicial_nuevo, cantidad_inventario_nuevo = prod_nuevo

        if cantidad_inventario_actual <= 0:
            messagebox.showerror("Error", "No hay unidades disponibles del producto actual para cambiar.")
            return

        # Compra más reciente del producto actual
        cursor.execute("""
            SELECT IdCompra, CantidadComprada 
            FROM Compras 
            WHERE CodigoProducto = ?
            ORDER BY IdCompra DESC
        """, (codigo_actual,))
        compra_reciente = cursor.fetchone()

        if not compra_reciente:
            messagebox.showerror("Error", "No se encontró una compra registrada para este producto.")
            return

        id_compra_anterior, cantidad_comprada = compra_reciente

        # Ventana emergente para registrar nueva compra
        ventana = tk.Toplevel(root)
        ventana.title(f"Registrar cambio de producto: {codigo_nuevo}")
        ventana.geometry("400x500")
        ventana.lift()
        ventana.focus_force()

        ttk.Label(ventana, text="📅 Fecha de Compra (YYYY-MM-DD):").pack(pady=10)
        entry_fecha = DateEntry(ventana, width=20, date_pattern='yyyy-MM-dd')
        entry_fecha.pack()

        ttk.Label(ventana, text="📅 Fecha del Cambio (YYYY-MM-DD):").pack(pady=10)
        entry_fecha_cambio = DateEntry(ventana, width=20, date_pattern='yyyy-MM-dd')
        entry_fecha_cambio.set_date(date.today())
        entry_fecha_cambio.pack()

        ttk.Label(ventana, text="🏢 Proveedor:").pack(pady=10)
        entry_proveedor = ttk.Entry(ventana, width=30)
        entry_proveedor.pack()

        def registrar_compra():
            try:
                fecha_compra = entry_fecha.get_date().strftime("%Y-%m-%d")
                fecha_cambio = entry_fecha_cambio.get_date().strftime("%Y-%m-%d")
                proveedor = entry_proveedor.get().strip()

                if not fecha_compra or not proveedor or not fecha_cambio:
                    messagebox.showerror("Error", "Debe ingresar todos los campos requeridos.")
                    return

                # Obtener datos del nuevo producto
                cursor.execute("""
                    SELECT NombreProducto, CostoConIVA, IVA 
                    FROM Productos
                    WHERE CodigoProducto = ?
                """, (codigo_nuevo,))
                datos_producto = cursor.fetchone()

                if not datos_producto:
                    messagebox.showerror("Error", "No se encontró el producto para registrar la compra.")
                    return

                nombre_producto, costo_con_iva, iva = datos_producto

                costo_con_iva_total = costo_con_iva
                valor_iva = costo_con_iva_total - (costo_con_iva_total / (1 + iva))
                costo_sin_iva = costo_con_iva_total - valor_iva

                # Insertar nueva compra y obtener ID insertado
                cursor.execute("""
                    INSERT INTO Compras (
                        FechaCompra, CodigoProducto, NombreProducto, CantidadComprada,
                        CostoSinIVA, PorcentajeIVA, IVA, CostoConIVA, Proveedor
                    )
                    OUTPUT INSERTED.IdCompra
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    fecha_compra, codigo_nuevo, nombre_producto, 1,
                    costo_sin_iva, iva, valor_iva, costo_con_iva_total, proveedor
                ))
                id_compra_nueva = cursor.fetchone()[0]

                # Actualizar compra anterior
                if cantidad_comprada > 1:
                    nueva_cantidad = cantidad_comprada - 1
                    cursor.execute("""
                        UPDATE Compras
                        SET CantidadComprada = ?
                        WHERE IdCompra = ?
                    """, (nueva_cantidad, id_compra_anterior))
                else:
                    cursor.execute("DELETE FROM Compras WHERE IdCompra = ?", (id_compra_anterior,))

                # Actualizar inventario del producto anterior
                nueva_inicial_actual = max(0, cantidad_inicial_actual - 1)
                nueva_inventario_actual = max(0, cantidad_inventario_actual - 1)
                cursor.execute("""
                    UPDATE Productos
                    SET CantidadInicial = ?, CantidadInventario = ?
                    WHERE CodigoProducto = ?
                """, (nueva_inicial_actual, nueva_inventario_actual, codigo_actual))

                # Actualizar inventario del nuevo producto
                nueva_inicial_nuevo = cantidad_inicial_nuevo + 1
                nueva_inventario_nuevo = cantidad_inventario_nuevo + 1
                cursor.execute("""
                    UPDATE Productos
                    SET CantidadInicial = ?, CantidadInventario = ?
                    WHERE CodigoProducto = ?
                """, (nueva_inicial_nuevo, nueva_inventario_nuevo, codigo_nuevo))

                # Insertar en tabla CambiarProducto incluyendo IdCompra nueva
                cursor.execute("""
                    INSERT INTO CambiarProducto (CodigoProductoAnterior, CodigoProductoNuevo, FechaCambio, IdCompra)
                    VALUES (?, ?, ?, ?)
                """, (codigo_actual, codigo_nuevo, fecha_cambio, id_compra_nueva))

                conn.commit()

                messagebox.showinfo(
                    "✅ Cambio realizado",
                    f"Se registró correctamente el cambio:\n\n{codigo_actual} ➡️ {codigo_nuevo}\n\n"
                    f"Fecha de cambio: {fecha_cambio}\nID Compra nueva: {id_compra_nueva}"
                )
                ventana.destroy()

            except Exception as e:
                messagebox.showerror("Error", f"Ocurrió un error al registrar el cambio: {str(e)}")

        tk.Button(ventana, text="✅ Confirmar Registro", bg="lightgreen", command=registrar_compra).pack(pady=20)

    except Exception as e:
        messagebox.showerror("Error", f"Ocurrió un error al confirmar el cambio: {str(e)}")








def abrir_ventana_registrar_compra(codigo_nuevo):
    """Ventana emergente para registrar la compra de 1 unidad al cambiar producto."""
    try:
        ventana_registro = tk.Toplevel(root)
        ventana_registro.title(f"Registrar Compra - Producto: {codigo_nuevo}")
        ventana_registro.geometry("400x250")
        ventana_registro.resizable(False, False)

        # Fecha de Compra
        ttk.Label(ventana_registro, text="📅 Fecha de Compra:").pack(pady=(20, 5))
        entry_fecha = DateEntry(ventana_registro, width=20, background='darkblue', foreground='white', date_pattern='yyyy-mm-dd')
        entry_fecha.pack()

        # Proveedor
        ttk.Label(ventana_registro, text="🏢 Proveedor:").pack(pady=(20, 5))
        entry_proveedor = ttk.Entry(ventana_registro, width=30)
        entry_proveedor.pack()

        def confirmar_registro():
            fecha_compra = entry_fecha.get_date().strftime("%Y-%m-%d")
            proveedor = entry_proveedor.get().strip()

            if not proveedor:
                messagebox.showwarning("Atención", "Debe ingresar el nombre del proveedor.")
                return

            # Obtener datos del nuevo producto
            cursor.execute("SELECT NombreProducto, CostoConIVA, IVA FROM Productos WHERE CodigoProducto = ?", (codigo_nuevo,))
            producto = cursor.fetchone()

            if not producto:
                messagebox.showerror("Error", "No se encontró el producto seleccionado.")
                return

            nombre_producto, costo_con_iva_unitario, porcentaje_iva = producto

            # Calcular valores
            costo_con_iva_total = costo_con_iva_unitario * 1
            valor_iva = costo_con_iva_total - (costo_con_iva_total / (1 + porcentaje_iva))
            costo_sin_iva = costo_con_iva_total - valor_iva

            # Insertar en Compras
            cursor.execute("""
                INSERT INTO Compras 
                (FechaCompra, CodigoProducto, NombreProducto, CantidadComprada, CostoSinIVA, PorcentajeIVA, IVA, CostoConIVA, Proveedor)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                fecha_compra, codigo_nuevo, nombre_producto, 1, 
                costo_sin_iva, porcentaje_iva, valor_iva, costo_con_iva_total, proveedor
            ))

            conn.commit()

            messagebox.showinfo("✅ Éxito", f"Compra registrada correctamente para el producto {codigo_nuevo}.")

            ventana_registro.destroy()

        # Botón Confirmar
        tk.Button(ventana_registro, text="💾 Confirmar Registro", bg="lightgreen", command=confirmar_registro).pack(pady=20)

    except Exception as e:
        messagebox.showerror("Error", f"Ocurrió un error: {str(e)}")



tk.Button(tab_compras, text="🔄 Cambiar 1 Producto", bg="#00BFFF", fg="white", command=buscar_productos_similares).grid(row=6, column=2, pady=5,)



def exportar_excel3(datos, columnas):
    import pandas as pd
    from tkinter import filedialog
    from openpyxl.utils import get_column_letter
    from openpyxl import load_workbook
    from openpyxl.styles import numbers

    # Convertir strings numéricos/fechas a tipos reales
    def convertir_valor(valor):
        if isinstance(valor, str):
            valor = valor.strip()
            # Probar si es fecha en formato YYYY-MM-DD
            try:
                if "-" in valor and len(valor) == 10:
                    return pd.to_datetime(valor).date()
            except:
                pass
            # Probar si es numérico
            try:
                if "." in valor:
                    return float(valor)
                else:
                    return int(valor)
            except ValueError:
                return valor
        return valor

    # Convertir datos
    datos_convertidos = [[convertir_valor(celda) for celda in fila] for fila in datos]

    # Crear DataFrame
    df = pd.DataFrame(datos_convertidos, columns=columnas)

    # Guardar archivo
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        df.to_excel(file_path, index=False)

        # Formato profesional
        wb = load_workbook(file_path)
        ws = wb.active

        for col_idx, col_cells in enumerate(ws.columns, start=1):
            max_length = 0
            for cell in col_cells:
                if cell.value is not None:
                    # Ajustar ancho de columna
                    max_length = max(max_length, len(str(cell.value)))

                    # Aplicar formato si es numérico o fecha
                    if isinstance(cell.value, int):
                        cell.number_format = numbers.FORMAT_NUMBER
                    elif isinstance(cell.value, float):
                        cell.number_format = numbers.FORMAT_NUMBER_00  # dos decimales
                    elif isinstance(cell.value, (pd.Timestamp, pd._libs.tslibs.timestamps.Timestamp, type(pd.to_datetime("2020-01-01").date()))):
                        cell.number_format = numbers.FORMAT_DATE_YYYYMMDD2

            # Autoajustar ancho
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max_length + 2

        wb.save(file_path)
from datetime import date
from tkcalendar import DateEntry

def mostrar_cambioproducto():
    try:
        ventana_productos = tk.Toplevel(root)
        ventana_productos.title("Historial de Cambios de Producto")
        ventana_productos.geometry("1920x1080")

        # === FILTROS ===
        frame_filtros = ttk.Frame(ventana_productos)
        frame_filtros.pack(pady=10)

        ttk.Label(frame_filtros, text="🔁 Código Producto Anterior:").grid(row=0, column=0, padx=5, sticky="e")
        entry_cod_anterior = ttk.Entry(frame_filtros, width=20)
        entry_cod_anterior.grid(row=0, column=1, padx=5)

        ttk.Label(frame_filtros, text="✅ Código Producto Nuevo:").grid(row=0, column=2, padx=5, sticky="e")
        entry_cod_nuevo = ttk.Entry(frame_filtros, width=20)
        entry_cod_nuevo.grid(row=0, column=3, padx=5)

        ttk.Label(frame_filtros, text="📅 Fecha Inicio:").grid(row=0, column=4, padx=5, sticky="e")
        entry_fecha_inicio = DateEntry(frame_filtros, width=15, date_pattern="yyyy-mm-dd")
        entry_fecha_inicio.set_date(date(2025, 1, 1))  # ✅ FECHA INICIO DEFAULT
        entry_fecha_inicio.grid(row=0, column=5, padx=5)

        ttk.Label(frame_filtros, text="📅 Fecha Fin:").grid(row=0, column=6, padx=5, sticky="e")
        entry_fecha_fin = DateEntry(frame_filtros, width=15, date_pattern="yyyy-mm-dd")
        entry_fecha_fin.set_date(date.today())
        entry_fecha_fin.grid(row=0, column=7, padx=5)

        # === TREEVIEW ===
        frame_tree = ttk.Frame(ventana_productos)
        frame_tree.pack(fill=tk.BOTH, expand=True)

        scrollbar_x = ttk.Scrollbar(frame_tree, orient=tk.HORIZONTAL)
        scrollbar_y = ttk.Scrollbar(frame_tree, orient=tk.VERTICAL)

        columnas_tree = (
            "IdCambioProducto", "CodigoProductoAnterior", "NombreAnterior",
            "CodigoProductoNuevo", "NombreNuevo", "FechaCambio", "IdCompra"
        )

        tree = ttk.Treeview(
            frame_tree,
            columns=columnas_tree,
            show="headings",
            xscrollcommand=scrollbar_x.set,
            yscrollcommand=scrollbar_y.set
        )
        scrollbar_x.config(command=tree.xview)
        scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)
        scrollbar_y.config(command=tree.yview)
        scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)

        style = ttk.Style()
        style.configure("Treeview.Heading", font=("Arial", 13, "bold"))
        style.configure("Treeview", font=("Arial", 10))

        encabezados = [
            ("IdCambioProducto", "ID Cambio", 300),
            ("CodigoProductoAnterior", "Código Anterior", 300),
            ("NombreAnterior", "Nombre Producto Anterior", 600),
            ("CodigoProductoNuevo", "Código Nuevo", 300),
            ("NombreNuevo", "Nombre Producto Nuevo", 600),
            ("FechaCambio", "Fecha del Cambio", 300),
            ("IdCompra", "ID Compra", 200)
        ]

        for col, text, width in encabezados:
            tree.heading(col, text=text)
            tree.column(col, width=width, anchor="center")

        tree.pack(fill=tk.BOTH, expand=True)

        datos_exportacion = []

        def ejecutar_filtro():
            try:
                tree.delete(*tree.get_children())
                datos_exportacion.clear()

                cod_anterior = entry_cod_anterior.get().strip()
                cod_nuevo = entry_cod_nuevo.get().strip()
                fecha_inicio = entry_fecha_inicio.get_date()
                fecha_fin = entry_fecha_fin.get_date()

                query = """
                    SELECT 
                        cp.IdCambioProducto,
                        cp.CodigoProductoAnterior,
                        p1.NombreProducto AS NombreAnterior,
                        cp.CodigoProductoNuevo,
                        p2.NombreProducto AS NombreNuevo,
                        cp.FechaCambio,
                        cp.IdCompra
                    FROM CambiarProducto cp
                    LEFT JOIN Productos p1 ON cp.CodigoProductoAnterior = p1.CodigoProducto
                    LEFT JOIN Productos p2 ON cp.CodigoProductoNuevo = p2.CodigoProducto
                    WHERE cp.FechaCambio BETWEEN ? AND ?
                """

                params = [fecha_inicio, fecha_fin]

                if cod_anterior:
                    query += " AND cp.CodigoProductoAnterior = ?"
                    params.append(cod_anterior)

                if cod_nuevo:
                    query += " AND cp.CodigoProductoNuevo = ?"
                    params.append(cod_nuevo)

                query += " ORDER BY cp.FechaCambio DESC"

                cursor.execute(query, params)
                resultados = cursor.fetchall()

                for fila in resultados:
                    fila_str = [str(x) if x is not None else "" for x in fila]
                    tree.insert("", tk.END, values=fila_str)
                    datos_exportacion.append(fila_str)

                if not resultados:
                    messagebox.showinfo("Sin resultados", "No se encontraron cambios de producto con esos filtros.")

            except Exception as e:
                messagebox.showerror("Error", f"No se pudo aplicar el filtro: {e}")

        # Botón Buscar
        btn_buscar = ttk.Button(frame_filtros, text="🔍 Buscar", command=ejecutar_filtro)
        btn_buscar.grid(row=0, column=8, padx=10)

        # Ejecutar consulta inicial
        ejecutar_filtro()

        columnas = [col for col, _, _ in encabezados]

        # === EXPORTACIÓN ===
        frame_botones_exportar = ttk.Frame(ventana_productos)
        frame_botones_exportar.pack(pady=10)

        btn_excel = ttk.Button(
            frame_botones_exportar,
            text="📥 Exportar a Excel",
            command=lambda: exportar_excel(datos_exportacion, columnas)
        )
        btn_excel.pack(side="left", padx=10)

    except Exception as e:
        messagebox.showerror("Error", f"No se pudo cargar el historial de cambios: {e}")





btn_mostrar_cambioproducto = tk.Button(
    tab_compras, text="🔄 Mostrar Cambio Producto ", command=mostrar_cambioproducto,bg="#5CD6C0",fg="black",font=("Arial", 10, "bold"),relief="raised",padx=10,pady=3,activebackground="#47B7A7",cursor="hand2"
)
btn_mostrar_cambioproducto.grid(row=7, column=0, pady=10)


# ---------------------- PESTAÑA OrdenCompra ----------------------



productos_agregados = []
# Crear un frame para organizar la distribución en la pestaña de ordencompra
frame_ordencompra_contenedor = ttk.Frame(tab_orden_compra)
frame_ordencompra_contenedor.grid(row=0, column=0, sticky="nsew", padx=10, pady=(5, 2))


# Dividir la pantalla en dos partes: IZQUIERDA (Formulario) y DERECHA (Productos Agregados)
frame_izquierdo = ttk.Frame(frame_ordencompra_contenedor)
frame_izquierdo.grid(row=0, column=0, sticky="nw", padx=10, pady=(0, 5))

frame_derecho = ttk.Frame(frame_ordencompra_contenedor)
frame_derecho.grid(row=0, column=1, sticky="ne", padx=10, pady=(0, 5))
# Labels y Entradas

campos_ordencompra = [
    "📜 Orden de Compra*", "📅 Fecha Orden De Compra (YYYY-MM-DD)*","👤 Nombre del Cliente*", "☎️ Telefono", "🌆 Ciudad" , "👤 20% Consultor",  "👤 5% Padre Empresarial","💲 IVA (%)*",
    "🏷️ Código del Producto*", "📦 Nombre del Producto", "📊 Cantidad Vendida*",
]

entries_ordencompra = {}

# Crear entradas y labels con mayor ancho
for i, campo in enumerate(campos_ordencompra):
    ttk.Label(frame_izquierdo, text=f"{campo}:").grid(column=0, row=i, padx=5, pady=(1, 2), sticky="w")

    if "Fecha" in campo:
        entry = DateEntry(frame_izquierdo, width=20, background='darkblue', foreground='white', borderwidth=2,
                        year=datetime.datetime.now().year, month=datetime.datetime.now().month,
                        day=datetime.datetime.now().day, date_pattern='yyyy-MM-dd', state="readonly")

    else:
        entry = ttk.Entry(frame_izquierdo, width=70)  

    entry.grid(column=1, row=i, padx=5, pady=3, sticky="w")
    entries_ordencompra[campo] = entry

    # Crear un frame para agrupar los tres campos de fechas de pago en una única fila
frame_fechas = ttk.Frame(frame_izquierdo)
# Se posiciona en la fila siguiente a los 18 campos
frame_fechas.grid(row=len(campos_ordencompra), column=0, columnspan=2, sticky="w", padx=5, pady=3)



# Área de texto para mostrar productos agregados
ttk.Label(frame_derecho, text="Productos Agregados a la Orden:", font=("Arial", 12, "bold")).pack(anchor="w", pady=5)

text_area = tk.Text(frame_derecho, width=70, height=24, wrap="word")
text_area.pack(side="left", fill="both", expand=True, padx=5, pady=5)

scrollbar = ttk.Scrollbar(frame_derecho, command=text_area.yview)
scrollbar.pack(side="right", fill="y")
text_area.config(yscrollcommand=scrollbar.set)

# Función para completar el nombre del producto automáticamente
def completar_nombre_producto(event):
    codigo = entries_ordencompra["🏷️ Código del Producto*"].get().strip()
    if codigo:
        cursor.execute("SELECT NombreProducto FROM Productos WHERE CodigoProducto = ?", (codigo,))
        producto = cursor.fetchone()
        if producto:
            entries_ordencompra["📦 Nombre del Producto"].delete(0, tk.END)
            entries_ordencompra["📦 Nombre del Producto"].insert(0, producto[0])
        else:
            entries_ordencompra["📦 Nombre del Producto"].delete(0, tk.END)
            messagebox.showerror("Error", "El código de producto ingresado no existe.")


entries_ordencompra["🏷️ Código del Producto*"].bind("<FocusOut>", completar_nombre_producto)
def actualizar_text_area():
    text_area.delete("1.0", tk.END)
    for i, producto in enumerate(productos_agregados, start=1):
        text_area.insert(tk.END, f"\n===== 🛒 Producto #{i} =====\n"
                                 f"📜 Orden de Compra: {producto['NumOrdenCompra']}\n"
                                 f"📅 Fecha Orden Compra: {producto['FechaOrdenCompra']}\n"
                                 f"👤 Nombre del Cliente: {producto['NombreCliente']}\n"
                                 f"☎️ Teléfono: {producto['Telefono']}\n"
                                 f"🌆 Ciudad: {producto['Ciudad']}\n"
                                 f"👤 20% Consultor: {producto['NombreConsultor']}\n"
                                 f"👤 5% Padre Empresarial: {producto['NombrePadreEmpresarial']}\n"
                                 f"💲 IVA (%): {producto['PorcentajeIVA'] * 100:.2f}\n"
                                 f"🏷️ Código del Producto: {producto['CodigoProducto']}\n"
                                 f"📦 Nombre del Producto: {producto['NombreProducto']}\n"
                                 f"📊 Cantidad Vendida: {producto['CantidadVendida']}\n"
                                 f"Valor a Cobrar con IVA: {producto['ValorXCobrarConIVA']}\n"
                                 "------------------------------------------\n")
    text_area.yview_moveto(1)

def deshacer_ultimo_producto():
    if not productos_agregados:
        messagebox.showinfo("Nada que eliminar", "No hay productos agregados aún.")
        return
    productos_agregados.pop()  # Quita el último
    actualizar_text_area()
    messagebox.showinfo("Deshecho", "Se eliminó el último producto agregado.")


orden_contador = 1  # Contador de órdenes
productos_agregados = []  # Lista para almacenar los productos agregados a la orden

def agregar_producto_a_orden():
    orden_compra = entries_ordencompra["📜 Orden de Compra*"].get().strip()

    # Verificar si ya existe una orden con ese número
    cursor.execute("SELECT COUNT(*) FROM OrdenCompra WHERE NumOrdenCompra = ?", (orden_compra,))
    existe = cursor.fetchone()[0]
    if existe > 0:
        messagebox.showerror("Error", f"Ya existe una orden de compra con el número {orden_compra}. No se puede duplicar.")
        return

    """Agrega un producto a la orden con cálculos completos y lo muestra en el área de texto."""
    global orden_contador
    
    try:
        # Validar campos obligatorios
        if not entries_ordencompra["📜 Orden de Compra*"].get().strip():
            messagebox.showerror("Error", "Falta llenar Orden de Compra.")
            return
        if not entries_ordencompra["📅 Fecha Orden De Compra (YYYY-MM-DD)*"].get().strip():
            messagebox.showerror("Error", "Falta llenar Fecha OrdenCompra.")
            return
        if not entries_ordencompra["👤 Nombre del Cliente*"].get().strip():
            messagebox.showerror("Error", "Falta llenar Nombre del Cliente.")
            return
        if not entries_ordencompra["💲 IVA (%)*"].get().strip():
            messagebox.showerror("Error", "Falta llenar IVA.")
            return
        if not entries_ordencompra["🏷️ Código del Producto*"].get().strip():
            messagebox.showerror("Error", "Falta llenar Código del Producto.")
            return
        if not entries_ordencompra["📊 Cantidad Vendida*"].get().strip():
            messagebox.showerror("Error", "Falta llenar Cantidad Vendida.")
            return

        # Extraer datos del formulario
        orden_compra = entries_ordencompra["📜 Orden de Compra*"].get().strip()
        fecha_orden = entries_ordencompra["📅 Fecha Orden De Compra (YYYY-MM-DD)*"].get_date().strftime('%Y-%m-%d')
        nombre_cliente = entries_ordencompra["👤 Nombre del Cliente*"].get().strip()
        telefono = entries_ordencompra["☎️ Telefono"].get().strip()
        ciudad = entries_ordencompra["🌆 Ciudad"].get().strip()
        consultor = entries_ordencompra["👤 20% Consultor"].get().strip()
        padre_empresarial = entries_ordencompra["👤 5% Padre Empresarial"].get().strip()
        iva = entries_ordencompra["💲 IVA (%)*"].get().strip()
        codigo_producto = entries_ordencompra["🏷️ Código del Producto*"].get().strip()
        nombre_producto = entries_ordencompra["📦 Nombre del Producto"].get().strip()
        cantidad_vendida = entries_ordencompra["📊 Cantidad Vendida*"].get().strip()

        # Verificar si el producto ya existe en la orden actual
        for producto in productos_agregados:
            if producto["CodigoProducto"] == codigo_producto:
                messagebox.showerror("Error", f"El producto con código {codigo_producto} ya existe en la orden con una cantidad de {producto['CantidadVendida']} unidades.")
                return

        # === VALIDACIONES LÓGICAS DE LOS CAMPOS ===

        if not orden_compra.isdigit():
            messagebox.showerror("Error", "El número de Orden de Compra debe ser un número entero.")
            return

        if not re.match(r"^[a-zA-ZÁÉÍÓÚáéíóúñÑ\s]+$", nombre_cliente):
            messagebox.showerror("Error", "El Nombre del Cliente solo debe contener letras.")
            return


        if ciudad and not re.match(r"^[a-zA-ZÁÉÍÓÚáéíóúñÑ\s]+$", ciudad):
            messagebox.showerror("Error", "La Ciudad solo debe contener letras.")
            return

        if consultor and not re.match(r"^[a-zA-ZÁÉÍÓÚáéíóúñÑ\s]+$", consultor):
            messagebox.showerror("Error", "El nombre del Consultor debe contener solo letras.")
            return

        if padre_empresarial and not re.match(r"^[a-zA-ZÁÉÍÓÚáéíóúñÑ\s]+$", padre_empresarial):
            messagebox.showerror("Error", "El nombre del Padre Empresarial debe contener solo letras.")
            return

        try:
            iva_decimal = Decimal(iva)
            if iva_decimal < 0 or iva_decimal > 100:
                messagebox.showerror("Error", "El IVA debe estar entre 0 y 100.")
                return
        except:
            messagebox.showerror("Error", "Ingrese un valor válido para el IVA.")
            return

                
        try:
            cantidad_vendida = Decimal(cantidad_vendida)
            if cantidad_vendida <= 0:
                messagebox.showerror("Error", "La cantidad vendida debe ser mayor que cero.")
                return
        except InvalidOperation:
            messagebox.showerror("Error", "Ingrese un número válido en Cantidad Vendida.")
            return

        
        # Obtener detalles del producto desde la base de datos
        cursor.execute("SELECT NombreProducto, PrecioVentaConIVA, PVPSinIVA, CostoConIVA, CantidadInventario FROM Productos WHERE CodigoProducto = ?", (codigo_producto,))
        producto = cursor.fetchone()

        if not producto:
            messagebox.showerror("Error", "El código de producto ingresado no existe.")
            return

        nombre_producto, precio_venta_con_iva, pvp_sin_iva, costo_con_iva, cantidad_inventario = producto
        cantidad_inventario = int(cantidad_inventario)  # 🔧 Conversión segura

        precio_venta_con_iva = Decimal(precio_venta_con_iva)
        pvp_sin_iva = Decimal(pvp_sin_iva)
        costo_con_iva = Decimal(costo_con_iva)

        # Convertir todas las variables necesarias a Decimal antes de operar
        cantidad_vendida = Decimal(cantidad_vendida)
        iva = Decimal(iva) / 100  # Para usar 15% como 0.15 
        porcentaje_descuento = Decimal("0")  # Se inicializa en 0, y se recalcula después





        # Validación de stock disponible
        if cantidad_vendida <= 0 or cantidad_vendida > cantidad_inventario:
            messagebox.showerror("Error", "Cantidad no disponible en inventario.")
            return
        

        # Cálculos de valores asegurando Decimales
        precio_venta_con_iva = precio_venta_con_iva
        pvp_sin_iva = pvp_sin_iva
        valor_descuento = pvp_sin_iva * porcentaje_descuento
        Base_Retencion = (pvp_sin_iva - valor_descuento) * cantidad_vendida
        Valor_Base_Retencion = Base_Retencion * iva
        valor_x_cobrar_con_iva = (Base_Retencion + (Base_Retencion * iva)) 
        costo_con_iva= costo_con_iva * cantidad_vendida

        # Comisión Consultor
        if consultor == "":
            porcentaje_comision_consultor = Decimal("0") 
            comision_por_pagar_consultor = Decimal("0")
        else:
            porcentaje_comision_consultor = Decimal("0.20")
            comision_por_pagar_consultor = round(Base_Retencion * porcentaje_comision_consultor, 2)

        # Comisión Padre Empresarial
        if padre_empresarial == "":
            porcentaje_padre_empresarial = Decimal("0")
            comision_por_pagar_padre_empresarial = Decimal("0")
        else:
            porcentaje_padre_empresarial = Decimal("0.05")
            comision_por_pagar_padre_empresarial = round(Base_Retencion * porcentaje_padre_empresarial, 2)


        valor_cliente= precio_venta_con_iva * cantidad_vendida



        # Agregar producto a la lista
        productos_agregados.append({
        "NumOrdenCompra": orden_compra,
        "FechaOrdenCompra": fecha_orden,
        "NombreCliente": nombre_cliente,
        "Telefono": telefono,
        "Ciudad": ciudad,
        "NombreConsultor": consultor,
        "PorcentajeComisionConsultor": porcentaje_comision_consultor,
        "ComisionPorPagarConsultor": comision_por_pagar_consultor,
        "NombrePadreEmpresarial": padre_empresarial,
        "PorcentajePadreEmpresarial": porcentaje_padre_empresarial,
        "ComisionPorPagarPadreEmpresarial": comision_por_pagar_padre_empresarial,
        "PorcentajeIVA": iva,
        "CodigoProducto": codigo_producto,
        "NombreProducto": nombre_producto,
        "CantidadVendida": cantidad_vendida,
        "PorcentajeDescuento": porcentaje_descuento ,
        "PrecioVentaConIVA": precio_venta_con_iva,
        "PVPSinIVA": pvp_sin_iva,
        "ValorDescuento": valor_descuento,
        "BaseRetencion": Base_Retencion,
        "ValorBaseRetencion": Valor_Base_Retencion,
        "ValorCliente": valor_cliente,
        "ValorXCobrarConIVA": valor_x_cobrar_con_iva,
    
        })
        actualizar_text_area()

        messagebox.showinfo("Éxito", "Producto agregado a la orden exitosamente.")
        orden_contador += 1

        # Campos que NO se deben limpiar al agregar producto
        campos_no_limpiar = {
            "📜 Orden de Compra*",
            "📅 Fecha Orden De Compra (YYYY-MM-DD)*",
            "👤 Nombre del Cliente*",
            "☎️ Telefono",
            "🌆 Ciudad",
            "👤 20% Consultor",
            "👤 5% Padre Empresarial",
            "💲 IVA (%)*"
        }

        for key, entry in entries_ordencompra.items():
            if key in campos_no_limpiar:
                continue  # No limpiar estos campos
            if isinstance(entry, tk.Entry):
                entry.delete(0, tk.END)
            elif isinstance(entry, ttk.Combobox):
                entry.set("")
            elif isinstance(entry, ttk.DateEntry):
                entry.set_date("")

    except (ValueError, InvalidOperation) as e:
        messagebox.showerror("Error", f"Ocurrió un error: {e}")

def confirmar_ordencompra():
    if not productos_agregados:
        messagebox.showerror("Error", "No hay productos agregados a la orden.")
        return

    total_por_cobrar_con_iva = sum(producto["ValorXCobrarConIVA"] for producto in productos_agregados)
    descuento_actual = productos_agregados[0]["PorcentajeDescuento"] / 100 if productos_agregados else Decimal(0)

    confirm_window = tk.Toplevel()
    confirm_window.title("Confirmar Venta")

    tk.Label(confirm_window, text="Confirmar Venta", font=("Arial", 12, "bold")).pack(pady=5)
    lbl_total = tk.Label(confirm_window, text=f"Total por Cobrar con IVA: ${round(total_por_cobrar_con_iva, 2):,.2f}")
    lbl_total.pack(pady=5)

    tk.Label(confirm_window, text="Valor por Cobrar Cliente:").pack()
    entry_valor_por_cobrar_cliente = tk.Entry(confirm_window)
    entry_valor_por_cobrar_cliente.pack(pady=5)

    def on_editar_valor(event):
        btn_aceptar.config(state="disabled")

    entry_valor_por_cobrar_cliente.bind("<KeyRelease>", on_editar_valor)
    entry_valor_por_cobrar_cliente.insert(0, str(round(total_por_cobrar_con_iva, 2)))

    lbl_descuento = tk.Label(confirm_window, text=f"% Descuento: {descuento_actual * 100:.2f}%")
    lbl_descuento.pack(pady=5)

    def calcular_descuento():
        try:
            valor_cliente_str = entry_valor_por_cobrar_cliente.get().strip()
            valor_cliente = Decimal(valor_cliente_str or "0").quantize(Decimal("0.01"))
            total_redondeado = Decimal(total_por_cobrar_con_iva).quantize(Decimal("0.01"))

            if valor_cliente < 0:
                messagebox.showerror("Error", "El valor ingresado no puede ser negativo.")
                return descuento_actual, None

            if valor_cliente == total_redondeado:
                porcentaje_mostrado = Decimal("0.00")
                descuento_exacto = Decimal("0.00")
            elif valor_cliente == 0:
                porcentaje_mostrado = Decimal("100.00")
                descuento_exacto = Decimal("1.00")
            else:
                descuento_exacto = 1 - (valor_cliente / total_redondeado)
                porcentaje_mostrado = descuento_exacto * 100

            lbl_descuento.config(text=f"% Descuento: {porcentaje_mostrado:.2f}%")
            btn_aceptar.config(state="normal")

            return descuento_exacto, valor_cliente

        except Exception as e:
            messagebox.showerror("Error", f"Error al calcular descuento: {e}")
            return descuento_actual, None

    tk.Button(confirm_window, text="Calcular Descuento", command=calcular_descuento).pack(pady=5)

    def limitar_decimal(valor, precision=10, escala=2):
        if not isinstance(valor, Decimal):
            valor = Decimal(valor)
        factor = Decimal("1" + ("0" * escala))
        return (valor * factor).to_integral_value(rounding="ROUND_HALF_UP") / factor

    def confirmar_y_procesar():
        nuevo_descuento, valor_ingresado_cliente = calcular_descuento()
        try:
            total_original = sum(p["ValorXCobrarConIVA"] for p in productos_agregados)

            # Distribuir valores de forma precisa
            valores_distribuidos = []
            total_asignado = Decimal("0.00")
            for i, producto in enumerate(productos_agregados):
                participacion = producto["ValorXCobrarConIVA"] / total_original if total_original else Decimal("0.00")
                if i < len(productos_agregados) - 1:
                    nuevo_valor_individual = (valor_ingresado_cliente * participacion).quantize(Decimal("0.01"))
                    total_asignado += nuevo_valor_individual
                else:
                    nuevo_valor_individual = valor_ingresado_cliente - total_asignado
                valores_distribuidos.append(nuevo_valor_individual)

            for i, producto in enumerate(productos_agregados):
                cantidad_vendida = Decimal(producto["CantidadVendida"])
                pvp_sin_iva = Decimal(producto["PVPSinIVA"])
                iva = Decimal(producto["PorcentajeIVA"])
                porcentaje_comision_consultor = Decimal(producto["PorcentajeComisionConsultor"])
                porcentaje_padre_empresarial = Decimal(producto["PorcentajePadreEmpresarial"])

                nuevo_valor_individual = valores_distribuidos[i]

                valor_original_producto = producto["ValorXCobrarConIVA"]
                if abs(valor_original_producto - nuevo_valor_individual) < Decimal("0.01"):
                    producto["PorcentajeDescuento"] = Decimal("0.0000")
                    producto["ValorDescuento"] = Decimal("0.00")
                else:
                    base_recalculada = nuevo_valor_individual / (1 + iva)
                    nuevo_descuento_unitario = (pvp_sin_iva - (base_recalculada / cantidad_vendida)) / pvp_sin_iva
                    producto["PorcentajeDescuento"] = round(nuevo_descuento_unitario * 100, 4)
                    producto["ValorDescuento"] = round(pvp_sin_iva * nuevo_descuento_unitario, 2)

                producto["ValorXCobrarConIVA"] = nuevo_valor_individual

                base_retencion = (pvp_sin_iva - producto["ValorDescuento"]) * cantidad_vendida
                producto["BaseRetencion"] = round(base_retencion, 6)
                producto["ValorBaseRetencion"] = round(base_retencion * iva, 2)
                producto["ComisionPorPagarConsultor"] = round(base_retencion * porcentaje_comision_consultor, 2)
                producto["ComisionPorPagarPadreEmpresarial"] = round(base_retencion * porcentaje_padre_empresarial, 2)

                cursor.execute("SELECT CostoConIVA FROM Productos WHERE CodigoProducto = ?", (producto["CodigoProducto"],))
                costo_unitario_con_iva = Decimal(cursor.fetchone()[0])
                producto["CostoConIVA"] = round(costo_unitario_con_iva * cantidad_vendida, 2)

            columnas_ordencompra = [
                "NumOrdenCompra", "FechaOrdenCompra", "NombreCliente", "Telefono",
                "Ciudad", "NombreConsultor", "PorcentajeComisionConsultor", "ComisionPorPagarConsultor", "NombrePadreEmpresarial",
                "PorcentajePadreEmpresarial", "ComisionPorPagarPadreEmpresarial", "PorcentajeIVA", "CodigoProducto",
                "NombreProducto", "CantidadVendida", "PorcentajeDescuento", "PrecioVentaConIVA", "PVPSinIVA",
                "ValorDescuento", "BaseRetencion", "ValorBaseRetencion",
                "ValorCliente", "ValorXCobrarConIVA"
            ]

            def convertir_a_sql(campo, valor):
                if isinstance(valor, Decimal):
                    if campo in ["PorcentajeDescuento", "PorcentajeComisionConsultor", "PorcentajePadreEmpresarial"]:
                        return max(min(round(valor, 4), Decimal("9999")), Decimal("-9999"))
                    elif campo == "BaseRetencion":
                        return max(min(round(valor, 6), Decimal("9999999999.999999")), Decimal("-9999999999.999999"))
                    elif campo in ["ValorDescuento", "ValorCliente", "ValorXCobrarConIVA", "ValorBaseRetencion", "PVPSinIVA", "PrecioVentaConIVA"]:
                        return max(min(round(valor, 2), Decimal("9999999999.99")), Decimal("-9999999999.99"))
                    else:
                        return limitar_decimal(valor)
                return valor

            for producto in productos_agregados:
                cursor.execute(f"""
                    INSERT INTO OrdenCompra (
                        {', '.join(columnas_ordencompra)}
                    ) VALUES (
                        {', '.join(['?'] * len(columnas_ordencompra))}
                    )
                """, tuple(convertir_a_sql(campo, producto.get(campo)) for campo in columnas_ordencompra))

                cursor.execute("""
                    UPDATE Productos 
                    SET CantidadInventario = CantidadInventario - ?, 
                        CantidadVendida = CantidadVendida + ? 
                    WHERE CodigoProducto = ?
                """, (producto["CantidadVendida"], producto["CantidadVendida"], producto["CodigoProducto"]))

            conn.commit()
            messagebox.showinfo("Éxito", "Orden Compra confirmada correctamente.")
            productos_agregados.clear()
            text_area.delete("1.0", tk.END)
            confirm_window.destroy()
            entries_ordencompra["📜 Orden de Compra*"].delete(0, tk.END)

            for key, entry in entries_ordencompra.items():
                if isinstance(entry, tk.Entry):
                    entry.delete(0, tk.END)
                elif isinstance(entry, ttk.Combobox):
                    entry.set("")
                elif isinstance(entry, ttk.DateEntry):
                    entry.set_date("")

        except Exception as e:
            messagebox.showerror("Error", f"Error al confirmar orden compra: {e}")

    btn_aceptar = tk.Button(confirm_window, text="Aceptar", state="disabled", command=confirmar_y_procesar)
    btn_aceptar.pack(pady=10)

    confirm_window.geometry("500x400")




def cargar_datos_ordencompra():
    """Carga los datos de una orden compra existente y los llena en los campos de entrada."""
    try:
        id_ordencompra = entries_ordencompra["🆔 IdOrdenCompra"].get().strip()
        if not id_ordencompra:
            messagebox.showerror("Error", "Debe ingresar el IdOrdenCompra para cargar los datos.")
            return

        # Convertir a entero antes de usarlo
        try:
            id_ordencompra = int(id_ordencompra)
        except ValueError:
            messagebox.showerror("Error", "El IdOrdenCompra debe ser un número entero válido.")
            return

        cursor.execute("""
        SELECT IdOrdenCompra,NumOrdenCompra, FechaOrdenCompra, NombreCliente, Telefono, Ciudad,
            NombreConsultor, PorcentajeComisionConsultor, ComisionPorPagarConsultor, NombrePadreEmpresarial, PorcentajePadreEmpresarial, 
            ComisionPorPagarPadreEmpresarial, PorcentajeIVA, CodigoProducto, NombreProducto, CantidadVendida,
            PorcentajeDescuento, PrecioVentaConIVA, PVPSinIVA, ValorDescuento, BaseRetencion, ValorBaseRetencion,
            ValorCliente, ValorXCobrarConIVA

        FROM OrdenCompra
        WHERE IdOrdenCompra = ?
        """, (id_ordencompra,))

        ordencompra = cursor.fetchone()

        if not ordencompra:
            messagebox.showerror("Error", "No se encontró una orden de compra con ese Id.")
            return

        # Desempaquetar correctamente los valores de la consulta SQL
        (
            id_ordencompra, orden_compra, fecha_orden, nombre_cliente, telefono, ciudad,
            nombre_consultor, porcentaje_comision_consultor, comision_por_pagar_consultor,
            nombre_padre_empresarial, porcentaje_padre_empresarial, comision_por_pagar_padre_empresarial,
            porcentaje_iva, codigo_producto, nombre_producto, cantidad_vendida,
            porcentaje_descuento, precio_venta_con_iva, pvp_sin_iva, valor_descuento,
            base_retencion, valor_base_retencion, valor_cliente, 
            valor_x_cobrar_con_iva
            
            
        ) = ordencompra

        # Asegurar que los valores no sean None (evita errores en la interfaz)
        def valor_defecto(valor, default=0):
            return valor if valor is not None else default

        cantidad_vendida = Decimal(valor_defecto(cantidad_vendida))
        precio_venta_con_iva = Decimal(valor_defecto(precio_venta_con_iva))
        pvp_sin_iva = Decimal(valor_defecto(pvp_sin_iva))
        porcentaje_descuento = Decimal(valor_defecto(porcentaje_descuento))
        valor_descuento = Decimal(valor_defecto(valor_descuento))
        base_retencion = Decimal(valor_defecto(base_retencion))
        porcentaje_iva = Decimal(valor_defecto(porcentaje_iva))
        valor_base_retencion = Decimal(valor_defecto(valor_base_retencion))
        valor_x_cobrar_con_iva = Decimal(valor_defecto(valor_x_cobrar_con_iva))




        # Asignar valores a los campos en la interfaz
        entries_ordencompra["📜 Orden de Compra*"].delete(0, tk.END)
        entries_ordencompra["📜 Orden de Compra*"].insert(0, orden_compra)
        entries_ordencompra["📅 Fecha Orden De Compra (YYYY-MM-DD)*"].set_date(fecha_orden)
        entries_ordencompra["👤 Nombre del Cliente*"].delete(0, tk.END)
        entries_ordencompra["👤 Nombre del Cliente*"].insert(0, nombre_cliente)
        entries_ordencompra["☎️ Telefono"].delete(0, tk.END)
        entries_ordencompra["☎️ Telefono"].insert(0, telefono)
        entries_ordencompra["🌆 Ciudad"].delete(0, tk.END)
        entries_ordencompra["🌆 Ciudad"].insert(0, ciudad)
        entries_ordencompra["👤 20% Consultor"].delete(0, tk.END)
        entries_ordencompra["👤 20% Consultor"].insert(0, nombre_consultor)
        entries_ordencompra["👤 5% Padre Empresarial"].delete(0, tk.END)
        entries_ordencompra["👤 5% Padre Empresarial"].insert(0, nombre_padre_empresarial)
        entries_ordencompra["💲 IVA (%)*"].delete(0, tk.END)
        entries_ordencompra["💲 IVA (%)*"].insert(0, f"{porcentaje_iva * 100:.2f}")
        entries_ordencompra["🏷️ Código del Producto*"].delete(0, tk.END)
        entries_ordencompra["🏷️ Código del Producto*"].insert(0, codigo_producto)
        entries_ordencompra["📦 Nombre del Producto"].delete(0, tk.END)
        entries_ordencompra["📦 Nombre del Producto"].insert(0, nombre_producto)
        entries_ordencompra["📊 Cantidad Vendida*"].delete(0, tk.END)
        entries_ordencompra["📊 Cantidad Vendida*"].insert(0, str(cantidad_vendida))



        messagebox.showinfo("Éxito", "Datos cargados correctamente.")

    except Exception as e:
        messagebox.showerror("Error", f"Error al cargar los datos: {e}")
            
def modificar_ordencompra():
    try:
        id_ordencompra = entries_ordencompra["🆔 IdOrdenCompra"].get().strip()
        if not id_ordencompra:
            messagebox.showerror("Error", "Debe ingresar el IdOrdenCompra para modificar la orden de compra.")
            return

        orden_compra = entries_ordencompra["📜 Orden de Compra*"].get().strip()
        fecha_orden = entries_ordencompra["📅 Fecha Orden De Compra (YYYY-MM-DD)*"].get_date().strftime('%Y-%m-%d')
        nombre_cliente = entries_ordencompra["👤 Nombre del Cliente*"].get().strip()
        telefono = entries_ordencompra["☎️ Telefono"].get().strip()
        ciudad = entries_ordencompra["🌆 Ciudad"].get().strip()
        consultor = entries_ordencompra[" 20% Consultor"].get().strip()
        padre_empresarial = entries_ordencompra["👤 5% Padre Empresarial"].get().strip()
        iva_nuevo = Decimal(entries_ordencompra["💲 IVA (%)*"].get().strip()) / 100
        codigo_producto_nuevo = entries_ordencompra["🏷️ Código del Producto*"].get().strip()
        cantidad_vendida_nueva = Decimal(entries_ordencompra["📊 Cantidad Vendida*"].get().strip() or "0")

        # Validaciones
        if not orden_compra.isdigit():
            messagebox.showerror("Error", "El número de Orden de Compra debe ser un número entero.")
            return
        if not re.match(r"^[a-zA-ZÁÉÍÓÚáéíóúñÑ\s]+$", nombre_cliente):
            messagebox.showerror("Error", "El Nombre del Cliente solo debe contener letras.")
            return
        if telefono and not re.match(r"^\d{6,15}$", telefono):
            messagebox.showerror("Error", "El Teléfono debe contener entre 6 y 15 dígitos.")
            return
        if ciudad and not re.match(r"^[a-zA-ZÁÉÍÓÚáéíóúñÑ\s]+$", ciudad):
            messagebox.showerror("Error", "La Ciudad solo debe contener letras.")
            return

        # Obtener datos originales
        cursor.execute("""
            SELECT CodigoProducto, CantidadVendida, PorcentajeIVA, NombreConsultor, NombrePadreEmpresarial,
                   NumOrdenCompra, FechaOrdenCompra, NombreCliente, Telefono, Ciudad
            FROM OrdenCompra
            WHERE IdOrdenCompra = ?
        """, (id_ordencompra,))
        datos_anteriores = cursor.fetchone()

        if not datos_anteriores:
            messagebox.showerror("Error", "No se encontró una orden de compra con ese ID.")
            return

        (
            codigo_producto_original, cantidad_vendida_anterior, iva_anterior, consultor_ant, padre_emp_ant,
            num_orden_ant, fecha_ant, nombre_cliente_ant, telefono_ant, ciudad_ant
        ) = datos_anteriores

        # Verificar modificación de campos no permitidos
        campos_prohibidos_modificados = (
            codigo_producto_nuevo != codigo_producto_original or
            cantidad_vendida_nueva != Decimal(cantidad_vendida_anterior) or
            iva_nuevo != Decimal(iva_anterior) or
            consultor != (consultor_ant or "") or
            padre_empresarial != (padre_emp_ant or "")
        )

        if campos_prohibidos_modificados:
            cursor.execute("SELECT COUNT(*) FROM CuentasPorCobrar WHERE NumOrdenCompra = ?", (orden_compra,))
            cuentas_cobrar = cursor.fetchone()[0]
            cursor.execute("SELECT COUNT(*) FROM CuentasPorPagarConsultor WHERE NombreConsultor = ?", (consultor_ant,))
            cuentas_pagar_consultor = cursor.fetchone()[0]
            cursor.execute("SELECT COUNT(*) FROM CuentasPorPagarPadreEmpresarial WHERE NombrePadreEmpresarial = ?", (padre_emp_ant,))
            cuentas_pagar_padre = cursor.fetchone()[0]

            messagebox.showwarning(
                "Modificación no permitida",
                f"No se puede modificar Código Producto, IVA, Consultor o Padre Empresarial.\n\n"
                f"⚠ Para hacerlo, elimine todos los productos de la orden o la orden completa.\n\n"
                f"Esto eliminaría:\n"
                f"🧾 {cuentas_cobrar} registros en CuentasPorCobrar\n"
                f"👤 {cuentas_pagar_consultor} en CuentasPorPagarConsultor\n"
                f"🏢 {cuentas_pagar_padre} en CuentasPorPagarPadreEmpresarial"
            )
            return

        # Si se modificaron los campos permitidos, actualizar en todas las filas
        if (
            fecha_orden != fecha_ant or
            nombre_cliente != nombre_cliente_ant or
            telefono != telefono_ant or
            ciudad != ciudad_ant
        ):
            # Actualizar en OrdenCompra
            cursor.execute("""
                UPDATE OrdenCompra
                SET FechaOrdenCompra = ?, NombreCliente = ?, Telefono = ?, Ciudad = ?
                WHERE NumOrdenCompra = ?
            """, (fecha_orden, nombre_cliente, telefono, ciudad, orden_compra))
            # Actualizar también en CuentasPorCobrar para mantener sincronizado el NombreCliente
            cursor.execute("""
                UPDATE CuentasPorCobrar
                SET NombreCliente = ?
                WHERE NumOrdenCompra = ?
            """, (nombre_cliente, orden_compra))


            conn.commit()
            messagebox.showinfo("✅ Éxito", "Datos modificados correctamente en todas las órdenes y cuentas por cobrar.")

        # Limpiar formulario
        for key, entry in entries_ordencompra.items():
            if isinstance(entry, tk.Entry): entry.delete(0, tk.END)
            elif isinstance(entry, ttk.Combobox): entry.set("")
            elif isinstance(entry, ttk.DateEntry): entry.set_date("")

    except Exception as e:
        messagebox.showerror("❌ Error", f"Error al modificar la orden: {e}")


def eliminar_ordencompra():
    try:
        id_ordencompra = entries_ordencompra["🆔 IdOrdenCompra"].get().strip()
        if not id_ordencompra:
            messagebox.showerror("Error", "Debe ingresar el IdOrdenCompra para eliminar.")
            return

        cursor.execute("""
            SELECT CodigoProducto, CantidadVendida, NumOrdenCompra, NombreConsultor, NombrePadreEmpresarial
            FROM OrdenCompra WHERE IdOrdenCompra = ?
        """, (id_ordencompra,))
        orden = cursor.fetchone()

        if not orden:
            messagebox.showerror("Error", "No se encontró una orden con ese ID.")
            return

        codigo, cantidad, num_orden, consultor, padre = orden
        cantidad = Decimal(cantidad)

        # Obtener registros relacionados solo de esta orden
        cursor.execute("SELECT * FROM CuentasPorCobrar WHERE NumOrdenCompra = ?", (num_orden,))
        cxc = cursor.fetchall()

        cursor.execute("""
            SELECT * FROM CuentasPorPagarConsultor
            WHERE NombreConsultor = ? AND NumOrdenCompra = ?
        """, (consultor, num_orden))
        cppc = cursor.fetchall()

        cursor.execute("""
            SELECT * FROM CuentasPorPagarPadreEmpresarial
            WHERE NombrePadreEmpresarial = ? AND NumOrdenCompra = ?
        """, (padre, num_orden))
        cppp = cursor.fetchall()

        # Mensaje de confirmación detallado
        mensaje = f"""
🧾 Orden #{id_ordencompra} con número: {num_orden}
📦 Producto: {codigo}
📤 Cantidad vendida: {cantidad}

👤 Consultor vinculado: {consultor}
👥 Padre Empresarial vinculado: {padre}

🔴 Registros relacionados que se eliminarán solo para esta orden:
- {len(cxc)} en CuentasPorCobrar (por NumOrdenCompra)
- {len(cppc)} en CuentasPorPagarConsultor (por NombreConsultor y NumOrdenCompra)
- {len(cppp)} en CuentasPorPagarPadreEmpresarial (por NombrePadreEmpresarial y NumOrdenCompra)

⚠️ Todos estos datos se eliminarán permanentemente solo para esta orden de compra.
Deberás ingresarlos nuevamente si fue un error.
"""

        confirm = messagebox.askyesno("Confirmar eliminación", mensaje)
        if not confirm:
            return

        # Restaurar inventario
        cursor.execute("SELECT CantidadInventario, CantidadVendida FROM Productos WHERE CodigoProducto = ?", (codigo,))
        inv, vend = cursor.fetchone()
        inv_nuevo = Decimal(inv) + cantidad
        vend_nuevo = max(Decimal(vend) - cantidad, 0)

        cursor.execute("""
            UPDATE Productos
            SET CantidadInventario = ?, CantidadVendida = ?
            WHERE CodigoProducto = ?
        """, (inv_nuevo, vend_nuevo, codigo))

        # Eliminar registros
        cursor.execute("DELETE FROM OrdenCompra WHERE IdOrdenCompra = ?", (id_ordencompra,))
        cursor.execute("DELETE FROM CuentasPorCobrar WHERE NumOrdenCompra = ?", (num_orden,))
        cursor.execute("""
            DELETE FROM CuentasPorPagarConsultor
            WHERE NombreConsultor = ? AND NumOrdenCompra = ?
        """, (consultor, num_orden))
        cursor.execute("""
            DELETE FROM CuentasPorPagarPadreEmpresarial
            WHERE NombrePadreEmpresarial = ? AND NumOrdenCompra = ?
        """, (padre, num_orden))
        
        conn.commit()

        messagebox.showinfo("Éxito", "Orden y registros relacionados eliminados correctamente.")

        for entry in entries_ordencompra.values():
            if isinstance(entry, (tk.Entry, ttk.Combobox)):
                entry.delete(0, tk.END)
            elif isinstance(entry, DateEntry):
                entry.set_date("")

    except Exception as e:
        messagebox.showerror("Error", f"Error al eliminar orden: {str(e)}")


def eliminar_por_numordencompra():
    try:
        num_orden = entries_ordencompra["🔢 NumOrdenCompra"].get().strip()
        if not num_orden:
            messagebox.showerror("Error", "Debe ingresar un número de orden.")
            return

        cursor.execute("""
            SELECT IdOrdenCompra, CodigoProducto, CantidadVendida, NombreConsultor, NombrePadreEmpresarial
            FROM OrdenCompra WHERE NumOrdenCompra = ?
        """, (num_orden,))
        ordenes = cursor.fetchall()

        if not ordenes:
            messagebox.showinfo("Sin resultados", f"No se encontraron órdenes con número {num_orden}.")
            return

        codigos = []
        total_vendidas = Decimal(0)
        consultores = set()
        padres = set()

        for o in ordenes:
            codigos.append(o[1])
            total_vendidas += Decimal(o[2])
            consultores.add(o[3])
            padres.add(o[4])

        # Obtener registros relacionados
        cursor.execute("SELECT COUNT(*) FROM CuentasPorCobrar WHERE NumOrdenCompra = ?", (num_orden,))
        total_cxc = cursor.fetchone()[0]

        total_cppc = sum(cursor.execute(
            "SELECT COUNT(*) FROM CuentasPorPagarConsultor WHERE NombreConsultor = ?", (c,)
        ).fetchone()[0] for c in consultores)

        total_cppp = sum(cursor.execute(
            "SELECT COUNT(*) FROM CuentasPorPagarPadreEmpresarial WHERE NombrePadreEmpresarial = ?", (p,)
        ).fetchone()[0] for p in padres)

        # Mensaje confirmación detallado
        mensaje = f"""
🔢 Número de Orden: {num_orden}
🧾 Total de órdenes con ese número: {len(ordenes)}
📦 Productos involucrados: {', '.join(set(codigos))}
📤 Total unidades vendidas a devolver al inventario: {total_vendidas}

👤 Consultores vinculados en esta orden: {', '.join(consultores)}
👥 Padres Empresariales vinculados: {', '.join(padres)}

🔴 Registros relacionados que se eliminarán solo para esta orden:
- {total_cxc} en CuentasPorCobrar (por NumOrdenCompra)
- Pagos del consultor en CuentasPorPagarConsultor (solo NumOrdenCompra={num_orden})
- Pagos del padre empresarial en CuentasPorPagarPadreEmpresarial (solo NumOrdenCompra={num_orden})

⚠️ Estos datos serán eliminados permanentemente solo para este número de orden.
Deberás volver a ingresarlos si fue un error.
"""


        confirm = messagebox.askyesno("Confirmar eliminación masiva", mensaje)
        if not confirm:
            return

        for id_orden, cod, cant, *_ in ordenes:
            cursor.execute("SELECT CantidadInventario, CantidadVendida FROM Productos WHERE CodigoProducto = ?", (cod,))
            inv, vend = cursor.fetchone()
            nueva_inv = Decimal(inv) + Decimal(cant)
            nueva_vend = max(Decimal(vend) - Decimal(cant), 0)

            cursor.execute("""
                UPDATE Productos
                SET CantidadInventario = ?, CantidadVendida = ?
                WHERE CodigoProducto = ?
            """, (nueva_inv, nueva_vend, cod))

            cursor.execute("DELETE FROM OrdenCompra WHERE IdOrdenCompra = ?", (id_orden,))

        cursor.execute("DELETE FROM CuentasPorCobrar WHERE NumOrdenCompra = ?", (num_orden,))

        cursor.execute("DELETE FROM CuentasPorPagarConsultor WHERE NumOrdenCompra = ?", (num_orden,))
        cursor.execute("DELETE FROM CuentasPorPagarPadreEmpresarial WHERE NumOrdenCompra = ?", (num_orden,))


        conn.commit()

        messagebox.showinfo("Éxito", "Todas las órdenes y registros relacionados han sido eliminados.")

        for entry in entries_ordencompra.values():
            if isinstance(entry, (tk.Entry, ttk.Combobox)):
                entry.delete(0, tk.END)
            elif isinstance(entry, DateEntry):
                entry.set_date("")

    except Exception as e:
        messagebox.showerror("Error", f"Error general al eliminar: {str(e)}")




# ---------- BOTONES DE ACCIÓN (Debajo del Formulario) ------------
frame_botones = ttk.Frame(tab_orden_compra)
frame_botones.grid(row=1, column=0, columnspan=2, pady=10, sticky="w")  # <- Añade sticky="w" aquí

# Botones alineados a la izquierda

tk.Button(frame_botones, text="⏪ Deshacer Último Producto", command=deshacer_ultimo_producto, bg="tomato").grid(column=5, row=0, padx=5, pady=5, sticky="w")
tk.Button(frame_botones, text="➕ Agregar Producto a Orden", command=agregar_producto_a_orden,bg="lightgreen").grid(column=6, row=0, padx=5, pady=5, sticky="w")
tk.Button(frame_botones, text="✅ Confirmar Orden de Compra", command=confirmar_ordencompra,bg="lightgreen").grid(column=7, row=0, padx=5, pady=30, sticky="w")

# Nuevo campo IdOrdenCompra alineado
ttk.Label(frame_botones, text="🆔 IdOrdenCompra:").grid(column=0, row=1, padx=5, pady=5, sticky="w")  # <- sticky="w"
entry_id_ordencompra = ttk.Entry(frame_botones, width=15)
entry_id_ordencompra.grid(column=2, row=1, padx=5, pady=5, sticky="w")  # <- sticky="w"


# ¡IMPORTANTE! Aquí es donde se añade al diccionario para que cargar_datos2 y otras funciones lo puedan leer.
entries_ordencompra["🆔 IdOrdenCompra"] = entry_id_ordencompra

# Campo NumOrdenCompra y botón eliminar en la misma fila, con distribución estética
frame_numorden = ttk.Frame(frame_botones)
frame_numorden.grid(column=6, row=1, columnspan=3, padx=10, pady=5, sticky="w")

ttk.Label(frame_numorden, text="🆔 Orden de Compra:").grid(row=0, column=0, padx=(0, 5), sticky="w")
entry_num_ordencompra = ttk.Entry(frame_numorden, width=15)
entry_num_ordencompra.grid(row=0, column=1, padx=(0, 15), sticky="w")
btn_eliminar_numorden = tk.Button(frame_numorden, text="🗑️ Eliminar Orden de Compra", command=eliminar_por_numordencompra,bg="tomato")
btn_eliminar_numorden.grid(row=0, column=2, sticky="w")

entries_ordencompra["🔢 NumOrdenCompra"] = entry_num_ordencompra



# Botones alineados a la izquierda
tk.Button(frame_botones, text="🔄 Cargar Datos", command=cargar_datos_ordencompra, bg="lightgreen").grid(column=3, row=1, padx=5, pady=5, sticky="w")
tk.Button(frame_botones, text="✏️ Modificar Producto", command=modificar_ordencompra, bg="skyblue").grid(column=4, row=1, padx=5, pady=5, sticky="w")
tk.Button(frame_botones, text="🗑️ Eliminar Producto", command=eliminar_ordencompra, bg="tomato").grid(column=5, row=1, padx=5, pady=5, sticky="w")

# Crear un marco para la búsqueda de productos
frame_buscar_productos2 = ttk.Frame(tab_orden_compra)
frame_buscar_productos2.grid(row=1, column=0, columnspan=3, pady=10, sticky="w")

ttk.Label(frame_buscar_productos2, text="🔍 Buscar Producto:").grid(row=0, column=0, padx=5)
entry_buscar_producto2 = ttk.Entry(frame_buscar_productos2, width=30)
entry_buscar_producto2.grid(row=0, column=1, padx=5)

# Botón para buscar productos
btn_buscar_producto2 = tk.Button(frame_buscar_productos2, text="🔍 Buscar", bg="lightgreen", command=lambda: buscar_producto2(entry_buscar_producto2.get().strip()))
btn_buscar_producto2.grid(row=0, column=2, padx=5)

def buscar_producto2(nombre_producto):
    """Busca productos por nombre y muestra los resultados en una nueva ventana con scroll horizontal y vertical."""
    if not nombre_producto:
        messagebox.showwarning("Atención", "Por favor, ingrese un nombre de producto para buscar.")
        return

    cursor.execute("""
        SELECT CodigoProducto, NombreProducto 
        FROM Productos 
        WHERE NombreProducto LIKE ?
    """, ('%' + nombre_producto + '%',))
    
    rows = cursor.fetchall()

    if not rows:
        messagebox.showinfo("Resultado", "No se encontraron productos con ese nombre.")
        return

    ventana_resultados = tk.Toplevel(root)
    ventana_resultados.title("Resultados de Búsqueda")
    ventana_resultados.geometry("800x500")

    frame_resultados = tk.Frame(ventana_resultados)
    frame_resultados.pack(fill="both", expand=True, padx=10, pady=10)

    scroll_y = tk.Scrollbar(frame_resultados, orient="vertical")
    scroll_x = tk.Scrollbar(frame_resultados, orient="horizontal")
    text_resultados = tk.Text(frame_resultados, wrap="none", font=("Arial", 11),
                              yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)

    scroll_y.config(command=text_resultados.yview)
    scroll_x.config(command=text_resultados.xview)

    scroll_y.pack(side="right", fill="y")
    scroll_x.pack(side="bottom", fill="x")
    text_resultados.pack(side="left", fill="both", expand=True)

    for row in rows:
        codigo_producto, nombre_producto = row
        text_resultados.insert(tk.END, f"📌 Código: {codigo_producto}\n🔹 Producto: {nombre_producto}\n\n")

    text_resultados.config(state="disabled")




# ---------------------- PESTAÑA CUENTAS POR COBRAR ----------------------

productos_cxc = []

frame_cxc_contenedor = ttk.Frame(tab_cuentas_cobrar)
frame_cxc_contenedor.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)

frame_izquierdo_cxc = ttk.Frame(frame_cxc_contenedor)
frame_izquierdo_cxc.grid(row=0, column=0, sticky="nw", padx=10, pady=10)

frame_derecho_cxc = ttk.Frame(frame_cxc_contenedor)
frame_derecho_cxc.grid(row=0, column=1, sticky="nsew", padx=10, pady=10)

frame_textarea_y_id = ttk.Frame(frame_derecho_cxc)
frame_textarea_y_id.pack(fill="both", expand=True)

frame_idcuenta_seccion = ttk.Frame(frame_derecho_cxc)
frame_idcuenta_seccion.pack(fill="x", pady=10)



# TextArea para mostrar los pagos agregados
ttk.Label(frame_derecho_cxc, text="Pagos Agregados a la Orden:", font=("Arial", 12, "bold")).pack(anchor="w", pady=5)

text_area_cxc = tk.Text(frame_textarea_y_id, width=85, height=29, wrap="word")
text_area_cxc.pack(side="left", fill="both", expand=True, padx=5, pady=5)
text_area_cxc.config(state="disabled")  # ⛔️ Lo vuelve de solo lectura


scrollbar_cxc = ttk.Scrollbar(frame_textarea_y_id, command=text_area_cxc.yview)
scrollbar_cxc.pack(side="right", fill="y")
text_area_cxc.config(yscrollcommand=scrollbar_cxc.set)



# Campo: Número de Orden de Compra
ttk.Label(frame_izquierdo_cxc, text="🔎 Número Orden de Compra:").grid(row=0, column=0, sticky="w")
entry_orden_cxc = ttk.Entry(frame_izquierdo_cxc, width=30)
entry_orden_cxc.valor_anterior = ""
entry_orden_cxc.grid(row=0, column=1, padx=5, pady=5)

# Nuevo frame debajo del TextArea
frame_idcuenta_seccion = ttk.Frame(frame_derecho_cxc)
frame_idcuenta_seccion.pack(fill="x", pady=10)

# Campo IdCuenta
ttk.Label(frame_idcuenta_seccion, text="🆔 IdCuenta:").pack(side="left", padx=5)
entry_id_cuenta = ttk.Entry(frame_idcuenta_seccion, width=20)
entry_id_cuenta.pack(side="left", padx=5)



def actualizar_encabezado_saldo():
    ab1 = sum(p["valor"] for p in pagos_cxc if p["tipo"] == "Efectivo / Transferencia")
    ab2 = sum(p["valor"] for p in pagos_cxc if p["tipo"] == "Tarjeta de Crédito")
    total_pagado = Decimal(ab1 + ab2)

    cursor.execute("""
        SELECT SUM(ValorXCobrarConIVA)
        FROM OrdenCompra
        WHERE NumOrdenCompra = ?
    """, (entry_orden_cxc.get().strip(),))
    total_orden = cursor.fetchone()[0] or 0

    saldo = Decimal(total_orden) - total_pagado
    cursor.execute("""
        SELECT TOP 1 NombreCliente FROM OrdenCompra WHERE NumOrdenCompra = ?
    """, (entry_orden_cxc.get().strip(),))
    nombre_cliente = cursor.fetchone()[0] or "Desconocido"

    # 🔁 RECONSTRUIR TODO
    text_area_cxc.config(state="normal")
    text_area_cxc.delete("1.0", tk.END)

    # Encabezado
    text_area_cxc.insert(tk.END, f"👤 Nombre Cliente: {nombre_cliente}\n")
    text_area_cxc.insert(tk.END, f"💰 Saldo por Cobrar Cliente: ${saldo:,.2f}\n\n")

    abono_num = 1
    tarjeta_num = 1
    for pago in pagos_cxc:
        if pago["tipo"] == "Efectivo / Transferencia":
            text_area_cxc.insert(tk.END, f"🟢 Abono #{abono_num} - Tipo: Efectivo/Transferencia\n")
            text_area_cxc.insert(tk.END, f"   💲 Valor: ${pago['valor']:,.2f} - 📅 Fecha: {pago['fecha']}\n\n")
            abono_num += 1
        elif pago["tipo"] == "Tarjeta de Crédito":
            text_area_cxc.insert(tk.END, f"💳 Tarjeta #{tarjeta_num} - Tipo: {pago['modo']}\n")
            text_area_cxc.insert(tk.END, f"🏦 Banco: {pago['banco']}\n💲 Valor: ${pago['valor']:,.2f} - 🔢 Lote: {pago['lote']} - 📅 Fecha: {pago['fecha']}\n")
            if pago['banco'].startswith("Personalizado"):
                porcentaje = porcentaje_comision_banco_dict.get(pago['banco'], Decimal(0)) * 100
                text_area_cxc.insert(tk.END, f"💡 Comisión personalizada aplicada: {porcentaje:.2f}%\n")
            text_area_cxc.insert(tk.END, "\n")
            tarjeta_num += 1

    text_area_cxc.config(state="disabled")

def reconstruir_text_area():
    text_area_cxc.config(state="normal")
    text_area_cxc.delete("3.0", tk.END)  # Mantener encabezado intacto (líneas 1 y 2)

    abono_num = 1
    tarjeta_num = 1

    for pago in pagos_cxc:
        if pago["tipo"] == "Efectivo / Transferencia":
            text_area_cxc.insert(tk.END, f"🟢 Abono #{abono_num} - Tipo: Efectivo/Transferencia\n")
            text_area_cxc.insert(tk.END, f"   💲 Valor: ${pago['valor']:,.2f} - 📅 Fecha: {pago['fecha']}\n\n")
            abono_num += 1
        elif pago["tipo"] == "Tarjeta de Crédito":
            text_area_cxc.insert(tk.END, f"💳 Tarjeta #{tarjeta_num} - Tipo: {pago['modo']}\n")
            text_area_cxc.insert(tk.END, f"🏦 Banco: {pago['banco']}\n💲 Valor: ${pago['valor']:,.2f} - 🔢 Lote: {pago['lote']} - 📅 Fecha: {pago['fecha']}\n")
            if pago['banco'].startswith("Personalizado"):
                porcentaje = porcentaje_comision_banco_dict.get(pago['banco'], Decimal(0)) * 100
                text_area_cxc.insert(tk.END, f"💡 Comisión personalizada aplicada: {porcentaje:.2f}%\n")
            text_area_cxc.insert(tk.END, "\n")
            tarjeta_num += 1

    text_area_cxc.config(state="disabled")







# Botón para validar la orden
def verificar_orden_compra():
    num_orden = entry_orden_cxc.get().strip()
    if not num_orden:
        messagebox.showerror("Error", "Debe ingresar un número de orden de compra.")
        return

    cursor.execute("SELECT * FROM OrdenCompra WHERE NumOrdenCompra = ?", (num_orden,))
    datos_orden = cursor.fetchall()

    if not datos_orden:
        messagebox.showerror("No encontrado", f"No existe ninguna orden con número {num_orden}.")
        return
    # 🔐 Validar si ya existe en cuentas por cobrar
    cursor.execute("SELECT COUNT(*) FROM CuentasPorCobrar WHERE NumOrdenCompra = ?", (num_orden,))
    existe_cuenta = cursor.fetchone()[0]

    if existe_cuenta > 0:
        messagebox.showerror(
            "Ya registrado",
            f"Ya existe una cuenta por cobrar para la orden #{num_orden}.\n\n"
            "🔎 BUSQUE el IdCuenta en la pestaña 'Reportes' aplastando el botón 'Mostrar Cuentas por Cobrar Pagadas'.\n\n"
            "💳 Con ese Id, podrá registrar un nuevo abono y cargar los datos en la ventana de edición para continuar con el pago de la orden."
            
        )
        return


    messagebox.showinfo("Orden encontrada", f"Se encontraron {len(datos_orden)} registros para la orden #{num_orden}")
    combo_tipo_pago.config(state="readonly")
    btn_confirmar_pago.config(state="normal")
    cursor.execute("""
        SELECT TOP 1 NombreCliente FROM OrdenCompra WHERE NumOrdenCompra = ?
    """, (num_orden,))
    nombre_cliente = cursor.fetchone()[0] or "Desconocido"

    cursor.execute("""
        SELECT SUM(ValorXCobrarConIVA)
        FROM OrdenCompra
        WHERE NumOrdenCompra = ?
    """, (num_orden,))
    total = cursor.fetchone()[0] or 0

    # ✅ Mostramos ambos datos en el TextArea, con protección de solo lectura
    text_area_cxc.config(state="normal")
    text_area_cxc.delete("1.0", tk.END)
    text_area_cxc.insert("1.0", f"👤 Nombre Cliente: {nombre_cliente}\n💰 Saldo por Cobrar Cliente: ${Decimal(total):,.2f}\n\n")
    text_area_cxc.config(state="disabled")

    label_recordatorio_buscar.grid_remove()



    # Aquí después se cargan internamente los datos necesarios para cálculos posteriores

btn_verificar_orden = tk.Button(frame_izquierdo_cxc, text="Buscar Orden", command=verificar_orden_compra,bg="lightgreen", state="disabled")
btn_verificar_orden.grid(row=0, column=2, padx=5, pady=5)


label_recordatorio_buscar = ttk.Label(frame_izquierdo_cxc, text="⚠️ Presione 'Buscar Orden'", foreground="red")
label_recordatorio_buscar.grid(row=0, column=3, padx=5)
label_recordatorio_buscar.grid_remove()  # 🔒 Oculto al inicio

# 🔁 Asociar cambio en entry_orden_cxc para revalidar
def validar_entrada_numero_orden(event=None):
    valor = entry_orden_cxc.get().strip()

    # Si ya había un valor anterior y este cambió
    if hasattr(entry_orden_cxc, "valor_anterior") and entry_orden_cxc.valor_anterior != valor:
        text_area_cxc.config(state="normal")
        text_area_cxc.delete("1.0", tk.END)
        text_area_cxc.config(state="disabled")

        pagos_cxc.clear()
        global contador_global_pagos
        contador_global_pagos = 1
        combo_tipo_pago.set("")
        combo_tipo_pago.config(state="disabled")
        btn_confirmar_pago.config(state="disabled")

    # Habilita o deshabilita el botón según si es numérico
    if valor.isdigit():
        if not hasattr(entry_orden_cxc, "valor_anterior") or entry_orden_cxc.valor_anterior != valor:
            btn_verificar_orden.config(state="normal")
            label_recordatorio_buscar.grid()
        else:
            btn_verificar_orden.config(state="disabled")
            label_recordatorio_buscar.grid_remove()
    else:
        btn_verificar_orden.config(state="disabled")
        label_recordatorio_buscar.grid_remove()


    entry_orden_cxc.valor_anterior = valor



entry_orden_cxc.bind("<KeyRelease>", validar_entrada_numero_orden)
btn_verificar_orden.config(state="disabled")  # 🔒 por defecto




# Tipo de Pago
ttk.Label(frame_izquierdo_cxc, text="💳 Tipo de Pago:").grid(row=1, column=0, sticky="w")
combo_tipo_pago = ttk.Combobox(frame_izquierdo_cxc, values=["Efectivo / Transferencia", "Tarjeta de Crédito"], state="readonly", width=30)
combo_tipo_pago.grid(row=1, column=1, padx=5, pady=5)
combo_tipo_pago.config(state="disabled")  # ⬅️ Agrega esto
# ESTA LÍNEA ES FUNDAMENTAL
combo_tipo_pago.bind("<<ComboboxSelected>>", lambda e: mostrar_formulario_pago(combo_tipo_pago.get()))


pagos_cxc = []
abonos_efectivo_transferencia = 0
pagos_tarjeta_credito = 0
contador_global_pagos = 1


opciones_banco = {
    "Efectivo": [],
    "Transferencia": [
        "BCO. GUAYAQUIL MAESTRO Y ELECTRON",
        "BCO. PACIFICO MAESTRO Y ELECTRON",
        "BCO. PICHINCHA MAESTRO Y ELECTRON",
        "Otros Bancos"
    ],
    "Corriente": [
        "BCO. GUAYAQUIL AMEX (4.62%)",
        "BCO. GUAYAQUIL MASTERCARD Y VISA (4.62%)",
        "BCO. PACÍFICO MASTERCARD Y VISA (4.62%)",
        "BCO. DINERS CLUB DINERS Y DISCOVER (4.62%)",
        "BCO. PICHINCHA MASTERCARD Y VISA (4.62%)",
        "Otros Bancos (4.62%)",
        "Ingresar Porcentaje Personalizado..."
    ],
    "Diferido": [
        "BCO. GUAYAQUIL AMEX (5.75%)",
        "BCO. GUAYAQUIL MASTERCARD Y VISA (5.75%)",
        "PACIFICARD MASTERCARD Y VISA (5.75%)",
        "BCO. DINERS CLUB DINERS Y DISCOVER (5.75%)",
        "BANCO PICHINCHA MASTERCARD Y VISA (5.75%)",
        "Otros Bancos (5.75%)",
        "Ingresar Porcentaje Personalizado..."
    ]
}

porcentaje_comision_banco_dict = {
    "BCO. GUAYAQUIL AMEX (5.75%)": Decimal("0.0575"),
    "BCO. GUAYAQUIL MASTERCARD Y VISA (5.75%)": Decimal("0.0575"),
    "PACIFICARD MASTERCARD Y VISA (5.75%)": Decimal("0.0575"),
    "BCO. DINERS CLUB DINERS Y DISCOVER (5.75%)": Decimal("0.0575"),
    "BANCO PICHINCHA MASTERCARD Y VISA (5.75%)": Decimal("0.0575"),
    "Otros Bancos (5.75%)": Decimal("0.0575"),
    "BCO. GUAYAQUIL AMEX (4.62%)": Decimal("0.0462"),
    "BCO. GUAYAQUIL MASTERCARD Y VISA (4.62%)": Decimal("0.0462"),
    "BCO. PACÍFICO MASTERCARD Y VISA (4.62%)": Decimal("0.0462"),
    "BCO. DINERS CLUB DINERS Y DISCOVER (4.62%)": Decimal("0.0462"),
    "BCO. PICHINCHA MASTERCARD Y VISA (4.62%)": Decimal("0.0462"),
    "Otros Bancos (4.62%)": Decimal("0.0462")
}

from decimal import Decimal, ROUND_HALF_UP

def mostrar_formulario_pago(tipo_pago):
    num_orden = entry_orden_cxc.get().strip()

    if not num_orden:
        messagebox.showerror("Error", "Debe ingresar primero una orden de compra válida.")
        return

    cursor.execute("""
        SELECT SUM(ValorXCobrarConIVA)
        FROM OrdenCompra
        WHERE NumOrdenCompra = ?
    """, (num_orden,))
    total_orden = Decimal(cursor.fetchone()[0] or 0)

    cursor.execute("""
        SELECT 
            AbonoEfectivoTransferencia1, AbonoEfectivoTransferencia2, AbonoEfectivoTransferencia3,
            ValorPagadoTarjeta2, ValorPagadoTarjeta3
        FROM CuentasPorCobrar
        WHERE NumOrdenCompra = ?
    """, (num_orden,))
    registros = cursor.fetchall()

    pagos_db = sum(Decimal(p or 0) for fila in registros for p in fila)
    pagos_sesion = sum(Decimal(p["valor"]) for p in pagos_cxc)
    total_pagado = pagos_db + pagos_sesion
    restante = total_orden - total_pagado

    if restante <= 0:
        messagebox.showwarning("Límite Alcanzado", f"Ya se ha alcanzado el total por cobrar (${total_orden:,.2f}) para esta orden.")
        return

    def redondear(valor):
        return Decimal(valor).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    if tipo_pago == "Efectivo / Transferencia":
        cursor.execute("""
            SELECT AbonoEfectivoTransferencia1, AbonoEfectivoTransferencia2, AbonoEfectivoTransferencia3
            FROM CuentasPorCobrar
            WHERE NumOrdenCompra = ?
        """, (num_orden,))
        registros_abonos = cursor.fetchall()

        abonos_existentes = sum(1 for fila in registros_abonos for pago in fila if pago is not None)
        abonos_en_sesion = sum(1 for p in pagos_cxc if p["tipo"] == "Efectivo / Transferencia")
        total_abonos = abonos_existentes + abonos_en_sesion

        if total_abonos >= 3:
            messagebox.showwarning("Límite Alcanzado", "Solo se permiten hasta 3 abonos en Efectivo o Transferencia por orden.")
            return

        ventana_pago = tk.Toplevel()
        ventana_pago.title("Agregar Pago en Efectivo / Transferencia")

        tk.Label(ventana_pago, text="💵 Valor Abono ($):").grid(row=0, column=0, padx=10, pady=10)
        entry_valor = ttk.Entry(ventana_pago)
        entry_valor.grid(row=0, column=1, padx=10, pady=10)
        entry_valor.insert(0, str(redondear(restante)))

        tk.Label(ventana_pago, text="📅 Fecha:").grid(row=1, column=0, padx=10, pady=10)
        entry_fecha = DateEntry(ventana_pago, width=20, background='darkblue', foreground='white', date_pattern='yyyy-mm-dd')
        entry_fecha.grid(row=1, column=1, padx=10, pady=10)

        def agregar_pago():
            try:
                valor = Decimal(entry_valor.get().strip())
                valor = redondear(valor)
                if valor <= 0:
                    raise ValueError("El valor debe ser mayor a 0.")

                if valor > restante:
                    messagebox.showwarning("Exceso de pago", f"Este abono excede el total restante por cobrar (${restante:,.2f}).")
                    return

                fecha = entry_fecha.get_date().strftime('%Y-%m-%d')
                global contador_global_pagos
                pagos_cxc.append({
                    "tipo": "Efectivo / Transferencia",
                    "valor": valor,
                    "fecha": fecha,
                    "orden": contador_global_pagos
                })
                contador_global_pagos += 1
                reconstruir_text_area()


                actualizar_encabezado_saldo()
                
                ventana_pago.destroy()

            except Exception as e:
                messagebox.showerror("Error", f"Ingrese un valor numérico válido. Detalle: {e}")

        # Usa tk.Button para aplicar bg
        tk.Button(ventana_pago, text="Agregar pago a orden", command=agregar_pago, bg="lightgreen").grid(row=2, column=0, columnspan=2, pady=10)
        ventana_pago.grab_set()


    # === TARJETA DE CRÉDITO ===
    elif tipo_pago == "Tarjeta de Crédito":
        cursor.execute("""
            SELECT ValorPagadoTarjeta2, ValorPagadoTarjeta3
            FROM CuentasPorCobrar
            WHERE NumOrdenCompra = ?
        """, (num_orden,))
        registros_tarjeta = cursor.fetchall()

        tarjetas_existentes = sum(1 for fila in registros_tarjeta for pago in fila if pago is not None)
        tarjetas_sesion = sum(1 for p in pagos_cxc if p["tipo"] == "Tarjeta de Crédito")
        total_tarjetas = tarjetas_existentes + tarjetas_sesion

        if total_tarjetas >= 2:
            messagebox.showwarning("Límite Alcanzado", "Solo se permiten hasta 2 pagos con Tarjeta de Crédito por orden.")
            return

        ventana_tarjeta = tk.Toplevel()
        ventana_tarjeta.title("Agregar Pago con Tarjeta de Crédito")

        tk.Label(ventana_tarjeta, text="🏦 Tipo de Tarjeta:").grid(row=0, column=0, padx=10, pady=5, sticky="w")
        combo_tipo_tarjeta = ttk.Combobox(ventana_tarjeta, values=["Corriente", "Diferido"], state="readonly", width=30)
        combo_tipo_tarjeta.grid(row=0, column=1, padx=10, pady=5)

        tk.Label(ventana_tarjeta, text="🏦 Banco:").grid(row=1, column=0, padx=10, pady=5, sticky="w")
        combo_banco = ttk.Combobox(ventana_tarjeta, state="readonly", width=50)
        combo_banco.grid(row=1, column=1, padx=10, pady=5)

        def actualizar_bancos_credito(event=None):
            tipo = combo_tipo_tarjeta.get()
            bancos = opciones_banco.get(tipo, [])
            combo_banco["values"] = bancos
            if bancos:
                combo_banco.current(0)

        def verificar_banco_personalizado(event=None):
            banco = combo_banco.get()
            if "Ingresar Porcentaje Personalizado" in banco:
                ventana_porcentaje = tk.Toplevel()
                ventana_porcentaje.title("Porcentaje Personalizado")
                ventana_porcentaje.geometry("300x130")

                tk.Label(ventana_porcentaje, text="Ingrese el porcentaje (%):", font=("Arial", 10)).pack(pady=10)
                entry_porcentaje = ttk.Entry(ventana_porcentaje)
                entry_porcentaje.pack(pady=5)

                def aceptar_porcentaje():
                    try:
                        valor = float(entry_porcentaje.get().strip())
                        if valor <= 0 or valor >= 100:
                            raise ValueError()
                        porcentaje_decimal = Decimal(valor) / 100

                        # Nombre único con clave visible
                        clave = f"Personalizado ({valor:.2f}%)"
                        porcentaje_comision_banco_dict[clave] = porcentaje_decimal

                        # Añadir la opción personalizada al combo y seleccionarla
                        nuevas_opciones = list(combo_banco["values"]) + [clave]
                        combo_banco["values"] = nuevas_opciones
                        combo_banco.set(clave)

                        ventana_porcentaje.destroy()
                    except:
                        messagebox.showerror("Error", "Ingrese un porcentaje válido entre 0 y 100.")

                ttk.Button(ventana_porcentaje, text="Aceptar", command=aceptar_porcentaje).pack(pady=10)
                ventana_porcentaje.grab_set()


        combo_tipo_tarjeta.bind("<<ComboboxSelected>>", actualizar_bancos_credito)
        combo_banco.bind("<<ComboboxSelected>>", verificar_banco_personalizado)

        tk.Label(ventana_tarjeta, text="💵 Valor Pagado:").grid(row=2, column=0, padx=10, pady=5, sticky="w")
        entry_valor_tarjeta = ttk.Entry(ventana_tarjeta)
        entry_valor_tarjeta.grid(row=2, column=1, padx=10, pady=5)
        entry_valor_tarjeta.insert(0, str(redondear(restante)))

        tk.Label(ventana_tarjeta, text="🔢 Lote:").grid(row=3, column=0, padx=10, pady=5, sticky="w")
        entry_lote = ttk.Entry(ventana_tarjeta)
        entry_lote.grid(row=3, column=1, padx=10, pady=5)

        tk.Label(ventana_tarjeta, text="📅 Fecha:").grid(row=4, column=0, padx=10, pady=5, sticky="w")
        entry_fecha_tarjeta = DateEntry(ventana_tarjeta, width=20, background='darkblue', foreground='white', date_pattern='yyyy-mm-dd')
        entry_fecha_tarjeta.grid(row=4, column=1, padx=10, pady=5)

        def agregar_pago_tarjeta():
            global pagos_tarjeta_credito, contador_global_pagos

            try:
                valor = Decimal(entry_valor_tarjeta.get().strip())
                valor = redondear(valor)

                if valor <= 0:
                    raise ValueError("El valor pagado debe ser mayor a 0.")
                if valor > restante:
                    messagebox.showwarning("Exceso de pago", f"Este pago excede el total restante por cobrar (${restante:,.2f}).")
                    return

                tipo_tarjeta = combo_tipo_tarjeta.get().strip()
                banco = combo_banco.get().strip()
                lote = entry_lote.get().strip()
                fecha = entry_fecha_tarjeta.get_date().strftime('%Y-%m-%d')

                if not tipo_tarjeta or not banco or not lote:
                    messagebox.showerror("Error", "Todos los campos son obligatorios.")
                    return

                if not lote.isdigit():
                    messagebox.showerror("Error", "El número de lote debe ser un número entero.")
                    return

                pagos_cxc.append({
                    "tipo": "Tarjeta de Crédito",
                    "modo": tipo_tarjeta,
                    "banco": banco,
                    "valor": valor,
                    "lote": lote,
                    "fecha": fecha,
                    "orden": contador_global_pagos
                })
                contador_global_pagos += 1

                pagos_tarjeta_credito += 1
                reconstruir_text_area()
                actualizar_encabezado_saldo()
                ventana_tarjeta.destroy()

            except Exception as e:
                messagebox.showerror("Error", f"Ingrese un valor numérico válido. Detalle: {e}")

        tk.Button(ventana_tarjeta, text="Agregar pago a orden", command=agregar_pago_tarjeta, bg="lightgreen").grid(row=5, column=0, columnspan=2, pady=10)
        ventana_tarjeta.grab_set()



def confirmar_pago_cliente():
    if not pagos_cxc:
        messagebox.showerror("Error", "Debe registrar al menos un pago antes de confirmar.")
        return

    ventana_confirmar = tk.Toplevel()
    ventana_confirmar.title("Confirmar Pago del Cliente")

    tk.Label(ventana_confirmar, text="¿Se emite Factura?").grid(row=0, column=0, padx=10, pady=5)
    combo_factura = ttk.Combobox(ventana_confirmar, values=["Sí", "No"], state="readonly", width=10)
    combo_factura.grid(row=0, column=1, padx=10, pady=5)

    label_num_factura = tk.Label(ventana_confirmar, text="Número de Factura:")
    entry_num_factura = ttk.Entry(ventana_confirmar)

    def mostrar_campo_factura(event=None):
        if combo_factura.get() == "Sí":
            label_num_factura.grid(row=1, column=0, padx=10, pady=5)
            entry_num_factura.grid(row=1, column=1, padx=10, pady=5)
        else:
            label_num_factura.grid_remove()
            entry_num_factura.grid_remove()

    combo_factura.bind("<<ComboboxSelected>>", mostrar_campo_factura)

    def guardar_en_bd():
        num_orden = entry_orden_cxc.get().strip()
        if not num_orden:
            messagebox.showerror("Error", "Debe ingresar un número de Orden de Compra válido.")
            return

        factura = combo_factura.get()
        num_factura = entry_num_factura.get().strip() if factura == "Sí" else None
        if factura == "Sí":
            if not num_factura.isdigit():
                messagebox.showerror("Error", "El número de factura debe ser un número entero.")
                return
            num_factura = int(num_factura)  # ✅ lo convertimos a entero para insertar correctamente

        # Inicializar campos
        campos = {
            "AbonoEfectivoTransferencia1": None, "FechaPagadoEfectivo1": None,
            "AbonoEfectivoTransferencia2": None, "FechaPagadoEfectivo2": None,
            "AbonoEfectivoTransferencia3": None, "FechaPagadoEfectivo3": None,
            "TipoPago2": None, "Banco2": None, "ValorPagadoTarjeta2": None,
            "Lote2": None, "FechaPagado2": None,
            "TipoPago3": None, "Banco3": None, "ValorPagadoTarjeta3": None,
            "Lote3": None, "FechaPagado3": None,
            "Factura": factura,
            "NumeroFactura": num_factura,
            "NumOrdenCompra": num_orden,
            "NombreCliente": None  # NUEVO
            
        }
        # Obtener el nombre del cliente desde OrdenCompra
        cursor.execute("SELECT TOP 1 NombreCliente FROM OrdenCompra WHERE NumOrdenCompra = ?", (num_orden,))
        cliente_row = cursor.fetchone()
        campos["NombreCliente"] = cliente_row[0] if cliente_row else "Desconocido"

        # Clasificar pagos
        ef_index = 1
        tarjeta_index = 2
        for pago in pagos_cxc:
            if pago["tipo"] == "Efectivo / Transferencia" and ef_index <= 3:
                campos[f"AbonoEfectivoTransferencia{ef_index}"] = pago["valor"]
                campos[f"FechaPagadoEfectivo{ef_index}"] = pago["fecha"]
                ef_index += 1
            elif pago["tipo"] == "Tarjeta de Crédito" and tarjeta_index <= 3:
                campos[f"TipoPago{tarjeta_index}"] = pago["modo"]
                campos[f"Banco{tarjeta_index}"] = pago["banco"]
                campos[f"ValorPagadoTarjeta{tarjeta_index}"] = pago["valor"]
                campos[f"Lote{tarjeta_index}"] = pago["lote"]
                campos[f"FechaPagado{tarjeta_index}"] = pago["fecha"]
                tarjeta_index += 1

        if abonos_efectivo_transferencia > 0:
            campos["TipoPagoEfecTrans"] = "Efectivo/Transferencia"

        # ========= CALCULOS FINANCIEROS =========
        cursor.execute("""
            SELECT MAX(PorcentajeIVA), SUM(BaseRetencion), SUM(ValorBaseRetencion), 
                SUM(ValorXCobrarConIVA), SUM(oc.CantidadVendida * ISNULL(p.CostoConIVA, 0)), 
                SUM(ComisionPorPagarConsultor), SUM(ComisionPorPagarPadreEmpresarial)
            FROM OrdenCompra oc
            LEFT JOIN Productos p ON oc.CodigoProducto = p.CodigoProducto
            WHERE oc.NumOrdenCompra = ?
        """, (num_orden,))
        datos = cursor.fetchone()

        if not datos:
            messagebox.showerror("Error", "No se encontraron datos financieros para la orden.")
            return

        iva = Decimal(datos[0])
        base_retencion = Decimal(datos[1])
        valor_base_retencion = Decimal(datos[2])
        valor_x_cobrar_con_iva = Decimal(datos[3])
        costo_con_iva = Decimal(datos[4])
        comision_consultor = Decimal(datos[5] or 0)
        comision_padre = Decimal(datos[6] or 0)

        # Abonos y tarjetas
        ab1 = Decimal(campos["AbonoEfectivoTransferencia1"] or 0)
        ab2 = Decimal(campos["AbonoEfectivoTransferencia2"] or 0)
        ab3 = Decimal(campos["AbonoEfectivoTransferencia3"] or 0)
        campos["TotalEfectivo"] = ab1 + ab2 + ab3
        tc2 = Decimal(campos["ValorPagadoTarjeta2"] or 0)
        tc3 = Decimal(campos["ValorPagadoTarjeta3"] or 0)

        # IVA Pago Efectivo
        campos["IVAPagoEfectivoFactura"] = ((ab1 + ab2 + ab3) - (ab1 + ab2 + ab3) / (1 + iva)) if factura == "Sí" else Decimal("0")

        for i, valor_tc in enumerate([tc2, tc3], start=2):
            banco = campos.get(f"Banco{i}")
            porcentaje_banco = porcentaje_comision_banco_dict.get(banco, Decimal("0"))
            comision = (base_retencion * porcentaje_banco) + (base_retencion * porcentaje_banco * iva)
            irf = (valor_tc / (1 + iva)) * Decimal("0.02") if valor_tc > 0 else Decimal("0")
            ret_iva = (base_retencion * iva) * Decimal("0.30") if valor_tc > 0 else Decimal("0")
            total_banco = comision + irf + ret_iva
            valor_neto = valor_tc - total_banco if valor_tc > 0 else Decimal("0")

            campos[f"PorcentajeComisionBanco{i}"] = porcentaje_banco * 100
            campos[f"ComisionTCFactura{i}"] = comision
            campos[f"PorcentajeIRF{i}"] = Decimal("2.0") if valor_tc > 0 else Decimal("0")
            campos[f"IRF{i}"] = irf
            campos[f"PorcentajeRetIVA{i}"] = Decimal("30.0") if valor_tc > 0 else Decimal("0")
            campos[f"RetIVAPagoTarjetaCredito{i}"] = ret_iva
            campos[f"TotalComisionBanco{i}"] = total_banco
            campos[f"ValorNetoTC{i}"] = valor_neto

        # Totales
        campos["ComisionBancoTotales"] = campos["TotalComisionBanco2"] + campos["TotalComisionBanco3"]
        campos["TotalesValorNetoTC"] = campos["ValorNetoTC2"] + campos["ValorNetoTC3"]
        campos["ValorXCobrarConIVATotal"] = valor_x_cobrar_con_iva
        campos["BaseRetencionTotal"] = base_retencion
        campos["SaldoXCobrarCliente"] = valor_x_cobrar_con_iva - ab1 - ab2 - ab3 - tc2 - tc3
        campos["CostoConIVA"] = costo_con_iva
        utilidad = (
            valor_x_cobrar_con_iva
            - campos["SaldoXCobrarCliente"]
            - costo_con_iva
            - campos["IVAPagoEfectivoFactura"]
            - comision_consultor
            - comision_padre
        )
        campos["UtilidadDescontadoIVASRI"] = utilidad
        campos["PorcentajeGanancia"] = (utilidad / costo_con_iva * 100) if costo_con_iva > 0 else Decimal("0")


        # ========== INSERTAR EN TABLA ==========
        columnas = ", ".join(campos.keys())
        placeholders = ", ".join(["?"] * len(campos))
        valores = [float(x) if isinstance(x, Decimal) else x for x in campos.values()]  

        cursor.execute(f"INSERT INTO CuentasPorCobrar ({columnas}) VALUES ({placeholders})", valores)
        conn.commit()
        messagebox.showinfo("Éxito", "Pago registrado correctamente en la base de datos.")

        # Limpiar text area
        text_area_cxc.config(state="normal")
        text_area_cxc.delete("1.0", tk.END)
        text_area_cxc.config(state="disabled")

        # Limpiar campos y deshabilitar controles
        pagos_cxc.clear()
        entry_orden_cxc.delete(0, tk.END)
        btn_verificar_orden.config(state="disabled")  # 🟢 ESTA ES NUEVA
        combo_tipo_pago.set("")
        combo_tipo_pago.config(state="disabled")
        btn_confirmar_pago.config(state="disabled")
        label_recordatorio_buscar.grid_remove()

        ventana_confirmar.destroy()





    tk.Button(ventana_confirmar, text="Confirmar", command=guardar_en_bd,bg="lightgreen").grid(row=2, column=0, columnspan=2, pady=10)


btn_confirmar_pago = tk.Button(frame_izquierdo_cxc, text="✅ Confirmar pago del cliente", command=confirmar_pago_cliente,bg="lightgreen")
btn_confirmar_pago.grid(row=2, column=0, columnspan=2, pady=10)


def abrir_ventana_porcentaje_personalizado(combo_banco):
    ventana = tk.Toplevel()
    ventana.title("Ingresar Porcentaje Personalizado")

    tk.Label(ventana, text="Ingrese el porcentaje de comisión (%):").pack(pady=10)
    entry_porcentaje = ttk.Entry(ventana)
    entry_porcentaje.pack(pady=5)

    def aceptar():
        try:
            valor = float(entry_porcentaje.get().strip())
            if valor <= 0 or valor >= 100:
                raise ValueError("Debe estar entre 0 y 100.")
            nombre_opcion = f"Personalizado ({valor:.2f}%)"
            porcentaje_comision_banco_dict[nombre_opcion] = Decimal(valor / 100)
            valores_actuales = list(combo_banco["values"])
            if nombre_opcion not in valores_actuales:
                combo_banco["values"] = valores_actuales + [nombre_opcion]
            combo_banco.set(nombre_opcion)
            ventana.destroy()
        except:
            messagebox.showerror("Error", "Ingrese un porcentaje válido (entre 0 y 100).")

    ttk.Button(ventana, text="Aceptar", command=aceptar).pack(pady=10)
    ventana.grab_set()


def cargar_datos_cuentaporcobrar():
    
    id_cuenta = entry_id_cuenta.get().strip()
    if not id_cuenta.isdigit():
        messagebox.showerror("Error", "Ingrese un IdCuenta válido (número entero).")
        return

    cursor.execute("SELECT * FROM CuentasPorCobrar WHERE IdCuenta = ?", (int(id_cuenta),))
    cuenta = cursor.fetchone()

    if not cuenta:
        messagebox.showerror("No encontrado", "No se encontró una cuenta con ese ID.")
        return

    entry_orden_cxc.delete(0, tk.END)
    entry_orden_cxc.insert(0, str(cuenta[1]))

    # Mostramos la nueva ventana de edición
    mostrar_ventana_edicion_completa(cuenta)

def mostrar_ventana_edicion_completa(cuenta):
    
    from datetime import datetime
    ventana = tk.Toplevel()
    ventana.title(f"✏️ Editar Cuenta por Cobrar - ID {cuenta[0]}")
    ventana.geometry("950x700")
    # Validación de solo números para abonos y lotes
    def solo_numeros(valor):
        return valor.replace(".", "", 1).isdigit() or valor == ""
    
    vcmd = ventana.register(solo_numeros)

    entries_abonos = []
    entries_tarjetas = []

    # ==== SALDO DINÁMICO CORREGIDO ====
    valor_total_con_iva = Decimal(cuenta[42] or 0)  # Este es el total real a cobrar
    nombre_cliente = cuenta[2]  # Este índice 2 corresponde a NombreCliente según tu estructura

    label_nombre = ttk.Label(
        ventana,
        text=f"👤 Nombre Cliente: {nombre_cliente}",
        font=("Arial", 12, "bold")
    )
    label_nombre.pack(pady=(10, 0))  # Un poco de margen superior, sin margen inferior

    label_saldo = ttk.Label(ventana, text=f"💰 Saldo por Cobrar Cliente: ${cuenta[44]:,.2f}", font=("Arial", 12, "bold"))
    label_saldo.pack(pady=10)

    def actualizar_saldo_dinamico():
        try:
            ab1 = Decimal(entries_abonos[0][0].get() or 0)
            ab2 = Decimal(entries_abonos[1][0].get() or 0)
            ab3 = Decimal(entries_abonos[2][0].get() or 0)
            tc2 = Decimal(entries_tarjetas[0][2].get() or 0)
            tc3 = Decimal(entries_tarjetas[1][2].get() or 0)

            total_pagado = ab1 + ab2 + ab3 + tc2 + tc3
            nuevo_saldo = valor_total_con_iva - total_pagado

            label_saldo.config(text=f"💰 Saldo por Cobrar Cliente: ${nuevo_saldo:,.2f}")
        except:
            pass  # Silencia errores momentáneos de entrada


    # ===== Abonos Efectivo / Transferencia =====
    frame_abonos = ttk.LabelFrame(ventana, text="🟢 Abonos Efectivo / Transferencia")
    frame_abonos.pack(fill="x", padx=10, pady=10)

    for i in range(1, 4):
        ttk.Label(frame_abonos, text=f"Abono #{i}:").grid(row=i-1, column=0, padx=5, pady=5)
        entry_valor = ttk.Entry(frame_abonos, width=15, validate="key", validatecommand=(vcmd, "%P"))
        entry_valor.grid(row=i-1, column=1, padx=5)

        entry_fecha = DateEntry(frame_abonos, width=15, date_pattern='yyyy-mm-dd')
        entry_fecha.grid(row=i-1, column=2, padx=5)

        valor = cuenta[3 + (i - 1) * 2 + 1]  
        fecha = cuenta[3 + (i - 1) * 2 + 2]  

        if valor is not None:
            entry_valor.insert(0, str(valor))
        if isinstance(fecha, (datetime, date)):
            entry_fecha.set_date(fecha)

        entry_valor.bind("<KeyRelease>", lambda e, ev=entry_valor: actualizar_saldo_dinamico())
        entries_abonos.append((entry_valor, entry_fecha))

    # ===== Pagos con Tarjeta de Crédito =====
    frame_tarjetas = ttk.LabelFrame(ventana, text="💳 Pagos con Tarjeta de Crédito")
    frame_tarjetas.pack(fill="x", padx=10, pady=10)

    indices_tarjetas = [
        (14, "Tarjeta #1"),
        (27, "Tarjeta #2")
    ]

    for i, (idx, label_texto) in enumerate(indices_tarjetas):
        fila = i * 3
        ttk.Label(frame_tarjetas, text=label_texto).grid(row=fila, column=0, columnspan=6, sticky="w")

        # Tipo
        ttk.Label(frame_tarjetas, text="Tipo:").grid(row=fila+1, column=0, padx=5)
        combo_tipo = ttk.Combobox(frame_tarjetas, values=["Corriente", "Diferido"], width=15, state="readonly")
        combo_tipo.grid(row=fila+1, column=1)

        # Banco
        ttk.Label(frame_tarjetas, text="Banco:").grid(row=fila+1, column=2, padx=5)
        combo_banco = ttk.Combobox(frame_tarjetas, width=40, state="disabled")
        combo_banco.grid(row=fila+1, column=3, columnspan=2)

        # Valor
        ttk.Label(frame_tarjetas, text="Valor:").grid(row=fila+2, column=0, padx=5)
        entry_valor = ttk.Entry(frame_tarjetas, width=12, state="disabled")
        entry_valor.grid(row=fila+2, column=1)

        # Lote
        ttk.Label(frame_tarjetas, text="Lote:").grid(row=fila+2, column=2, padx=5)
        entry_lote = ttk.Entry(frame_tarjetas, width=15, state="disabled")
        entry_lote.grid(row=fila+2, column=3)

        # Fecha
        ttk.Label(frame_tarjetas, text="Fecha:").grid(row=fila+2, column=4, padx=5)
        entry_fecha = DateEntry(frame_tarjetas, width=15, date_pattern='yyyy-mm-dd', state="disabled")
        entry_fecha.grid(row=fila+2, column=5)

        # ==== Lógica de habilitación dinámica ====
        def habilitar_campos_valores(entry_valor, entry_lote, entry_fecha):
            entry_valor.config(state="normal")
            entry_lote.config(state="normal")
            entry_fecha.config(state="normal")

        def crear_callback_banco(cb_banco, entry_valor, entry_lote, entry_fecha):
            def callback(_event=None):
                banco = cb_banco.get()
                habilitar_campos_valores(entry_valor, entry_lote, entry_fecha)
                if banco == "Ingresar Porcentaje Personalizado...":
                    abrir_ventana_porcentaje_personalizado(cb_banco)
            return callback

        def habilitar_banco(e, cb_tipo, cb_banco, entry_valor, entry_lote, entry_fecha):
            tipo = cb_tipo.get()
            opciones = opciones_banco.get(tipo, [])
            cb_banco["values"] = opciones
            cb_banco.set("")
            cb_banco.config(state="readonly")
            entry_valor.config(state="disabled")
            entry_lote.config(state="disabled")
            entry_fecha.config(state="disabled")
            cb_banco.bind("<<ComboboxSelected>>", crear_callback_banco(cb_banco, entry_valor, entry_lote, entry_fecha))

        combo_tipo.bind("<<ComboboxSelected>>", lambda e, ct=combo_tipo, cb=combo_banco, ev=entry_valor, el=entry_lote, ef=entry_fecha:
                        habilitar_banco(e, ct, cb, ev, el, ef))

        

        entry_valor.bind("<KeyRelease>", lambda e: actualizar_saldo_dinamico())
        entry_lote.bind("<KeyRelease>", lambda e: None)  # puedes validar también aquí

        # ==== Insertar datos existentes si hay ====
        tipo = cuenta[idx]
        valor = cuenta[idx + 1]
        banco = cuenta[idx + 2]
        lote = cuenta[idx + 3]
        fecha = cuenta[idx + 4]

        if tipo:
            combo_tipo.set(tipo)
            combo_banco["values"] = opciones_banco.get(tipo, [])
            combo_banco.config(state="readonly")

        if banco:
            combo_banco.set(banco)
            entry_valor.config(state="normal")
            entry_lote.config(state="normal")
            entry_fecha.config(state="normal")

        if valor:
            entry_valor.insert(0, str(valor))
        if lote:
            entry_lote.insert(0, str(lote))
        from datetime import datetime

        if fecha:
            try:
                if isinstance(fecha, str):
                    fecha = datetime.strptime(fecha, "%Y-%m-%d").date()
                entry_fecha.set_date(fecha)
            except Exception as e:
                print(f"⚠️ Error al asignar fecha tarjeta #{i+1}: {e}")


        # ✅ Importante para actualizar el saldo
        entry_valor.bind("<KeyRelease>", lambda e, ev=entry_valor: actualizar_saldo_dinamico())

        entries_tarjetas.append((combo_tipo, combo_banco, entry_valor, entry_lote, entry_fecha))



    # ===== Facturación =====
    frame_factura = ttk.LabelFrame(ventana, text="📄 Facturación")
    frame_factura.pack(pady=5)

    ttk.Label(frame_factura, text="¿Factura emitida?:").grid(row=0, column=0, padx=5)
    combo_factura = ttk.Combobox(frame_factura, values=["Sí", "No"], width=10, state="readonly")
    combo_factura.grid(row=0, column=1, padx=5)
    combo_factura.set(cuenta[11] or "No")

    ttk.Label(frame_factura, text="Número de Factura:").grid(row=0, column=2, padx=5)

    # === Validación solo números ===
    def solo_numeros(valor):
        return valor.replace(".", "").isdigit() or valor == ""


    vcmd = ventana.register(solo_numeros)
    entry_factura = ttk.Entry(frame_factura, width=15, validate="key", validatecommand=(vcmd, "%P"))
    entry_factura.grid(row=0, column=3, padx=5)

    # === Habilitar o deshabilitar según factura ===
    def mostrar_campo_factura_edicion(event=None):
        if combo_factura.get() == "Sí":
            entry_factura.config(state="normal")
        else:
            entry_factura.delete(0, tk.END)
            entry_factura.config(state="disabled")

    combo_factura.bind("<<ComboboxSelected>>", mostrar_campo_factura_edicion)
    mostrar_campo_factura_edicion()  # Estado inicial

    # Cargar valor si ya existía
    if cuenta[12]:
        entry_factura.config(state="normal")
        entry_factura.insert(0, cuenta[12])



    # ===== Guardar Cambios =====
    btn_guardar = tk.Button(
        ventana,
        text="💾 Guardar Cambios",
        bg="lightgreen",
        command=lambda: guardar_cambios_desde_ventana(
            id_cuenta=cuenta[0],
            entries_abonos=entries_abonos,
            entries_tarjetas=entries_tarjetas,
            combo_factura=combo_factura,
            entry_factura=entry_factura,
            ventana=ventana,
            num_orden=cuenta[1]  # ← nuevo
        )
    )
    btn_guardar.pack(pady=20)

def guardar_cambios_desde_ventana(id_cuenta, entries_abonos, entries_tarjetas, combo_factura, entry_factura, ventana, num_orden):

    
    try:
        for i, (combo_tipo, combo_banco, entry_valor, entry_lote, entry_fecha) in enumerate(entries_tarjetas, start=2):
            tipo = combo_tipo.get().strip()
            banco = combo_banco.get().strip()
            valor = entry_valor.get().strip()
            lote = entry_lote.get().strip()

            if tipo and banco and (not valor or not lote):
                messagebox.showerror("Error", f"Debe completar los campos Valor y Lote para Tarjeta #{i-3}.")
                return

        campos = {
            "Factura": combo_factura.get(),
            "NumeroFactura": int(entry_factura.get().strip()) if combo_factura.get() == "Sí" and entry_factura.get().strip().isdigit() else None
        }

        # ===== ABONOS EFECTIVO / TRANSFERENCIA =====
        for i, (entry_valor, entry_fecha) in enumerate(entries_abonos, start=1):
            valor = entry_valor.get().strip()
            fecha = entry_fecha.get().strip()
            try:
                valor_float = float(valor.replace(",", "").strip()) if valor.strip() else 0
            except:
                valor_float = 0
            campos[f"AbonoEfectivoTransferencia{i}"] = valor_float
            campos[f"FechaPagadoEfectivo{i}"] = fecha if fecha else None

        # ===== TARJETAS DE CRÉDITO =====
        # Validar que si se ingresó tipo y banco, los demás campos no estén vacíos
        for i, (combo_tipo, combo_banco, entry_valor, entry_lote, entry_fecha) in enumerate(entries_tarjetas, start=2):
            tipo = combo_tipo.get().strip()
            banco = combo_banco.get().strip()
            valor = entry_valor.get().strip()
            lote = entry_lote.get().strip()
            fecha = entry_fecha.get().strip()

            if tipo and banco:  # Si hay intento de registrar tarjeta
                if not valor or not lote or not fecha:
                    messagebox.showerror("Campos incompletos", f"Debe completar todos los campos en la Tarjeta #{i - 3}.")
                    return

        for i, (combo_tipo, combo_banco, entry_valor, entry_lote, entry_fecha) in enumerate(entries_tarjetas, start=2):
            if i > 3:
                break  # Solo se permite Tarjeta #1 (2) y Tarjeta #2 (3)
            campos[f"TipoPago{i}"] = combo_tipo.get().strip()
            campos[f"Banco{i}"] = combo_banco.get().strip()
            try:
                campos[f"ValorPagadoTarjeta{i}"] = float(entry_valor.get().strip().replace(",", "")) if entry_valor.get().strip() else 0
            except:
                campos[f"ValorPagadoTarjeta{i}"] = 0
            campos[f"Lote{i}"] = entry_lote.get().strip()
            campos[f"FechaPagado{i}"] = entry_fecha.get().strip()

        campos["TipoPagoEfecTrans"] = "Efectivo/Transferencia" if any(campos[f"AbonoEfectivoTransferencia{i}"] > 0 for i in range(1, 4)) else None

        # ===== INFO ADICIONAL ORDEN COMPRA =====
        campos["NumOrdenCompra"] = num_orden

        cursor.execute("SELECT TOP 1 NombreCliente FROM OrdenCompra WHERE NumOrdenCompra = ?", (num_orden,))
        cliente_row = cursor.fetchone()
        campos["NombreCliente"] = cliente_row[0] if cliente_row else "Desconocido"

        cursor.execute("""
            SELECT MAX(PorcentajeIVA), SUM(BaseRetencion), SUM(ValorBaseRetencion), 
                   SUM(ValorXCobrarConIVA), SUM(oc.CantidadVendida * ISNULL(p.CostoConIVA, 0)), 
                   SUM(ComisionPorPagarConsultor), SUM(ComisionPorPagarPadreEmpresarial)
            FROM OrdenCompra oc
            LEFT JOIN Productos p ON oc.CodigoProducto = p.CodigoProducto
            WHERE oc.NumOrdenCompra = ?
        """, (num_orden,))
        datos = cursor.fetchone()

        iva = Decimal(datos[0])
        base_retencion = Decimal(datos[1])
        valor_base_retencion = Decimal(datos[2])
        valor_x_cobrar_con_iva = Decimal(datos[3])
        costo_con_iva = Decimal(datos[4])
        comision_consultor = Decimal(datos[5] or 0)
        comision_padre = Decimal(datos[6] or 0)

        ab1 = Decimal(campos["AbonoEfectivoTransferencia1"] or 0)
        ab2 = Decimal(campos["AbonoEfectivoTransferencia2"] or 0)
        ab3 = Decimal(campos["AbonoEfectivoTransferencia3"] or 0)
        tc2 = Decimal(campos["ValorPagadoTarjeta2"] or 0)
        tc3 = Decimal(campos["ValorPagadoTarjeta3"] or 0)

        total_efectivo = ab1 + ab2 + ab3
        campos["TotalEfectivo"] = total_efectivo
        campos["IVAPagoEfectivoFactura"] = (total_efectivo - total_efectivo / (1 + iva)) if campos["Factura"] == "Sí" else Decimal("0")

        for i, valor_tc in enumerate([tc2, tc3], start=2):
            banco = campos.get(f"Banco{i}")
            porcentaje_banco = porcentaje_comision_banco_dict.get(banco, Decimal("0"))
            comision = (base_retencion * porcentaje_banco) + (base_retencion * porcentaje_banco * iva)
            irf = (valor_tc / (1 + iva)) * Decimal("0.02") if valor_tc > 0 else Decimal("0")
            ret_iva = (base_retencion * iva) * Decimal("0.30") if valor_tc > 0 else Decimal("0")
            total_banco = comision + irf + ret_iva
            valor_neto = valor_tc - total_banco if valor_tc > 0 else Decimal("0")

            campos[f"PorcentajeComisionBanco{i}"] = porcentaje_banco * 100
            campos[f"ComisionTCFactura{i}"] = comision
            campos[f"PorcentajeIRF{i}"] = Decimal("2.0") if valor_tc > 0 else Decimal("0")
            campos[f"IRF{i}"] = irf
            campos[f"PorcentajeRetIVA{i}"] = Decimal("30.0") if valor_tc > 0 else Decimal("0")
            campos[f"RetIVAPagoTarjetaCredito{i}"] = ret_iva
            campos[f"TotalComisionBanco{i}"] = total_banco
            campos[f"ValorNetoTC{i}"] = valor_neto

        campos["ComisionBancoTotales"] = campos["TotalComisionBanco2"] + campos["TotalComisionBanco3"]
        campos["TotalesValorNetoTC"] = campos["ValorNetoTC2"] + campos["ValorNetoTC3"]
        campos["ValorXCobrarConIVATotal"] = valor_x_cobrar_con_iva
        campos["BaseRetencionTotal"] = base_retencion
        campos["SaldoXCobrarCliente"] = valor_x_cobrar_con_iva - ab1 - ab2 - ab3 - tc2 - tc3
        campos["CostoConIVA"] = costo_con_iva

        utilidad = (
            valor_x_cobrar_con_iva
            - campos["SaldoXCobrarCliente"]
            - costo_con_iva
            - campos["IVAPagoEfectivoFactura"]
            - comision_consultor
            - comision_padre
        )
        campos["UtilidadDescontadoIVASRI"] = utilidad
        campos["PorcentajeGanancia"] = (utilidad / costo_con_iva * 100) if costo_con_iva > 0 else Decimal("0")

        # ===== ACTUALIZAR EN BASE DE DATOS =====
        columnas = ", ".join([f"{col} = ?" for col in campos])
        valores = []
        for col, val in campos.items():
            try:
                if val is None or val == "":
                    valores.append(None)
                elif isinstance(val, Decimal):
                    valores.append(float(val))
                elif isinstance(val, (int, float)):
                    valores.append(val)
                elif isinstance(val, str):
                    if col.lower().startswith(("abono", "valor", "comision", "porcentaje", "ret", "saldo", "costo", "total", "utilidad")):
                        valores.append(float(val.strip().replace(",", "")))
                    else:
                        valores.append(val.strip())
                else:
                    valores.append(val)
            except Exception as e:
                print(f"⚠️ Error en columna {col} con valor '{val}': {e}")
                valores.append(None)

        valores.append(id_cuenta)

        cursor.execute(f"UPDATE CuentasPorCobrar SET {columnas} WHERE IdCuenta = ?", valores)
        conn.commit()

        messagebox.showinfo("✅ Éxito", f"Cuenta #{id_cuenta} modificada correctamente.")
        ventana.destroy()

    except Exception as e:
        messagebox.showerror("❌ Error", f"No se pudo guardar la cuenta:\n{str(e)}")




btn_cargar_cuenta = tk.Button(
    frame_idcuenta_seccion,
    text="🟡 Cargar Cuenta por Cobrar",
    bg="lightgreen",
    command=cargar_datos_cuentaporcobrar
)
btn_cargar_cuenta.pack(side="left", padx=10)

def eliminar_cuenta_por_cobrar():
    id_cuenta = entry_id_cuenta.get().strip()
    if not id_cuenta.isdigit():
        messagebox.showerror("Error", "Ingrese un IdCuenta válido (número entero).")
        return

    respuesta = messagebox.askyesno("Confirmar eliminación", f"¿Está seguro que desea eliminar la cuenta con ID {id_cuenta}? Esta acción no se puede deshacer.")
    if not respuesta:
        return

    cursor.execute("SELECT * FROM CuentasPorCobrar WHERE IdCuenta = ?", (int(id_cuenta),))
    cuenta = cursor.fetchone()
    if not cuenta:
        messagebox.showerror("No encontrado", f"No se encontró ninguna cuenta con ID {id_cuenta}.")
        return

    try:
        cursor.execute("DELETE FROM CuentasPorCobrar WHERE IdCuenta = ?", (int(id_cuenta),))
        conn.commit()
        messagebox.showinfo("Eliminado", f"Cuenta con ID {id_cuenta} eliminada correctamente.")
        entry_id_cuenta.delete(0, tk.END)
        entry_orden_cxc.delete(0, tk.END)
        text_area_cxc.delete("1.0", tk.END)
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo eliminar la cuenta:\n{e}")


btn_eliminar_cuenta = tk.Button(
    frame_idcuenta_seccion,
    text="🗑️ Eliminar Cuenta por Cobrar",
    bg="tomato",
    command=eliminar_cuenta_por_cobrar
)

btn_eliminar_cuenta.pack(side="left", padx=5)

def deshacer_ultimo_pago():
    global abonos_efectivo_transferencia, pagos_tarjeta_credito

    if not pagos_cxc:
        messagebox.showinfo("Nada que eliminar", "No hay pagos agregados aún.")
        return

    # Eliminar el último pago agregado
    ultimo_pago = pagos_cxc.pop()

    # Recalcular contadores
    abonos_efectivo_transferencia = sum(1 for p in pagos_cxc if p["tipo"] == "Efectivo / Transferencia")
    pagos_tarjeta_credito = sum(1 for p in pagos_cxc if p["tipo"] == "Tarjeta de Crédito")

    # Reconstruir visualmente el TextArea
    reconstruir_text_area()

    # Actualizar encabezado con saldo actualizado
    actualizar_encabezado_saldo()

    # Mostrar mensaje de confirmación
    messagebox.showinfo("Deshecho", "Se eliminó el último pago agregado.")






btn_deshacer_pago = tk.Button(
    frame_textarea_y_id,
    text="⏪ Deshacer Pago",
    command=deshacer_ultimo_pago,
    bg="tomato"
)
btn_deshacer_pago.pack(side="right", padx=10, pady=5)



# ---------------------- CAMPOS DIRECTAMENTE EN tab_cuentas_pagar ----------------------

# === Sección de Búsqueda (alineada y organizada) ===
ttk.Label(tab_cuentas_pagar, text="Número Orden Compra (Consultor):").grid(row=0, column=0, sticky="e", padx=(10, 2), pady=8)
frame_filtro_consultor = ttk.Frame(tab_cuentas_pagar)
frame_filtro_consultor.grid(row=0, column=1, padx=(2, 10), pady=8, sticky="w")
entry_buscar_consultor = ttk.Entry(frame_filtro_consultor, width=22)
entry_buscar_consultor.pack(side="left")
btn_buscar_consultor = tk.Button(frame_filtro_consultor, text="🔍", width=3,bg="lightgreen", command=lambda: buscar_datos_consultor(entry_buscar_consultor.get()))
btn_buscar_consultor.pack(side="left", padx=(5, 0))

# --- Búsqueda Padre Empresarial ---
ttk.Label(tab_cuentas_pagar, text="Número Orden Compra (Padre Empresarial):").grid(row=0, column=6, sticky="e", padx=(10, 2), pady=8)
frame_filtro_padre = ttk.Frame(tab_cuentas_pagar)
frame_filtro_padre.grid(row=0, column=7, padx=(2, 10), pady=8, sticky="w")
entry_buscar_padre = ttk.Entry(frame_filtro_padre, width=22)
entry_buscar_padre.pack(side="left")
btn_buscar_padre = tk.Button(frame_filtro_padre, text="🔍", width=3, bg="lightgreen",command=lambda: buscar_datos_padre(entry_buscar_padre.get()))
btn_buscar_padre.pack(side="left", padx=(5, 0))

# === Encabezados ===
ttk.Label(tab_cuentas_pagar, text="📋 Información del Consultor", font=("Arial", 10, "bold")).grid(row=1, column=0, columnspan=3, sticky="w", padx=10, pady=(10, 5))
ttk.Label(tab_cuentas_pagar, text="📋 Información del Padre Empresarial", font=("Arial", 10, "bold")).grid(row=1, column=6, columnspan=3, sticky="w", padx=100, pady=(10, 5))

# === Configuración estándar de ancho ===
ancho_campo = 22

# === Campos del Consultor ===
entry_nombre_consultor = ttk.Entry(tab_cuentas_pagar, state="readonly", width=ancho_campo)
entry_por_pagar_consultor = ttk.Entry(tab_cuentas_pagar, state="readonly", width=ancho_campo)
entry_valor_pago_consultor = ttk.Entry(tab_cuentas_pagar, state="disabled", width=ancho_campo)
entry_fecha_pago_consultor = DateEntry(tab_cuentas_pagar, width=ancho_campo, background='darkblue', foreground='white', date_pattern='yyyy-mm-dd', state="disabled")
entry_banco_consultor = ttk.Entry(tab_cuentas_pagar, state="disabled", width=ancho_campo)
entry_cuenta_consultor = ttk.Entry(tab_cuentas_pagar, state="disabled", width=ancho_campo)
entry_comprobante_consultor = ttk.Entry(tab_cuentas_pagar, state="disabled", width=ancho_campo)

campos_consultor = [
    ("Nombre Consultor:", entry_nombre_consultor),
    ("Por Pagar:", entry_por_pagar_consultor),
    ("Valor Pago:", entry_valor_pago_consultor),
    ("Fecha de Pago:", entry_fecha_pago_consultor),
    ("Banco:", entry_banco_consultor),
    ("Cuenta:", entry_cuenta_consultor),
    ("Comprobante #:", entry_comprobante_consultor),
]

for i, (label_text, widget) in enumerate(campos_consultor):
    ttk.Label(tab_cuentas_pagar, text=label_text).grid(row=i+2, column=0, sticky="e", padx=(10, 2), pady=3)
    widget.grid(row=i+2, column=1, padx=(2, 10), pady=3)


# === Campos del Padre Empresarial ===
entry_nombre_padre = ttk.Entry(tab_cuentas_pagar, state="readonly", width=ancho_campo)
entry_por_pagar_padre = ttk.Entry(tab_cuentas_pagar, state="readonly", width=ancho_campo)
entry_valor_pago_padre = ttk.Entry(tab_cuentas_pagar, state="disabled", width=ancho_campo)
entry_fecha_pago_padre = DateEntry(tab_cuentas_pagar, width=ancho_campo, background='darkblue', foreground='white', date_pattern='yyyy-mm-dd', state="disabled")
entry_banco_padre = ttk.Entry(tab_cuentas_pagar, state="disabled", width=ancho_campo)
entry_cuenta_padre = ttk.Entry(tab_cuentas_pagar, state="disabled", width=ancho_campo)
entry_comprobante_padre = ttk.Entry(tab_cuentas_pagar, state="readonly", width=ancho_campo)

campos_padre = [
    ("Nombre Padre Empresarial:", entry_nombre_padre),
    ("Por Pagar:", entry_por_pagar_padre),
    ("Valor Pago:", entry_valor_pago_padre),
    ("Fecha de Pago:", entry_fecha_pago_padre),
    ("Banco:", entry_banco_padre),
    ("Cuenta:", entry_cuenta_padre),
    ("Comprobante #:", entry_comprobante_padre),
]


for i, (label_text, widget) in enumerate(campos_padre):
    ttk.Label(tab_cuentas_pagar, text=label_text).grid(row=i+2, column=6, sticky="e", padx=(10, 2), pady=3)
    widget.grid(row=i+2, column=7, padx=(2, 10), pady=3)

def limpiar_campos_consultor():
    entry_nombre_consultor.config(state="normal")
    entry_nombre_consultor.delete(0, tk.END)
    entry_nombre_consultor.config(state="readonly")

    entry_por_pagar_consultor.config(state="normal")
    entry_por_pagar_consultor.delete(0, tk.END)
    entry_por_pagar_consultor.config(state="readonly")

    for widget in [entry_valor_pago_consultor, entry_banco_consultor, entry_cuenta_consultor, entry_comprobante_consultor]:
        widget.config(state="normal")
        widget.delete(0, tk.END)
        widget.config(state="disabled")

    entry_fecha_pago_consultor.config(state="normal")
    entry_fecha_pago_consultor.set_date(date.today())
    entry_fecha_pago_consultor.config(state="disabled")


# === Función para buscar consultor ===
def buscar_datos_consultor(num_orden):
    try:
        if not num_orden.isdigit():
            limpiar_campos_consultor()
            messagebox.showwarning("Advertencia", "Debe ingresar un número válido de orden de compra.")
            return

        num_orden = int(num_orden)

        cursor.execute("""
            SELECT SUM(ComisionPorPagarConsultor), MAX(NombreConsultor)
            FROM OrdenCompra
            WHERE NumOrdenCompra = ?
        """, (num_orden,))
        result = cursor.fetchone()
        total_por_pagar, nombre_consultor = result if result else (None, None)

        if total_por_pagar is None:
            limpiar_campos_consultor()
            messagebox.showerror("Error", f"No se encontró información para la orden #{num_orden}")
            return

        cursor.execute("""
            SELECT SUM(PagadoConsultor)
            FROM CuentasPorPagarConsultor
            WHERE NumOrdenCompra = ?
        """, (num_orden,))
        total_pagado = cursor.fetchone()[0] or 0

        restante = float(total_por_pagar) - float(total_pagado)

        # Mostrar nombre del consultor y valor por pagar
        entry_nombre_consultor.config(state="normal")
        entry_nombre_consultor.delete(0, tk.END)
        entry_nombre_consultor.insert(0, nombre_consultor)
        entry_nombre_consultor.config(state="readonly")

        entry_por_pagar_consultor.config(state="normal")
        entry_por_pagar_consultor.delete(0, tk.END)
        entry_por_pagar_consultor.insert(0, round(restante, 2))
        entry_por_pagar_consultor.config(state="readonly")

        for widget in [entry_valor_pago_consultor, entry_fecha_pago_consultor,
                       entry_banco_consultor, entry_cuenta_consultor, entry_comprobante_consultor]:
            widget.config(state="normal")

    except Exception as e:
        limpiar_campos_consultor()
        messagebox.showerror("Error", f"Error al buscar consultor: {e}")



def limpiar_campos_padre():
    entry_nombre_padre.config(state="normal")
    entry_nombre_padre.delete(0, tk.END)
    entry_nombre_padre.config(state="readonly")

    entry_por_pagar_padre.config(state="normal")
    entry_por_pagar_padre.delete(0, tk.END)
    entry_por_pagar_padre.config(state="readonly")

    for widget in [entry_valor_pago_padre, entry_banco_padre, entry_cuenta_padre, entry_comprobante_padre]:
        widget.config(state="normal")
        widget.delete(0, tk.END)
        widget.config(state="disabled")

    entry_fecha_pago_padre.config(state="normal")
    entry_fecha_pago_padre.set_date(date.today())
    entry_fecha_pago_padre.config(state="disabled")

# === Función para buscar padre empresarial ===
def buscar_datos_padre(num_orden):
    try:
        if not num_orden.isdigit():
            limpiar_campos_padre()
            messagebox.showwarning("Advertencia", "Debe ingresar un número válido de orden de compra.")
            return

        num_orden = int(num_orden)

        cursor.execute("""
            SELECT SUM(ComisionPorPagarPadreEmpresarial), MAX(NombrePadreEmpresarial)
            FROM OrdenCompra
            WHERE NumOrdenCompra = ?
        """, (num_orden,))
        result = cursor.fetchone()
        total_por_pagar, nombre_padre = result if result else (None, None)

        if total_por_pagar is None:
            limpiar_campos_padre()
            messagebox.showerror("Error", f"No se encontró información para la orden #{num_orden}")
            return

        cursor.execute("""
            SELECT SUM(PagadoPadreEmpresarial)
            FROM CuentasPorPagarPadreEmpresarial
            WHERE NumOrdenCompra = ?
        """, (num_orden,))
        total_pagado = cursor.fetchone()[0] or 0

        restante = float(total_por_pagar) - float(total_pagado)

        entry_nombre_padre.config(state="normal")
        entry_nombre_padre.delete(0, tk.END)
        entry_nombre_padre.insert(0, nombre_padre)
        entry_nombre_padre.config(state="readonly")

        entry_por_pagar_padre.config(state="normal")
        entry_por_pagar_padre.delete(0, tk.END)
        entry_por_pagar_padre.insert(0, round(restante, 2))
        entry_por_pagar_padre.config(state="readonly")

        for widget in [entry_valor_pago_padre, entry_fecha_pago_padre,
                       entry_banco_padre, entry_cuenta_padre, entry_comprobante_padre]:
            widget.config(state="normal")

    except Exception as e:
        limpiar_campos_padre()
        messagebox.showerror("Error", f"Ocurrió un error al buscar padre empresarial: {e}")




# === Confirmar Pago del Consultor ===
def confirmar_pago_consultor():
    try:
        num_orden = entry_buscar_consultor.get()
        if not num_orden.isdigit():
            messagebox.showwarning("Advertencia", "Debe ingresar un número de orden válido.")
            return
        num_orden = int(num_orden)

        nombre = entry_nombre_consultor.get()
        valor_pago = float(entry_valor_pago_consultor.get() or 0)
        banco = entry_banco_consultor.get()
        cuenta = entry_cuenta_consultor.get()
        fecha = entry_fecha_pago_consultor.get_date()
        comprobante = entry_comprobante_consultor.get()

        # === Validaciones ===
        if valor_pago <= 0:
            messagebox.showwarning("Advertencia", "Ingrese un valor válido para el pago.")
            return
        if not banco.isalpha():
            messagebox.showwarning("Advertencia", "El campo Banco solo debe contener letras.")
            return
        if not cuenta or not comprobante:
            messagebox.showwarning("Advertencia", "Debe llenar todos los campos.")
            return

        # === Cálculos ===
        cursor.execute("""
            SELECT SUM(ComisionPorPagarConsultor)
            FROM OrdenCompra
            WHERE NumOrdenCompra = ?
        """, (num_orden,))
        comision_total = float(cursor.fetchone()[0] or 0)

        cursor.execute("""
            SELECT SUM(PagadoConsultor)
            FROM CuentasPorPagarConsultor
            WHERE NumOrdenCompra = ?
        """, (num_orden,))
        total_pagado = float(cursor.fetchone()[0] or 0)

        from decimal import Decimal, ROUND_HALF_UP
        saldo = Decimal(str(comision_total)) - Decimal(str(total_pagado)) - Decimal(str(valor_pago))
        saldo = saldo.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

        if saldo < Decimal("0.00"):
            messagebox.showwarning("Advertencia", "El valor pagado excede el saldo por pagar.")
            return

        # === Insertar nuevo registro con NumOrdenCompra ===
        cursor.execute("""
            INSERT INTO CuentasPorPagarConsultor (
                NumOrdenCompra, NombreConsultor, ComisionPorPagarConsultorTotal, PagadoConsultor,
                BancoDistrConsultor, CuentaDistrConsultor, FechaPagoConsultor,
                NumComprobante, SaldoPorPagarConsultor
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            num_orden, nombre, comision_total, valor_pago, banco, cuenta, fecha, comprobante, float(saldo)
        ))
        conn.commit()
         # ✅ FALTABA ESTO
        recalcular_saldos_consultor(num_orden)

        messagebox.showinfo("Éxito", "Pago del consultor registrado correctamente.")
        recalcular_saldos_consultor(num_orden)
        buscar_datos_consultor(str(num_orden))

        # === Limpiar campos ===
        for widget in [entry_valor_pago_consultor, entry_banco_consultor, entry_cuenta_consultor,
                       entry_comprobante_consultor]:
            widget.delete(0, tk.END)
            widget.config(state="disabled")
        entry_fecha_pago_consultor.set_date(date.today())
        entry_fecha_pago_consultor.config(state="disabled")

        entry_nombre_consultor.config(state="normal")
        entry_nombre_consultor.delete(0, tk.END)
        entry_nombre_consultor.config(state="readonly")

        entry_por_pagar_consultor.config(state="normal")
        entry_por_pagar_consultor.delete(0, tk.END)
        entry_por_pagar_consultor.config(state="readonly")

    except Exception as e:
        messagebox.showerror("Error", f"No se pudo guardar el pago del consultor: {e}")


# === Confirmar Pago del Padre Empresarial ===
from decimal import Decimal, ROUND_HALF_UP

def confirmar_pago_padre():
    try:
        num_orden = entry_buscar_padre.get()
        if not num_orden.isdigit():
            messagebox.showwarning("Advertencia", "Debe ingresar un número de orden válido.")
            return
        num_orden = int(num_orden)

        nombre = entry_nombre_padre.get()
        valor_pago = Decimal(entry_valor_pago_padre.get() or "0").quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        banco = entry_banco_padre.get()
        cuenta = entry_cuenta_padre.get()
        fecha = entry_fecha_pago_padre.get_date()
        comprobante = entry_comprobante_padre.get()

        # === Validaciones ===
        if valor_pago <= Decimal("0.00"):
            messagebox.showwarning("Advertencia", "Ingrese un valor válido para el pago.")
            return
        if not banco.isalpha():
            messagebox.showwarning("Advertencia", "El campo Banco solo debe contener letras.")
            return
        if not cuenta or not comprobante:
            messagebox.showwarning("Advertencia", "Debe llenar todos los campos.")
            return

        # === Cálculos ===
        cursor.execute("""
            SELECT SUM(ComisionPorPagarPadreEmpresarial)
            FROM OrdenCompra
            WHERE NumOrdenCompra = ?
        """, (num_orden,))
        total_comision = Decimal(str(cursor.fetchone()[0] or 0)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

        cursor.execute("""
            SELECT SUM(PagadoPadreEmpresarial)
            FROM CuentasPorPagarPadreEmpresarial
            WHERE NumOrdenCompra = ?
        """, (num_orden,))
        total_pagado = Decimal(str(cursor.fetchone()[0] or 0)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

        saldo = (total_comision - total_pagado - valor_pago).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

        if saldo < Decimal("0.00"):
            messagebox.showwarning("Advertencia", "El valor pagado excede el saldo por pagar.")
            return

        # === Insertar nuevo registro ===
        cursor.execute("""
            INSERT INTO CuentasPorPagarPadreEmpresarial (
                NumOrdenCompra, NombrePadreEmpresarial, ComisionPorPagarPadreEmpresarialTotal,
                PagadoPadreEmpresarial, BancoDistrPadreEmpresarial, CuentaDistriPadreEmpresarial,
                FechaPagoPadreEmpresarial, NumComprobante, SaldoPorPagarPadreEmpresarial
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            num_orden, nombre, float(total_comision), float(valor_pago), banco, cuenta, fecha, comprobante, float(saldo)
        ))
        conn.commit()

        recalcular_saldos_padre(num_orden)

        messagebox.showinfo("Éxito", "Pago del padre empresarial registrado correctamente.")
        buscar_datos_padre(str(num_orden))

        # === Limpiar campos ===
        for widget in [entry_valor_pago_padre, entry_banco_padre, entry_cuenta_padre, entry_comprobante_padre]:
            widget.delete(0, tk.END)
            widget.config(state="disabled")

        entry_fecha_pago_padre.set_date(date.today())
        entry_fecha_pago_padre.config(state="disabled")

        entry_nombre_padre.config(state="normal")
        entry_nombre_padre.delete(0, tk.END)
        entry_nombre_padre.config(state="readonly")

        entry_por_pagar_padre.config(state="normal")
        entry_por_pagar_padre.delete(0, tk.END)
        entry_por_pagar_padre.config(state="readonly")

    except Exception as e:
        messagebox.showerror("Error", f"No se pudo guardar el pago del padre empresarial: {e}")





# === Botones Confirmar alineados debajo de sus secciones ===
btn_confirmar_pago_consultor = tk.Button(
    tab_cuentas_pagar,
    text="💼 Confirmar Pago del Consultor",
    command=confirmar_pago_consultor,
    bg="lightgreen"
)
btn_confirmar_pago_consultor.grid(row=9, column=1, columnspan=2, padx=(10, 10), pady=(10, 15), sticky="w")

btn_confirmar_pago_padre = tk.Button(
    tab_cuentas_pagar,
    text="🏢 Confirmar Pago del Padre Empresarial",
    command=confirmar_pago_padre,
    bg="lightgreen"
)
btn_confirmar_pago_padre.grid(row=9, column=7, columnspan=2, padx=(10, 10), pady=(10, 15), sticky="w")

# === Campo IdConsultor y botones de acción ===
ttk.Label(tab_cuentas_pagar, text="🔍 Id Cuentas Por Pagar Consultor:").grid(row=10, column=0, sticky="e", padx=(10, 2), pady=5)
entry_id_consultor = ttk.Entry(tab_cuentas_pagar, width=22)
entry_id_consultor.grid(row=10, column=1, padx=(2, 10), pady=5)



# === Función: Cargar datos desde Id ===
def cargar_datos_consultor():
    try:
        id_consultor = entry_id_consultor.get().strip()
        if not id_consultor.isdigit():
            messagebox.showerror("Error", "El Id del consultor debe ser un número.")
            return

        cursor.execute("SELECT * FROM CuentasPorPagarConsultor WHERE IdCuentasPorPagarConsultor = ?", id_consultor)
        datos = cursor.fetchone()
        if not datos:
            limpiar_campos_consultor()
            messagebox.showerror("Error", f"No se encontró información con el Id {id_consultor}")
            return

        num_orden = datos[1]  # NumOrdenCompra
        nombre_consultor = datos[2]

        # Buscar saldo actualizado real
        cursor.execute("""
            SELECT SUM(ComisionPorPagarConsultor)
            FROM OrdenCompra
            WHERE NumOrdenCompra = ?
        """, (num_orden,))
        total_por_pagar = cursor.fetchone()[0] or 0

        cursor.execute("""
            SELECT SUM(PagadoConsultor)
            FROM CuentasPorPagarConsultor
            WHERE NumOrdenCompra = ?
        """, (num_orden,))
        total_pagado = cursor.fetchone()[0] or 0

        restante = round(float(total_por_pagar) - float(total_pagado), 2)

        # Asignaciones visuales
        entry_nombre_consultor.config(state="normal")
        entry_nombre_consultor.delete(0, tk.END)
        entry_nombre_consultor.insert(0, nombre_consultor)
        entry_nombre_consultor.config(state="readonly")

        entry_por_pagar_consultor.config(state="normal")
        entry_por_pagar_consultor.delete(0, tk.END)
        entry_por_pagar_consultor.insert(0, restante)
        entry_por_pagar_consultor.config(state="readonly")

        entry_valor_pago_consultor.config(state="normal")
        entry_valor_pago_consultor.delete(0, tk.END)
        entry_valor_pago_consultor.insert(0, datos[4])  # PagadoConsultor

        entry_banco_consultor.config(state="normal")
        entry_banco_consultor.delete(0, tk.END)
        entry_banco_consultor.insert(0, datos[5])  # BancoDistrConsultor

        entry_cuenta_consultor.config(state="normal")
        entry_cuenta_consultor.delete(0, tk.END)
        entry_cuenta_consultor.insert(0, datos[6])  # CuentaDistrConsultor

        entry_fecha_pago_consultor.config(state="normal")
        try:
            entry_fecha_pago_consultor.set_date(datos[7])  # FechaPagoConsultor
        except:
            messagebox.showwarning("Advertencia", f"Fecha no válida: {datos[7]}")
            entry_fecha_pago_consultor.set_date(date.today())

        entry_comprobante_consultor.config(state="normal")
        entry_comprobante_consultor.delete(0, tk.END)
        if datos[8]:  # NumComprobante
            entry_comprobante_consultor.insert(0, datos[8])
        entry_comprobante_consultor.config(state="normal")

    except Exception as e:
        limpiar_campos_consultor()
        messagebox.showerror("Error", f"No se pudieron cargar los datos: {e}")




# === Función: Modificar pago consultor y recalcular saldo ===
def modificar_pago_consultor():
    try:
        id_consultor = entry_id_consultor.get().strip()
        if not id_consultor.isdigit():
            messagebox.showerror("Error", "Debe ingresar un Id válido.")
            return

        nuevo_valor = float(entry_valor_pago_consultor.get())
        banco = entry_banco_consultor.get()
        cuenta = entry_cuenta_consultor.get()
        fecha = entry_fecha_pago_consultor.get_date()
        comprobante = entry_comprobante_consultor.get()

        if not banco.isalpha():
            messagebox.showwarning("Advertencia", "Verifique que Banco sea texto")
            return

        # Obtener NumOrdenCompra
        cursor.execute("SELECT NumOrdenCompra FROM CuentasPorPagarConsultor WHERE IdCuentasPorPagarConsultor = ?", id_consultor)
        num_orden = cursor.fetchone()[0]

        cursor.execute("""
            UPDATE CuentasPorPagarConsultor
            SET PagadoConsultor = ?, BancoDistrConsultor = ?, CuentaDistrConsultor = ?, 
                FechaPagoConsultor = ?, NumComprobante = ?
            WHERE IdCuentasPorPagarConsultor = ?
        """, (nuevo_valor, banco, cuenta, fecha, comprobante, id_consultor))
        conn.commit()

        recalcular_saldos_consultor(num_orden)

        messagebox.showinfo("Éxito", "Pago del consultor modificado correctamente.")
        entry_id_consultor.delete(0, tk.END)
        limpiar_campos_consultor()

    except Exception as e:
        messagebox.showerror("Error", f"No se pudo modificar el pago: {e}")



# === Función: Eliminar pago consultor y recalcular saldos ===
def eliminar_pago_consultor():
    try:
        id_consultor = entry_id_consultor.get().strip()
        if not id_consultor.isdigit():
            messagebox.showerror("Error", "Debe ingresar un Id válido.")
            return

        cursor.execute("SELECT NumOrdenCompra, NombreConsultor, PagadoConsultor FROM CuentasPorPagarConsultor WHERE IdCuentasPorPagarConsultor = ?", id_consultor)
        data = cursor.fetchone()
        if not data:
            messagebox.showerror("Error", "No se encontró el pago con ese Id.")
            return
        num_orden, nombre, valor_eliminado = data

        confirm = messagebox.askyesno("Confirmar", f"¿Está seguro de eliminar el pago de {valor_eliminado} para {nombre}?")
        if not confirm:
            return

        cursor.execute("DELETE FROM CuentasPorPagarConsultor WHERE IdCuentasPorPagarConsultor = ?", id_consultor)
        conn.commit()

        messagebox.showinfo("Éxito", "Pago eliminado correctamente.")
        recalcular_saldos_consultor(num_orden)

        entry_id_consultor.delete(0, tk.END)
        limpiar_campos_consultor()

    except Exception as e:
        limpiar_campos_consultor()
        messagebox.showerror("Error", f"No se pudo eliminar el pago: {e}")




# === Botones debajo de IdConsultor ===
frame_botones_cpc = ttk.Frame(tab_cuentas_pagar)
frame_botones_cpc.grid(row=11, column=0, columnspan=3, padx=10, pady=10, sticky="w")

btn_cargar = tk.Button(frame_botones_cpc, text="📂 Cargar Datos", bg="lightgreen", command=cargar_datos_consultor)
btn_cargar.pack(side="left", padx=5)

btn_modificar = tk.Button(frame_botones_cpc, text="🛠️ Modificar Pago", bg="skyblue", command=modificar_pago_consultor)
btn_modificar.pack(side="left", padx=5)

btn_eliminar = tk.Button(frame_botones_cpc, text="🗑️ Eliminar Pago", bg="tomato", command=eliminar_pago_consultor)
btn_eliminar.pack(side="left", padx=5)

def recalcular_saldos_consultor(num_orden):
    try:
        cursor.execute("""
            SELECT IdCuentasPorPagarConsultor, PagadoConsultor
            FROM CuentasPorPagarConsultor
            WHERE NumOrdenCompra = ?
            ORDER BY FechaPagoConsultor ASC, IdCuentasPorPagarConsultor ASC
        """, (num_orden,))
        pagos = cursor.fetchall()

        cursor.execute("""
            SELECT SUM(ComisionPorPagarConsultor)
            FROM OrdenCompra
            WHERE NumOrdenCompra = ?
        """, (num_orden,))
        total_comision = cursor.fetchone()[0] or 0

        pagado_acumulado = 0
        for pago in pagos:
            id_pago = pago[0]
            valor_actual = pago[1] or 0

            saldo_previo = round(total_comision - pagado_acumulado, 2)
            saldo_final = round(saldo_previo - valor_actual, 2)

            cursor.execute("""
                UPDATE CuentasPorPagarConsultor
                SET SaldoPorPagarConsultor = ?, SaldoFinal = ?
                WHERE IdCuentasPorPagarConsultor = ?
            """, (saldo_previo, saldo_final, id_pago))

            pagado_acumulado += valor_actual

        conn.commit()

    except Exception as e:
        messagebox.showerror("Error", f"No se pudieron recalcular los saldos del consultor: {e}")



ttk.Label(tab_cuentas_pagar, text="🔍 Id Cuentas Por Pagar Padre Empresarial:").grid(row=10, column=6, sticky="e", padx=(10, 2), pady=5)
entry_id_padre = ttk.Entry(tab_cuentas_pagar, width=22)
entry_id_padre.grid(row=10, column=7, padx=(2, 10), pady=5)


def cargar_datos_padre():
    try:
        id_padre = entry_id_padre.get().strip()
        if not id_padre.isdigit():
            messagebox.showerror("Error", "El Id del padre empresarial debe ser un número.")
            return

        cursor.execute("""
            SELECT * FROM CuentasPorPagarPadreEmpresarial 
            WHERE IdCuentasPorPagarPadreEmpresarial = ?
        """, (id_padre,))
        datos = cursor.fetchone()
        if not datos:
            limpiar_campos_padre()
            messagebox.showerror("Error", f"No se encontró información con el Id {id_padre}")
            return

        # Índices correctos según la estructura:
        # 0: Id, 1: NumOrdenCompra, 2: Nombre, 3: ComisionTotal, 4: Pagado, 
        # 5: Banco, 6: Cuenta, 7: Fecha, 8: Comprobante

        num_orden = datos[1]
        nombre_padre = datos[2]

        # 🔵 Buscar saldo real actualizado
        cursor.execute("""
            SELECT SUM(ComisionPorPagarPadreEmpresarial)
            FROM OrdenCompra
            WHERE NumOrdenCompra = ?
        """, (num_orden,))
        total_por_pagar = cursor.fetchone()[0] or 0

        cursor.execute("""
            SELECT SUM(PagadoPadreEmpresarial)
            FROM CuentasPorPagarPadreEmpresarial
            WHERE NumOrdenCompra = ?
        """, (num_orden,))
        total_pagado = cursor.fetchone()[0] or 0

        restante = round(float(total_por_pagar) - float(total_pagado), 2)

        # 🔵 Asignaciones visuales
        entry_nombre_padre.config(state="normal")
        entry_nombre_padre.delete(0, tk.END)
        entry_nombre_padre.insert(0, nombre_padre)
        entry_nombre_padre.config(state="readonly")

        entry_por_pagar_padre.config(state="normal")
        entry_por_pagar_padre.delete(0, tk.END)
        entry_por_pagar_padre.insert(0, restante)
        entry_por_pagar_padre.config(state="readonly")

        entry_valor_pago_padre.config(state="normal")
        entry_valor_pago_padre.delete(0, tk.END)
        entry_valor_pago_padre.insert(0, datos[4])

        entry_banco_padre.config(state="normal")
        entry_banco_padre.delete(0, tk.END)
        entry_banco_padre.insert(0, datos[5])

        entry_cuenta_padre.config(state="normal")
        entry_cuenta_padre.delete(0, tk.END)
        entry_cuenta_padre.insert(0, datos[6])

        entry_fecha_pago_padre.config(state="normal")
        try:
            entry_fecha_pago_padre.set_date(datos[7])
        except:
            messagebox.showwarning("Advertencia", f"Fecha no válida: {datos[7]}")
            entry_fecha_pago_padre.set_date(date.today())

        entry_comprobante_padre.config(state="normal")
        entry_comprobante_padre.delete(0, tk.END)
        if datos[8] is not None:
            entry_comprobante_padre.insert(0, datos[8])
        # Permitir edición:
        entry_comprobante_padre.config(state="normal")

    except Exception as e:
        limpiar_campos_padre()
        messagebox.showerror("Error", f"No se pudieron cargar los datos: {e}")



from decimal import Decimal, ROUND_HALF_UP

def modificar_pago_padre():
    try:
        id_padre = entry_id_padre.get().strip()
        if not id_padre.isdigit():
            messagebox.showerror("Error", "Debe ingresar un Id válido.")
            return

        nuevo_valor = Decimal(entry_valor_pago_padre.get() or "0").quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        banco = entry_banco_padre.get()
        cuenta = entry_cuenta_padre.get()
        fecha = entry_fecha_pago_padre.get_date()
        comprobante = entry_comprobante_padre.get()

        if not banco.isalpha():
            messagebox.showwarning("Advertencia", "Verifique que Banco sea texto")
            return

        # Obtener NumOrdenCompra
        cursor.execute("""
            SELECT NumOrdenCompra FROM CuentasPorPagarPadreEmpresarial
            WHERE IdCuentasPorPagarPadreEmpresarial = ?
        """, (id_padre,))
        num_orden = cursor.fetchone()[0]

        cursor.execute("""
            UPDATE CuentasPorPagarPadreEmpresarial
            SET PagadoPadreEmpresarial = ?, BancoDistrPadreEmpresarial = ?, CuentaDistriPadreEmpresarial = ?, 
                FechaPagoPadreEmpresarial = ?, NumComprobante = ?
            WHERE IdCuentasPorPagarPadreEmpresarial = ?
        """, (float(nuevo_valor), banco, cuenta, fecha, comprobante, id_padre))
        conn.commit()

        # 🔁 Recalcular saldos
        recalcular_saldos_padre(num_orden)

        messagebox.showinfo("Éxito", "Pago del padre empresarial modificado correctamente.")
        entry_id_padre.delete(0, tk.END)
        limpiar_campos_padre()

    except Exception as e:
        messagebox.showerror("Error", f"No se pudo modificar el pago: {e}")

def eliminar_pago_padre():
    try:
        id_padre = entry_id_padre.get().strip()
        if not id_padre.isdigit():
            messagebox.showerror("Error", "Debe ingresar un Id válido.")
            return

        cursor.execute("""
            SELECT NumOrdenCompra, NombrePadreEmpresarial, PagadoPadreEmpresarial
            FROM CuentasPorPagarPadreEmpresarial
            WHERE IdCuentasPorPagarPadreEmpresarial = ?
        """, (id_padre,))
        data = cursor.fetchone()
        if not data:
            messagebox.showerror("Error", "No se encontró el pago con ese Id.")
            return

        num_orden, nombre, valor_eliminado = data

        confirm = messagebox.askyesno("Confirmar", f"¿Está seguro de eliminar el pago de {valor_eliminado} para {nombre}?")
        if not confirm:
            return

        cursor.execute("""
            DELETE FROM CuentasPorPagarPadreEmpresarial
            WHERE IdCuentasPorPagarPadreEmpresarial = ?
        """, (id_padre,))
        conn.commit()

        # 🔁 Recalcular saldos
        recalcular_saldos_padre(num_orden)

        messagebox.showinfo("Éxito", "Pago eliminado correctamente.")
        entry_id_padre.delete(0, tk.END)
        limpiar_campos_padre()

    except Exception as e:
        limpiar_campos_padre()
        messagebox.showerror("Error", f"No se pudo eliminar el pago: {e}")



from decimal import Decimal, ROUND_HALF_UP

def recalcular_saldos_padre(num_orden):
    try:
        cursor.execute("""
            SELECT IdCuentasPorPagarPadreEmpresarial, PagadoPadreEmpresarial
            FROM CuentasPorPagarPadreEmpresarial
            WHERE NumOrdenCompra = ?
            ORDER BY FechaPagoPadreEmpresarial ASC, IdCuentasPorPagarPadreEmpresarial ASC
        """, (num_orden,))
        pagos = cursor.fetchall()

        cursor.execute("""
            SELECT SUM(ComisionPorPagarPadreEmpresarial)
            FROM OrdenCompra
            WHERE NumOrdenCompra = ?
        """, (num_orden,))
        total_comision = Decimal(str(cursor.fetchone()[0] or 0)).quantize(Decimal("0.01"))

        pagado_acumulado = Decimal("0.00")

        for pago in pagos:
            id_pago = pago[0]
            valor_actual = Decimal(str(pago[1] or 0)).quantize(Decimal("0.01"))

            saldo_previo = (total_comision - pagado_acumulado).quantize(Decimal("0.01"))
            saldo_final = (saldo_previo - valor_actual).quantize(Decimal("0.01"))

            cursor.execute("""
                UPDATE CuentasPorPagarPadreEmpresarial
                SET SaldoPorPagarPadreEmpresarial = ?, SaldoFinal = ?
                WHERE IdCuentasPorPagarPadreEmpresarial = ?
            """, (float(saldo_previo), float(saldo_final), id_pago))

            pagado_acumulado += valor_actual

        conn.commit()

    except Exception as e:
        messagebox.showerror("Error", f"No se pudieron recalcular los saldos del padre empresarial: {e}")





frame_botones_padre = ttk.Frame(tab_cuentas_pagar)
frame_botones_padre.grid(row=11, column=6, columnspan=3, padx=280, pady=10, sticky="w")

btn_cargar_padre = tk.Button(frame_botones_padre, text="📂 Cargar Datos", bg="lightgreen", command=cargar_datos_padre)
btn_cargar_padre.pack(side="left", padx=5)

btn_modificar_padre = tk.Button(frame_botones_padre, text="🛠️ Modificar Pago", bg="skyblue", command=modificar_pago_padre)
btn_modificar_padre.pack(side="left", padx=5)

btn_eliminar_padre = tk.Button(frame_botones_padre, text="🗑️ Eliminar Pago", bg="tomato", command=eliminar_pago_padre)
btn_eliminar_padre.pack(side="left", padx=5)

from tkinter import ttk, messagebox
from tkcalendar import DateEntry
from datetime import date
from decimal import Decimal, InvalidOperation
import tkinter as tk

def safe_decimal(value):
    try:
        return Decimal(str(value).replace(",", "")).quantize(Decimal("0.01"))
    except (InvalidOperation, TypeError):
        return Decimal("0.00")

def mostrar_cuentasporpagarconsultor():
    try:
        ventana = tk.Toplevel(root)
        ventana.title("Lista de Cuentas por Pagar Consultor")
        ventana.geometry("1920x1080")

        # Filtros
        frame_filtros = ttk.Frame(ventana)
        frame_filtros.pack(pady=10)

        ttk.Label(frame_filtros, text="🔎 Nombre del Consultor:").grid(row=0, column=0, padx=5)
        entry_filtro_consultor = ttk.Entry(frame_filtros, width=25)
        entry_filtro_consultor.grid(row=0, column=1, padx=5)

        from datetime import date

        ttk.Label(frame_filtros, text="📅 Fecha Inicio:").grid(row=0, column=2, padx=5)
        entry_fecha_inicio = DateEntry(frame_filtros, width=15, date_pattern="yyyy-mm-dd")
        entry_fecha_inicio.set_date(date(2025, 1, 1))  # 1 de mayo 2025
        entry_fecha_inicio.grid(row=0, column=3, padx=5)

        ttk.Label(frame_filtros, text="📅 Fecha Fin:").grid(row=0, column=4, padx=5)
        entry_fecha_fin = DateEntry(frame_filtros, width=15, date_pattern="yyyy-mm-dd")
        entry_fecha_fin.set_date(date.today())
        entry_fecha_fin.grid(row=0, column=5, padx=5)

        # Tabla
        frame_tree = ttk.Frame(ventana)
        frame_tree.pack(fill=tk.BOTH, expand=True)

        columnas_tree = [
            "IdCuentasPorPagarConsultor", "NumOrdenCompra", "NombreConsultor",
            "ComisionPorPagarConsultorTotal", "PagadoConsultor",
            "BancoDistrConsultor", "CuentaDistrConsultor",
            "FechaPagoConsultor", "NumComprobante",
            "SaldoPorPagarConsultor", "SaldoFinal"
        ]

        scrollbar_x = ttk.Scrollbar(frame_tree, orient=tk.HORIZONTAL)
        scrollbar_y = ttk.Scrollbar(frame_tree, orient=tk.VERTICAL)

        tree = ttk.Treeview(
            frame_tree, columns=columnas_tree, show="headings",
            xscrollcommand=scrollbar_x.set, yscrollcommand=scrollbar_y.set
        )
        scrollbar_x.config(command=tree.xview)
        scrollbar_y.config(command=tree.yview)
        scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)
        scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)

        for col in columnas_tree:
            tree.heading(col, text=col)
            tree.column(col, width=300, anchor="center")
        tree.pack(fill=tk.BOTH, expand=True)

        datos_exportacion = []

        # Texto resumen global
        frame_text = ttk.Frame(ventana)
        frame_text.pack(fill="x", padx=10, pady=10)
        text_area = tk.Text(frame_text, height=6, font=("Arial", 12), wrap="word")
        text_area.config(state="disabled")
        text_area.pack(fill="both", expand=True)

        def ejecutar_consulta():
            try:
                tree.delete(*tree.get_children())
                datos_exportacion.clear()
                text_area.config(state="normal")
                text_area.delete("1.0", tk.END)

                filtro_nombre = entry_filtro_consultor.get().strip().lower()
                fecha_inicio = entry_fecha_inicio.get_date()
                fecha_fin = entry_fecha_fin.get_date()

                cursor.execute("""
                    SELECT * FROM CuentasPorPagarConsultor
                    WHERE FechaPagoConsultor BETWEEN ? AND ?
                    ORDER BY FechaPagoConsultor ASC, IdCuentasPorPagarConsultor ASC
                """, (fecha_inicio, fecha_fin))
                cuentasporpagar = cursor.fetchall()

                orden_map = {}
                for cuenta in cuentasporpagar:
                    nombre = (cuenta[2] or "").strip().lower()
                    if filtro_nombre and nombre != filtro_nombre:
                        continue
                    num_orden = cuenta[1]
                    if num_orden not in orden_map:
                        orden_map[num_orden] = []
                    orden_map[num_orden].append(cuenta)

                for num_orden, cuentas in orden_map.items():
                    cuentas_ordenadas = sorted(cuentas, key=lambda c: c[0])
                    comision_total = max(safe_decimal(c[3]) for c in cuentas_ordenadas)
                    pagos_acumulados = Decimal("0.00")

                    for i, cuenta in enumerate(cuentas_ordenadas):
                        pagado = safe_decimal(cuenta[4])
                        pagos_acumulados += pagado
                        saldo_final = (comision_total - pagos_acumulados).quantize(Decimal("0.01")) if i == len(cuentas_ordenadas) - 1 else Decimal("0.00")

                        fila = list(cuenta[:10])
                        fila.append(str(saldo_final))
                        fila = [str(x) if x is not None else "" for x in fila]
                        tree.insert("", tk.END, values=fila)
                        datos_exportacion.append(fila)

                # Resumen
                cursor.execute("""
                    SELECT NumOrdenCompra,
                           SUM(ComisionPorPagarConsultor) AS ComisionTotal,
                           (SELECT SUM(PagadoConsultor)
                            FROM CuentasPorPagarConsultor
                            WHERE NumOrdenCompra = oc.NumOrdenCompra) AS Pagado
                    FROM OrdenCompra oc
                    WHERE NombreConsultor IS NOT NULL AND NombreConsultor != ''
                    GROUP BY NumOrdenCompra
                    HAVING (SUM(ComisionPorPagarConsultor) - ISNULL((SELECT SUM(PagadoConsultor)
                             FROM CuentasPorPagarConsultor
                             WHERE NumOrdenCompra = oc.NumOrdenCompra), 0)) > 0
                """)
                rows = cursor.fetchall()

                ordenes = ", ".join([str(row[0]) for row in rows]) or "Ninguna"
                saldo_total = Decimal(sum(
                    safe_decimal(row[1]) - safe_decimal(row[2]) for row in rows
                )).quantize(Decimal("0.01"))

                texto = f"""
📋 Números de Orden de Compra con saldo pendiente:
{ordenes}

💰 Saldo total por pagar a consultores: ${saldo_total}
"""
                text_area.insert(tk.END, texto.strip())
                text_area.config(state="disabled")

            except Exception as e:
                messagebox.showerror("Error", f"No se pudo ejecutar la consulta: {str(e)}")

        # Botón consultar
        btn_consultar = ttk.Button(frame_filtros, text="🔍 Consultar", command=ejecutar_consulta)
        btn_consultar.grid(row=0, column=6, padx=10)

        # Botón exportar
        frame_exportar = ttk.Frame(ventana)
        frame_exportar.pack(pady=10)
        btn_excel = ttk.Button(
            frame_exportar,
            text="📥 Exportar a Excel",
            command=lambda: exportar_excel(datos_exportacion, columnas_tree)
        )
        btn_excel.pack()

        ejecutar_consulta()

    except Exception as e:
        messagebox.showerror("Error", f"No se pudo cargar la ventana: {str(e)}")




btn_mostrar_pagos_consultores = tk.Button(
    tab_cuentas_pagar,
    text="📂 Mostrar Pagos a Consultores",
    command=mostrar_cuentasporpagarconsultor,
    bg="#5CD6C0",
    fg="black",
    font=("Arial", 10, "bold"),
    relief="raised",
    padx=10,
    pady=1,
    activebackground="#47B7A7",
    cursor="hand2"
)
btn_mostrar_pagos_consultores.grid(row=100, column=0, columnspan=3, pady=(0, 15))

from decimal import Decimal, InvalidOperation
from tkinter import ttk, messagebox
import tkinter as tk
from tkcalendar import DateEntry
from datetime import date

def safe_decimal(value):
    try:
        return Decimal(str(value or "0").replace(",", "")).quantize(Decimal("0.01"))
    except (InvalidOperation, TypeError):
        return Decimal("0.00")

def mostrar_cuentasporpagarpadreempresarial():
    try:
        ventana = tk.Toplevel(root)
        ventana.title("Lista de Cuentas por Pagar Padre Empresarial")
        ventana.geometry("1920x1080")

        # Filtros
        frame_filtros = ttk.Frame(ventana)
        frame_filtros.pack(pady=10)

        ttk.Label(frame_filtros, text="🔎 Nombre del Padre Empresarial:").grid(row=0, column=0, padx=5)
        entry_nombre = ttk.Entry(frame_filtros, width=30)
        entry_nombre.grid(row=0, column=1, padx=5)

        from datetime import date

        ttk.Label(frame_filtros, text="📅 Fecha Inicio:").grid(row=0, column=2, padx=5)
        entry_inicio = DateEntry(frame_filtros, width=15, date_pattern="yyyy-mm-dd")
        entry_inicio.set_date(date(2025, 1, 1))  # 1 de mayo 2025
        entry_inicio.grid(row=0, column=3, padx=5)

        ttk.Label(frame_filtros, text="📅 Fecha Fin:").grid(row=0, column=4, padx=5)
        entry_fin = DateEntry(frame_filtros, width=15, date_pattern="yyyy-mm-dd")
        entry_fin.set_date(date.today())
        entry_fin.grid(row=0, column=5, padx=5)

        # Tabla
        frame_tree = ttk.Frame(ventana)
        frame_tree.pack(fill=tk.BOTH, expand=True)

        columnas = [
            "IdCuentasPorPagarPadreEmpresarial", "NumOrdenCompra", "NombrePadreEmpresarial",
            "ComisionPorPagarPadreEmpresarialTotal", "PagadoPadreEmpresarial", "BancoDistrPadreEmpresarial",
            "CuentaDistrPadreEmpresarial", "FechaPagoPadreEmpresarial", "NumComprobante",
            "SaldoPorPagarPadreEmpresarial", "SaldoFinal"
        ]

        scrollbar_x = ttk.Scrollbar(frame_tree, orient=tk.HORIZONTAL)
        scrollbar_y = ttk.Scrollbar(frame_tree, orient=tk.VERTICAL)

        tree = ttk.Treeview(
            frame_tree, columns=columnas, show="headings",
            xscrollcommand=scrollbar_x.set, yscrollcommand=scrollbar_y.set
        )
        scrollbar_x.config(command=tree.xview)
        scrollbar_y.config(command=tree.yview)
        scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)
        scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)

        for col in columnas:
            tree.heading(col, text=col)
            tree.column(col, anchor="center", width=450)

        tree.pack(fill=tk.BOTH, expand=True)
        datos_exportacion = []

        # Texto resumen
        frame_text = ttk.Frame(ventana)
        frame_text.pack(fill="x", padx=10, pady=10)
        text_area = tk.Text(frame_text, height=6, font=("Arial", 12), wrap="word")
        text_area.config(state="disabled")
        text_area.pack(fill="both", expand=True)

        def ejecutar_consulta():
            try:
                tree.delete(*tree.get_children())
                datos_exportacion.clear()
                text_area.config(state="normal")
                text_area.delete("1.0", tk.END)

                nombre_filtro = entry_nombre.get().strip().lower()
                fecha_ini = entry_inicio.get_date()
                fecha_fin = entry_fin.get_date()

                cursor.execute("""
                    SELECT * FROM CuentasPorPagarPadreEmpresarial
                    ORDER BY FechaPagoPadreEmpresarial ASC, IdCuentasPorPagarPadreEmpresarial ASC
                """)
                cuentas = cursor.fetchall()

                orden_map = {}
                for cuenta in cuentas:
                    nombre = str(cuenta[2] or "").strip().lower()
                    fecha_pago = cuenta[7]
                    if (not nombre_filtro or nombre == nombre_filtro) and fecha_ini <= fecha_pago <= fecha_fin:
                        num_orden = cuenta[1]
                        if num_orden not in orden_map:
                            orden_map[num_orden] = []
                        orden_map[num_orden].append(cuenta)

                for num_orden, cuentas_orden in orden_map.items():
                    cuentas_ordenadas = sorted(cuentas_orden, key=lambda c: c[0])
                    comision_total = max(safe_decimal(c[3]) for c in cuentas_ordenadas)
                    pagos_acumulados = Decimal("0.00")

                    for i, cuenta in enumerate(cuentas_ordenadas):
                        pagado = safe_decimal(cuenta[4])
                        pagos_acumulados += pagado
                        saldo_final = (comision_total - pagos_acumulados).quantize(Decimal("0.01")) if i == len(cuentas_ordenadas) - 1 else Decimal("0.00")

                        fila = list(cuenta[:10])
                        fila.append(str(saldo_final))
                        fila = [str(x) if x is not None else "" for x in fila]
                        tree.insert("", tk.END, values=fila)
                        datos_exportacion.append(fila)

                # Resumen
                cursor.execute("""
                    SELECT NumOrdenCompra,
                           SUM(ComisionPorPagarPadreEmpresarial) AS ComisionTotal,
                           (SELECT SUM(PagadoPadreEmpresarial)
                            FROM CuentasPorPagarPadreEmpresarial
                            WHERE NumOrdenCompra = oc.NumOrdenCompra) AS Pagado
                    FROM OrdenCompra oc
                    WHERE NombrePadreEmpresarial IS NOT NULL AND NombrePadreEmpresarial != ''
                    GROUP BY NumOrdenCompra
                    HAVING (SUM(ComisionPorPagarPadreEmpresarial) - ISNULL((SELECT SUM(PagadoPadreEmpresarial)
                             FROM CuentasPorPagarPadreEmpresarial
                             WHERE NumOrdenCompra = oc.NumOrdenCompra), 0)) > 0
                """)
                rows = cursor.fetchall()
                ordenes = ", ".join([str(row[0]) for row in rows]) or "Ninguna"
                saldo_total = Decimal(sum(
                    safe_decimal(row[1]) - safe_decimal(row[2]) for row in rows
                )).quantize(Decimal("0.01"))

                texto = f"""
📋 Números de Orden de Compra con saldo pendiente:
{ordenes}

💰 Saldo total por pagar a padres empresariales: ${saldo_total}
"""
                text_area.insert(tk.END, texto.strip())
                text_area.config(state="disabled")

            except Exception as err:
                messagebox.showerror("Error", f"No se pudo realizar la consulta: {err}")

        # Botón buscar
        btn_buscar = ttk.Button(frame_filtros, text="🔍 Buscar", command=ejecutar_consulta)
        btn_buscar.grid(row=0, column=6, padx=10)

        # Botón exportar
        frame_exportar = ttk.Frame(ventana)
        frame_exportar.pack(pady=10)
        btn_exportar = ttk.Button(
            frame_exportar,
            text="📥 Exportar a Excel",
            command=lambda: exportar_excel(datos_exportacion, columnas)
        )
        btn_exportar.pack()

        ejecutar_consulta()

    except Exception as e:
        messagebox.showerror("Error", f"No se pudo abrir la ventana: {e}")



btn_consultar_cxp_padre = tk.Button(
    tab_cuentas_pagar,
    text="📋 Mostrar Pagos a Padres Empresariales",
    command=mostrar_cuentasporpagarpadreempresarial,
    bg="#5CD6C0",
    fg="black",
    font=("Arial", 10, "bold"),
    relief="raised",
    padx=10,
    pady=1,
    activebackground="#47B7A7",
    cursor="hand2"
)
btn_consultar_cxp_padre.grid(row=100, column=6, columnspan=3)


import openpyxl
from openpyxl.utils import get_column_letter

from tkinter import filedialog
import openpyxl

def exportar_excel_consultores(datos_exportacion):
    try:
        ruta_archivo = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivos Excel", "*.xlsx")],
            title="Guardar archivo de saldo consultores",
            initialfile="Saldo_Consultores.xlsx"
        )

        if not ruta_archivo:
            return

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Saldo Consultores"

        # Nuevo encabezado con órdenes
        encabezados = ["Consultor", "Órdenes de Compra", "Comisión Total", "Pagado", "Saldo"]
        for col, encabezado in enumerate(encabezados, 1):
            sheet.cell(row=1, column=col).value = encabezado

        for fila_idx, fila in enumerate(datos_exportacion, 2):
            for col_idx, valor in enumerate(fila, 1):
                cell = sheet.cell(row=fila_idx, column=col_idx)
                cell.value = valor
                # Formatear como número si aplica
                if col_idx >= 3:  # Comisión, Pagado, Saldo
                    cell.number_format = '#,##0.00'

        workbook.save(ruta_archivo)
        messagebox.showinfo("Éxito", f"Archivo guardado exitosamente en:\n{ruta_archivo}")

    except Exception as e:
        messagebox.showerror("Error", f"No se pudo exportar a Excel: {e}")


def consultar_cuentas_por_pagar_consultor():
    try:
        ventana = tk.Toplevel(root)
        ventana.title("Consultar Cuentas por Pagar - Consultores")
        ventana.geometry("1920x1080")

        # Filtros
        frame_filtros = ttk.Frame(ventana)
        frame_filtros.pack(pady=10)

        ttk.Label(frame_filtros, text="🔎 Nombre del Consultor:").grid(row=0, column=0, padx=5)
        entry_filtro = ttk.Entry(frame_filtros, width=30)
        entry_filtro.grid(row=0, column=1, padx=5)

        ttk.Label(frame_filtros, text="🎯 Estado:").grid(row=0, column=2, padx=5)
        combo_estado = ttk.Combobox(frame_filtros, values=[
            "Mostrar todas", "Pagadas (Saldo = 0)", "Por pagar (Saldo > 0)"
        ], state="readonly", width=30)
        combo_estado.grid(row=0, column=3, padx=5)
        combo_estado.current(0)

        # Tabla
        frame_tree = ttk.Frame(ventana)
        frame_tree.pack(fill=tk.BOTH, expand=True)

        columnas = ["NumOrdenCompra", "NombreConsultor", "ComisionTotal", "Pagado", "Saldo"]
        datos_exportacion = []

        scrollbar_x = ttk.Scrollbar(frame_tree, orient="horizontal")
        scrollbar_y = ttk.Scrollbar(frame_tree, orient="vertical")

        tree = ttk.Treeview(frame_tree, columns=columnas, show="headings",
                            xscrollcommand=scrollbar_x.set, yscrollcommand=scrollbar_y.set)
        scrollbar_x.config(command=tree.xview)
        scrollbar_y.config(command=tree.yview)
        scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)
        scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)

        for col in columnas:
            tree.heading(col, text=col)
            tree.column(col, anchor="center", width=240)

        tree.pack(fill=tk.BOTH, expand=True)

        # Área de texto con resumen
        frame_text = ttk.Frame(ventana)
        frame_text.pack(fill="x", padx=10, pady=10)

        text_area = tk.Text(frame_text, height=10, font=("Arial", 11), wrap="word")
        text_area.pack(fill="both", expand=True)
        text_area.config(state="disabled")

        # Función para ejecutar consulta
        def ejecutar_consulta():
            try:
                for item in tree.get_children():
                    tree.delete(item)
                datos_exportacion.clear()

                filtro_nombre = entry_filtro.get().strip()
                filtro_estado = combo_estado.get()

                cursor.execute("""
                    SELECT 
                        oc.NumOrdenCompra,
                        oc.NombreConsultor,
                        SUM(oc.ComisionPorPagarConsultor) AS ComisionTotal,
                        ISNULL((
                            SELECT SUM(PagadoConsultor)
                            FROM CuentasPorPagarConsultor
                            WHERE NumOrdenCompra = oc.NumOrdenCompra
                        ), 0) AS Pagado
                    FROM OrdenCompra oc
                    WHERE oc.NombreConsultor IS NOT NULL AND oc.NombreConsultor != ''
                    GROUP BY oc.NumOrdenCompra, oc.NombreConsultor
                    ORDER BY oc.NumOrdenCompra
                """)
                filas = cursor.fetchall()

                total_general = 0
                texto = "📢 SALDO POR PAGAR A CONSULTORES\n\n"

                for fila in filas:
                    num_orden, nombre, comision, pagado = fila
                    saldo = round(comision - pagado, 2)

                    # Filtro por nombre exacto
                    if filtro_nombre and nombre.lower() != filtro_nombre.lower():
                        continue

                    # Filtro por estado
                    if filtro_estado == "Pagadas (Saldo = 0)" and saldo != 0:
                        continue
                    if filtro_estado == "Por pagar (Saldo > 0)" and saldo <= 0:
                        continue

                    fila_mostrar = [str(num_orden), nombre, f"{comision:.2f}", f"{pagado:.2f}", f"{saldo:.2f}"]
                    tree.insert("", tk.END, values=fila_mostrar)
                    datos_exportacion.append(fila_mostrar)

                    texto += f"👤 {nombre.upper()}\n🧾 Orden: {num_orden}\n💰 Comisión ${comision:.2f} - 💸 Pagado ${pagado:.2f} → 💼 Saldo ${saldo:.2f}\n\n"
                    total_general += saldo

                texto += f"💰 Total General de Saldo: ${total_general:.2f}"

                text_area.config(state="normal")
                text_area.delete("1.0", tk.END)
                text_area.insert(tk.END, texto.strip())
                text_area.config(state="disabled")

                if not datos_exportacion:
                    messagebox.showinfo("Sin resultados", "No se encontraron registros con los filtros aplicados.")

            except Exception as e:
                messagebox.showerror("Error", f"No se pudo realizar la consulta: {e}")

        # Botón Consultar
        btn_consultar = tk.Button(frame_filtros, text="🔍 Consultar", bg="lightgreen", font=("Arial", 10, "bold"), command=ejecutar_consulta)
        btn_consultar.grid(row=0, column=4, padx=10)

        # Exportar Excel
        frame_exportar = ttk.Frame(ventana)
        frame_exportar.pack(pady=10)

        btn_exportar_excel = ttk.Button(
            frame_exportar,
            text="📥 Exportar a Excel",
            command=lambda: exportar_excel_consultores(datos_exportacion)
        )
        btn_exportar_excel.pack()

        # Ejecutar por defecto
        ejecutar_consulta()

    except Exception as e:
        messagebox.showerror("Error", f"Ocurrió un error: {e}")



btn_consultar_cxp_consultor = tk.Button(
    tab_cuentas_pagar,
    text="🔍 Consultar Cuentas Por Pagar Consultor",
    command=consultar_cuentas_por_pagar_consultor,
    bg="#5CD6C0",
    fg="black",
    font=("Arial", 10, "bold"),
    relief="raised",
    padx=10,
    pady=1,
    activebackground="#47B7A7",
    cursor="hand2"
)
btn_consultar_cxp_consultor.grid(row=101, column=0, columnspan=3)


def exportar_excel_padres(datos_exportacion):
    try:
        ruta_archivo = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Archivos Excel", "*.xlsx")],
            title="Guardar archivo de saldo padres empresariales",
            initialfile="Saldo_PadresEmpresariales.xlsx"
        )

        if not ruta_archivo:
            return

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Saldo Padres Empresariales"

        encabezados = ["Padre Empresarial", "Órdenes de Compra", "Comisión Total", "Pagado", "Saldo"]
        for col, encabezado in enumerate(encabezados, 1):
            sheet.cell(row=1, column=col).value = encabezado

        for fila_idx, fila in enumerate(datos_exportacion, 2):
            for col_idx, valor in enumerate(fila, 1):
                cell = sheet.cell(row=fila_idx, column=col_idx)
                cell.value = valor
                if col_idx >= 3:  # Comisión, Pagado, Saldo
                    cell.number_format = '#,##0.00'

        workbook.save(ruta_archivo)
        messagebox.showinfo("Éxito", f"Archivo guardado exitosamente en:\n{ruta_archivo}")

    except Exception as e:
        messagebox.showerror("Error", f"No se pudo exportar a Excel: {e}")

def consultar_cuentas_por_pagar_padres():
    try:
        ventana_resultados = tk.Toplevel(root)
        ventana_resultados.title("Consultar Cuentas por Pagar - Padres Empresariales")
        ventana_resultados.geometry("1920x1080")

        # === FILTROS ===
        frame_filtros = ttk.Frame(ventana_resultados)
        frame_filtros.pack(pady=10)

        ttk.Label(frame_filtros, text="🔍 Nombre del Padre Empresarial:").grid(row=0, column=0, padx=5)
        entry_nombre = ttk.Entry(frame_filtros, width=30)
        entry_nombre.grid(row=0, column=1, padx=5)

        ttk.Label(frame_filtros, text="📊 Estado:").grid(row=0, column=2, padx=5)
        combo_estado = ttk.Combobox(frame_filtros, values=["Mostrar todas", "Pagadas (Saldo = 0)", "Por pagar (Saldo > 0)"], state="readonly", width=30)
        combo_estado.grid(row=0, column=3, padx=5)
        combo_estado.current(0)

        # === TREEVIEW ===
        frame_tree = ttk.Frame(ventana_resultados)
        frame_tree.pack(fill=tk.BOTH, expand=True)

        columnas = ["NumOrdenCompra", "NombrePadreEmpresarial", "ComisionTotal", "Pagado", "Saldo"]
        scrollbar_x = ttk.Scrollbar(frame_tree, orient=tk.HORIZONTAL)
        scrollbar_y = ttk.Scrollbar(frame_tree, orient=tk.VERTICAL)

        tree = ttk.Treeview(
            frame_tree, columns=columnas, show="headings",
            xscrollcommand=scrollbar_x.set, yscrollcommand=scrollbar_y.set
        )
        scrollbar_x.config(command=tree.xview)
        scrollbar_y.config(command=tree.yview)
        scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)
        scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)

        for col in columnas:
            tree.heading(col, text=col, anchor="center")
            tree.column(col, width=250, anchor="center")

        tree.pack(fill=tk.BOTH, expand=True)

        # === TEXTAREA RESULTADO ===
        frame_text = ttk.Frame(ventana_resultados)
        frame_text.pack(fill="x", padx=10, pady=10)

        text_area = tk.Text(frame_text, height=10, font=("Arial", 11), wrap="word")
        text_area.pack(fill="both", expand=True)
        text_area.config(state="disabled")

        # === EXPORTACIÓN ===
        datos_exportacion = []

        def ejecutar_consulta():
            try:
                nombre_filtro = entry_nombre.get().strip()
                estado_filtro = combo_estado.get()

                datos_exportacion.clear()
                for item in tree.get_children():
                    tree.delete(item)

                cursor.execute("""
                    SELECT 
                        oc.NumOrdenCompra,
                        oc.NombrePadreEmpresarial,
                        SUM(oc.ComisionPorPagarPadreEmpresarial) AS ComisionTotal,
                        ISNULL((
                            SELECT SUM(PagadoPadreEmpresarial)
                            FROM CuentasPorPagarPadreEmpresarial
                            WHERE NumOrdenCompra = oc.NumOrdenCompra
                        ), 0) AS Pagado
                    FROM OrdenCompra oc
                    WHERE oc.NombrePadreEmpresarial IS NOT NULL AND oc.NombrePadreEmpresarial != ''
                    GROUP BY oc.NumOrdenCompra, oc.NombrePadreEmpresarial
                    ORDER BY oc.NumOrdenCompra
                """)
                filas = cursor.fetchall()

                total_general = 0
                texto = "📢 SALDO POR PAGAR A PADRES EMPRESARIALES\n\n"

                for fila in filas:
                    num_orden, nombre, comision, pagado = fila
                    saldo = round(comision - pagado, 2)

                    # Filtro por nombre exacto (insensible a mayúsculas)
                    if nombre_filtro and nombre.lower() != nombre_filtro.lower():
                        continue

                    # Filtro por estado
                    if estado_filtro == "Pagadas (Saldo = 0)" and saldo != 0:
                        continue
                    if estado_filtro == "Por pagar (Saldo > 0)" and saldo <= 0:
                        continue

                    fila_mostrar = [str(num_orden), nombre, f"{comision:.2f}", f"{pagado:.2f}", f"{saldo:.2f}"]
                    tree.insert("", tk.END, values=fila_mostrar)
                    datos_exportacion.append(fila_mostrar)

                    texto += f"🏢 {nombre.upper()}\n🧾 Orden: {num_orden}\n💰 Comisión ${comision:.2f} - 💸 Pagado ${pagado:.2f} → 💼 Saldo ${saldo:.2f}\n\n"
                    total_general += saldo

                texto += f"💰 Total General de Saldo: ${total_general:.2f}"

                text_area.config(state="normal")
                text_area.delete("1.0", tk.END)
                text_area.insert(tk.END, texto.strip())
                text_area.config(state="disabled")

                if not datos_exportacion:
                    messagebox.showinfo("Sin resultados", "No se encontraron registros con los filtros seleccionados.")

            except Exception as e:
                messagebox.showerror("Error", f"Ocurrió un error al consultar: {e}")

        # === Botón Consultar
        btn_consultar = tk.Button(
            frame_filtros,
            text="🔍 Consultar",
            bg="lightgreen",
            font=("Arial", 10, "bold"),
            command=ejecutar_consulta
        )
        btn_consultar.grid(row=0, column=4, padx=10)

        # === Botón Exportar a Excel
        frame_botones = ttk.Frame(ventana_resultados)
        frame_botones.pack(pady=10)

        btn_exportar_excel = tk.Button(
            frame_botones,
            text="📥 Exportar a Excel",
            command=lambda: exportar_excel_padres(datos_exportacion),
            bg="lightgreen",
            font=("Arial", 10, "bold")
        )
        btn_exportar_excel.pack()

        # Ejecutar búsqueda inicial
        ejecutar_consulta()

    except Exception as e:
        messagebox.showerror("Error", f"Ocurrió un error: {e}")



btn_consultar_cxp_padre = tk.Button(
    tab_cuentas_pagar,
    text="🔍 Consultar Cuentas Por Pagar Padre Empresarial",
    command=consultar_cuentas_por_pagar_padres,
    bg="#5CD6C0",
    fg="black",
    font=("Arial", 10, "bold"),
    relief="raised",
    padx=10,
    pady=1,
    activebackground="#47B7A7",
    cursor="hand2"
)
btn_consultar_cxp_padre.grid(row=101, column=6, columnspan=3)



# ======================== Pestana prestamos / devoluciones ========================

# ======================== FORMULARIO DE PRÉSTAMOS ========================
ttk.Label(tab_prestamos, text="🔢 Código del Producto:").grid(row=0, column=0, padx=10, pady=5, sticky="w")
entry_codigo_prestamo = ttk.Entry(tab_prestamos, width=30)
entry_codigo_prestamo.grid(row=0, column=1, padx=10, pady=5, sticky="w")

ttk.Label(tab_prestamos, text="🏷️ Nombre del Producto:").grid(row=0, column=2, padx=10, pady=5, sticky="w")
entry_nombre_prestamo = ttk.Entry(tab_prestamos, width=100)
entry_nombre_prestamo.grid(row=0, column=3, padx=10, pady=5, sticky="w")

ttk.Label(tab_prestamos, text="📦 Cantidad Prestada:").grid(row=1, column=0, padx=10, pady=5, sticky="w")
entry_cantidad_prestada = ttk.Entry(tab_prestamos, width=30)
entry_cantidad_prestada.grid(row=1, column=1, padx=10, pady=5, sticky="w")

ttk.Label(tab_prestamos, text="📅 Fecha de Préstamo (YYYY-MM-DD):").grid(row=2, column=0, padx=10, pady=5, sticky="w")
entry_fecha_prestamo = DateEntry(tab_prestamos, width=30, date_pattern='yyyy-mm-dd')
entry_fecha_prestamo.grid(row=2, column=1, padx=10, pady=5, sticky="w")

ttk.Label(tab_prestamos, text="👤 Cliente:").grid(row=3, column=0, padx=10, pady=5, sticky="w")
entry_cliente_prestamo = ttk.Entry(tab_prestamos, width=30)
entry_cliente_prestamo.grid(row=3, column=1, padx=10, pady=5, sticky="w")



# Función para buscar el nombre del producto en la pestaña de prestamos
def buscar_nombre_producto_compra5(event=None):
    try:
        codigo_producto = entry_codigo_prestamo .get().strip()
        if not codigo_producto:
            entry_nombre_prestamo.delete(0, tk.END)
            return

        cursor.execute("SELECT NombreProducto FROM Productos WHERE CodigoProducto = ?", (codigo_producto,))
        resultado = cursor.fetchone()

        if resultado:
            nombre_producto = resultado[0]
            entry_nombre_prestamo.delete(0, tk.END)
            entry_nombre_prestamo.insert(0, nombre_producto)
        else:
            entry_nombre_prestamo.delete(0, tk.END)
            messagebox.showwarning("Advertencia", f"El producto con código '{codigo_producto}' no existe en la base de datos.")
    except Exception as e:
        messagebox.showerror("Error", f"Error al buscar el producto: {str(e)}")

entry_codigo_prestamo.bind("<FocusOut>", buscar_nombre_producto_compra5)

from datetime import date
def agregar_prestamo():
    try:
        codigo = entry_codigo_prestamo.get().strip()
        nombre = entry_nombre_prestamo.get().strip()
        cantidad_prestada = entry_cantidad_prestada.get().strip()
        fecha_prestamo = entry_fecha_prestamo.get_date().strftime("%Y-%m-%d")
        cliente = entry_cliente_prestamo.get().strip()

        if not codigo or not cantidad_prestada or not fecha_prestamo or not cliente:
            messagebox.showwarning("Advertencia", "Debe ingresar todos los campos requeridos.")
            return

        cantidad_prestada = int(cantidad_prestada)
        if cantidad_prestada <= 0:
            messagebox.showerror("Error", "La cantidad debe ser un número positivo.")
            return

        # Verificar que el producto existe y tiene inventario
        cursor.execute("SELECT NombreProducto, CantidadInventario, CantidadPrestada FROM Productos WHERE CodigoProducto = ?", (codigo,))
        producto = cursor.fetchone()

        if not producto:
            messagebox.showerror("Error", f"El producto con código '{codigo}' no existe.")
            return

        nombre_bd, inventario_actual, prestado_actual = producto
        if cantidad_prestada > inventario_actual:
            messagebox.showerror("Error", "No hay suficiente inventario disponible para este préstamo.")
            return

        # Actualizar inventario y prestado
        nuevo_inventario = inventario_actual - cantidad_prestada
        nuevo_prestado = prestado_actual + cantidad_prestada

        cursor.execute("""
            UPDATE Productos
            SET CantidadInventario = ?, CantidadPrestada = ?
            WHERE CodigoProducto = ?
        """, (nuevo_inventario, nuevo_prestado, codigo))

        # Insertar en Prestamos
        cursor.execute("""
            INSERT INTO Prestamos (CodigoProducto, NombreProducto, CantidadPrestadaTotal, CantidadPrestada, CantidadDevuelta, FechaPrestamo, Cliente)
            VALUES (?, ?, ?, ?, 0, ?, ?)
        """, (codigo, nombre if nombre else nombre_bd, cantidad_prestada, cantidad_prestada, fecha_prestamo, cliente))

        conn.commit()
        messagebox.showinfo("Éxito", "Préstamo registrado correctamente y stock actualizado.")

        # Limpiar campos
        entry_codigo_prestamo.delete(0, tk.END)
        entry_nombre_prestamo.delete(0, tk.END)
        entry_cantidad_prestada.delete(0, tk.END)
        entry_cliente_prestamo.delete(0, tk.END)
        from datetime import date
        entry_fecha_prestamo.set_date(date.today())


    except Exception as e:
        messagebox.showerror("Error", f"Ocurrió un error al registrar el préstamo: {str(e)}")


# Botón para registrar el préstamo
btn_agregar_prestamo = tk.Button(tab_prestamos, text="➕ Agregar Préstamo", command=agregar_prestamo,bg="lightgreen")
btn_agregar_prestamo.grid(row=4, column=0, columnspan=2, pady=10)

# Campo ID Prestamo
ttk.Label(tab_prestamos, text="🆔 IdPrestamo:").grid(row=5, column=0, padx=10, pady=10, sticky="w")
entry_id_prestamo = ttk.Entry(tab_prestamos, width=20)
entry_id_prestamo.grid(row=5, column=1, padx=10, pady=10, sticky="w")

# Frame para botones
frame_botones_prestamos = ttk.Frame(tab_prestamos)
frame_botones_prestamos.grid(row=5, column=2, columnspan=3, padx=10, pady=10, sticky="w")


def cargar_datos_prestamo():
    try:
        id_prestamo = entry_id_prestamo.get().strip()
        if not id_prestamo:
            messagebox.showerror("Error", "Debe ingresar un IdPrestamo válido.")
            return

        cursor.execute("""
            SELECT CodigoProducto, NombreProducto, CantidadPrestada, FechaPrestamo, Cliente
            FROM Prestamos
            WHERE IdPrestamo = ?
        """, (id_prestamo,))
        prestamo = cursor.fetchone()

        if not prestamo:
            messagebox.showerror("Error", "No se encontró el préstamo con ese ID.")
            return

        codigo, nombre, cantidad, fecha, cliente = prestamo

        entry_codigo_prestamo.delete(0, tk.END)
        entry_codigo_prestamo.insert(0, codigo)
        entry_nombre_prestamo.delete(0, tk.END)
        entry_nombre_prestamo.insert(0, nombre)
        entry_cantidad_prestada.delete(0, tk.END)
        entry_cantidad_prestada.insert(0, str(cantidad))
        entry_fecha_prestamo.set_date(fecha)
        entry_cliente_prestamo.delete(0, tk.END)
        entry_cliente_prestamo.insert(0, cliente)

        messagebox.showinfo("Éxito", "Datos del préstamo cargados correctamente.")

    except Exception as e:
        messagebox.showerror("Error", f"Error al cargar datos del préstamo: {str(e)}")


btn_cargar_prestamo = tk.Button(frame_botones_prestamos, bg="lightgreen", text="🔄 Cargar Datos", command=cargar_datos_prestamo)
btn_cargar_prestamo.pack(side=tk.LEFT, padx=5)

def modificar_prestamo():
    try:
        id_prestamo = entry_id_prestamo.get().strip()
        if not id_prestamo:
            messagebox.showerror("Error", "Debe ingresar un IdPrestamo para modificar.")
            return

        # ✅ VALIDACIÓN: ¿Existen devoluciones con este IdPrestamo?
        cursor.execute("""
            SELECT IdDevolucion FROM Devoluciones
            WHERE IdPrestamo = ?
        """, (id_prestamo,))
        devoluciones_asociadas = cursor.fetchall()

        if devoluciones_asociadas:
            ids_bloqueo = [str(d[0]) for d in devoluciones_asociadas]
            ids_str = ", ".join(ids_bloqueo)
            messagebox.showwarning(
                "Modificación Bloqueada",
                f"No se puede modificar este préstamo porque existen devoluciones asociadas.\n"
                f"Debe eliminar primero las siguientes devoluciones:\n\nID(s): {ids_str}"
            )

            # 🧹 LIMPIAR CAMPOS TRAS BLOQUEO
            entry_id_prestamo.delete(0, tk.END)
            entry_codigo_prestamo.delete(0, tk.END)
            entry_nombre_prestamo.delete(0, tk.END)
            entry_cantidad_prestada.delete(0, tk.END)
            entry_cliente_prestamo.delete(0, tk.END)
            from datetime import date
            entry_fecha_prestamo.set_date(date.today())

            return

        # 🔁 CONTINUAR CON MODIFICACIÓN
        codigo = entry_codigo_prestamo.get().strip()
        nombre = entry_nombre_prestamo.get().strip()
        nueva_cantidad_total = int(entry_cantidad_prestada.get().strip())
        fecha = entry_fecha_prestamo.get_date().strftime("%Y-%m-%d")
        cliente = entry_cliente_prestamo.get().strip()

        if nueva_cantidad_total <= 0:
            messagebox.showerror("Error", "La cantidad debe ser mayor a cero.")
            return

        # Obtener datos actuales del préstamo
        cursor.execute("""
            SELECT CodigoProducto, CantidadPrestadaTotal, CantidadPrestada, CantidadDevuelta
            FROM Prestamos
            WHERE IdPrestamo = ?
        """, (id_prestamo,))
        original = cursor.fetchone()

        if not original:
            messagebox.showerror("Error", "No se encontró el préstamo.")
            return

        codigo_original, total_anterior, pendiente_anterior, devuelto = original

        if nueva_cantidad_total < devuelto:
            messagebox.showerror("Error", f"No puede prestar menos de lo ya devuelto ({devuelto} unidades).")
            return

        nueva_pendiente = nueva_cantidad_total - devuelto
        diferencia_total = nueva_cantidad_total - total_anterior

        # Obtener inventario actual del producto
        cursor.execute("""
            SELECT CantidadInventario, CantidadPrestada
            FROM Productos
            WHERE CodigoProducto = ?
        """, (codigo,))
        inventario_actual, prestado_actual = cursor.fetchone()

        if diferencia_total > 0 and diferencia_total > inventario_actual:
            messagebox.showerror("Error", "No hay suficiente inventario para aumentar el préstamo.")
            return

        nuevo_inventario = inventario_actual - diferencia_total
        nuevo_prestado_productos = prestado_actual + diferencia_total

        # Actualizar tabla Productos
        cursor.execute("""
            UPDATE Productos
            SET CantidadInventario = ?, CantidadPrestada = ?
            WHERE CodigoProducto = ?
        """, (nuevo_inventario, nuevo_prestado_productos, codigo))

        # Actualizar tabla Prestamos
        cursor.execute("""
            UPDATE Prestamos
            SET CodigoProducto = ?, NombreProducto = ?, CantidadPrestadaTotal = ?, CantidadPrestada = ?, FechaPrestamo = ?, Cliente = ?
            WHERE IdPrestamo = ?
        """, (codigo, nombre, nueva_cantidad_total, nueva_pendiente, fecha, cliente, id_prestamo))

        conn.commit()
        messagebox.showinfo("Éxito", "Préstamo modificado correctamente.")

        # ✅ LIMPIAR CAMPOS
        entry_id_prestamo.delete(0, tk.END)
        entry_codigo_prestamo.delete(0, tk.END)
        entry_nombre_prestamo.delete(0, tk.END)
        entry_cantidad_prestada.delete(0, tk.END)
        entry_cliente_prestamo.delete(0, tk.END)
        from datetime import date
        entry_fecha_prestamo.set_date(date.today())


    except Exception as e:
        messagebox.showerror("Error", f"Error al modificar el préstamo: {str(e)}")




    


btn_modificar_prestamo = tk.Button(frame_botones_prestamos,bg="skyblue", text="✏️ Modificar Préstamo", command=modificar_prestamo)
btn_modificar_prestamo.pack(side=tk.LEFT, padx=5)


def eliminar_prestamo():
    try:
        id_prestamo = entry_id_prestamo.get().strip()
        if not id_prestamo:
            messagebox.showerror("Error", "Debe ingresar un IdPrestamo para eliminar.")
            return

        # === Obtener datos del préstamo ===
        cursor.execute("""
            SELECT CodigoProducto, CantidadPrestada, CantidadPrestadaTotal, Cliente
            FROM Prestamos WHERE IdPrestamo = ?
        """, (id_prestamo,))
        resultado = cursor.fetchone()

        if not resultado:
            messagebox.showerror("Error", "No se encontró el préstamo.")
            return

        codigo_producto, cantidad_pendiente, cantidad_total, cliente = resultado
        cantidad_devuelta = cantidad_total - cantidad_pendiente

        # === Buscar devoluciones asociadas al préstamo ===
        cursor.execute("""
            SELECT IdDevolucion, CantidadDevuelta
            FROM Devoluciones
            WHERE IdPrestamo = ?
        """, (id_prestamo,))
        devoluciones = cursor.fetchall()

        total_devueltas_encontradas = sum(dev[1] for dev in devoluciones)

        if total_devueltas_encontradas > cantidad_devuelta:
            messagebox.showerror("Error", f"Las devoluciones encontradas ({total_devueltas_encontradas}) superan lo devuelto registrado ({cantidad_devuelta}).")
            return

        # === Obtener estado actual del producto ===
        cursor.execute("""
            SELECT CantidadInventario, CantidadPrestada
            FROM Productos
            WHERE CodigoProducto = ?
        """, (codigo_producto,))
        producto = cursor.fetchone()

        if not producto:
            messagebox.showerror("Error", "El producto ya no existe.")
            return

        inventario_actual, prestado_actual = producto
        nuevo_inventario = inventario_actual + cantidad_pendiente
        nuevo_prestado = max(prestado_actual - cantidad_total, 0)

        # === Ventana de confirmación ===
        confirm_window = tk.Toplevel()
        confirm_window.title("Confirmar eliminación del préstamo")

        tk.Label(confirm_window, text=f"¿Deseas eliminar el préstamo #{id_prestamo}?", font=("Arial", 10, "bold")).pack(pady=5)
        tk.Label(confirm_window, text=f"📦 Producto: {codigo_producto}   |   👤 Cliente: {cliente}", font=("Arial", 10)).pack()
        tk.Label(confirm_window, text=f"\nDevoluciones asociadas a este préstamo: {len(devoluciones)}", font=("Arial", 10)).pack()

        entries_ids = []

        if devoluciones:
            for idx, (iddev, cantidad) in enumerate(devoluciones, start=1):
                var = tk.BooleanVar()
                frame = ttk.Frame(confirm_window)
                frame.pack(pady=2)
                ttk.Checkbutton(frame, text=f"ID Devolución #{iddev} (Cantidad {cantidad}):", variable=var).pack(anchor="w")
                entries_ids.append((iddev, var, cantidad))

        def confirmar_eliminacion():
            if devoluciones:
                ids_a_eliminar = [iddev for iddev, var, _ in entries_ids if var.get()]
                suma_eliminar = sum(cant for _, var, cant in entries_ids if var.get())

                if not ids_a_eliminar:
                    messagebox.showinfo("Cancelado", "No se seleccionó ninguna devolución. Operación cancelada.")
                    confirm_window.destroy()
                    return

                if suma_eliminar > cantidad_devuelta:
                    messagebox.showerror("Error", f"La suma seleccionada ({suma_eliminar}) excede lo devuelto en el préstamo ({cantidad_devuelta}).")
                    return

                for id_dev in ids_a_eliminar:
                    cursor.execute("DELETE FROM Devoluciones WHERE IdDevolucion = ?", (id_dev,))
            # === Actualizar producto ===
            cursor.execute("""
                UPDATE Productos
                SET CantidadInventario = ?, CantidadPrestada = ?
                WHERE CodigoProducto = ?
            """, (nuevo_inventario, nuevo_prestado, codigo_producto))

            # === Eliminar el préstamo ===
            cursor.execute("DELETE FROM Prestamos WHERE IdPrestamo = ?", (id_prestamo,))
            conn.commit()

            messagebox.showinfo("Éxito", f"Préstamo #{id_prestamo} eliminado correctamente.")
            confirm_window.destroy()

            # 🧹 Limpiar campos visuales
            entry_id_prestamo.delete(0, tk.END)
            entry_codigo_prestamo.delete(0, tk.END)
            entry_nombre_prestamo.delete(0, tk.END)
            entry_cantidad_prestada.delete(0, tk.END)
            entry_cliente_prestamo.delete(0, tk.END)
            from datetime import date
            entry_fecha_prestamo.set_date(date.today())


        ttk.Button(confirm_window, text="Aceptar", command=confirmar_eliminacion).pack(side=tk.LEFT, padx=20, pady=10)
        ttk.Button(confirm_window, text="Cancelar", command=confirm_window.destroy).pack(side=tk.RIGHT, padx=20, pady=10)
        confirm_window.geometry("550x300")
        confirm_window.transient(root)
        confirm_window.grab_set()

    except Exception as e:
        messagebox.showerror("Error", f"Error al eliminar el préstamo: {str(e)}")




btn_eliminar_prestamo = tk.Button(frame_botones_prestamos,bg="tomato", text="🗑️ Eliminar Préstamo", command=eliminar_prestamo)
btn_eliminar_prestamo.pack(side=tk.LEFT, padx=5)



# ======================== SECCIÓN DEVOLUCIÓN ========================
# Activador de sección
var_devolucion_activa = tk.BooleanVar()

frame_titulo_devolucion = ttk.Frame(tab_prestamos)
frame_titulo_devolucion.grid(row=6, column=0, columnspan=4, padx=10, pady=15, sticky="w")

ttk.Label(frame_titulo_devolucion, text="📥 Devolución", font=("Arial", 12, "bold")).pack(side="left", padx=(0, 10))
check_activar_devolucion = ttk.Checkbutton(frame_titulo_devolucion, variable=var_devolucion_activa, command=lambda: habilitar_devolucion(var_devolucion_activa.get()))
check_activar_devolucion.pack(side="left")

# Campos de devolución (inicialmente deshabilitados)
ttk.Label(tab_prestamos, text="📜 Orden de Compra:").grid(row=7, column=0, padx=10, pady=5, sticky="w")
entry_orden_devolucion = ttk.Entry(tab_prestamos, width=30, state="disabled")
entry_orden_devolucion.grid(row=7, column=1, padx=10, pady=5, sticky="w")

ttk.Label(tab_prestamos, text="📦 Cantidad Devolución:").grid(row=8, column=0, padx=10, pady=5, sticky="w")
entry_cantidad_devolucion = ttk.Entry(tab_prestamos, width=30, state="disabled")
entry_cantidad_devolucion.grid(row=8, column=1, padx=10, pady=5, sticky="w")

ttk.Label(tab_prestamos, text="📅 Fecha Devolución (YYYY-MM-DD):").grid(row=9, column=0, padx=10, pady=5, sticky="w")
entry_fecha_devolucion = DateEntry(tab_prestamos, width=30, date_pattern='yyyy-mm-dd', state="disabled")
entry_fecha_devolucion.grid(row=9, column=1, padx=10, pady=5, sticky="w")



# ID y botones
ttk.Label(tab_prestamos, text="🆔 IdDevolucion:").grid(row=11, column=0, padx=10, pady=10, sticky="w")
entry_id_devolucion = ttk.Entry(tab_prestamos, width=20)
entry_id_devolucion.grid(row=11, column=1, padx=10, pady=10, sticky="w")

frame_botones_devolucion = ttk.Frame(tab_prestamos)
frame_botones_devolucion.grid(row=11, column=2, columnspan=3, padx=10, pady=10, sticky="w")




def habilitar_devolucion(activa):
    if activa:
        estado = "normal"
    else:
        estado = "disabled"

    entry_orden_devolucion.config(state=estado)
    entry_cantidad_devolucion.config(state=estado)
    entry_fecha_devolucion.config(state=estado)
    btn_agregar_devolucion.config(state=estado)
    entry_id_devolucion.config(state=estado)
    btn_cargar_devolucion.config(state=estado)
    btn_modificar_devolucion.config(state=estado)
    btn_eliminar_devolucion.config(state=estado)



def agregar_devolucion():
    try:
        if not var_devolucion_activa.get():
            messagebox.showwarning("Advertencia", "Debe activar la sección de Devolución para continuar.")
            return

        id_prestamo = entry_id_prestamo.get().strip()
        if not id_prestamo:
            messagebox.showerror("Error", "Debe cargar primero un IdPrestamo válido.")
            return

        orden = entry_orden_devolucion.get().strip()
        cantidad_devolver = entry_cantidad_devolucion.get().strip()
        fecha = entry_fecha_devolucion.get_date().strftime("%Y-%m-%d")

        if not orden or not cantidad_devolver or not fecha:
            messagebox.showerror("Error", "Todos los campos de devolución son obligatorios.")
            return

        cantidad_devolver = int(cantidad_devolver)
        if cantidad_devolver <= 0:
            messagebox.showerror("Error", "La cantidad devuelta debe ser mayor que 0.")
            return

        # Obtener datos del préstamo incluyendo Cliente
        cursor.execute("""
            SELECT CodigoProducto, CantidadPrestadaTotal, CantidadPrestada, CantidadDevuelta, Cliente
            FROM Prestamos WHERE IdPrestamo = ?
        """, (id_prestamo,))
        prestamo = cursor.fetchone()
        if not prestamo:
            messagebox.showerror("Error", "No se encontró el préstamo con ese ID.")
            return

        codigo_producto, total, pendiente, devuelto, cliente = prestamo

        if cantidad_devolver > pendiente:
            messagebox.showerror("Error", f"No se puede devolver más de lo pendiente ({pendiente} unidades).")
            return

        # Obtener datos del producto
        cursor.execute("""
            SELECT CantidadInventario, CantidadPrestada
            FROM Productos WHERE CodigoProducto = ?
        """, (codigo_producto,))
        producto = cursor.fetchone()
        if not producto:
            messagebox.showerror("Error", "Producto no encontrado en tabla Productos.")
            return

        inventario, prestado_prod = producto

        # Actualizar tabla Productos
        cursor.execute("""
            UPDATE Productos SET CantidadInventario = ?, CantidadPrestada = ?
            WHERE CodigoProducto = ?
        """, (inventario + cantidad_devolver, prestado_prod - cantidad_devolver, codigo_producto))

        # Actualizar tabla Prestamos
        cursor.execute("""
            UPDATE Prestamos SET CantidadPrestada = ?, CantidadDevuelta = ?
            WHERE IdPrestamo = ?
        """, (pendiente - cantidad_devolver, devuelto + cantidad_devolver, id_prestamo))

        # ✅ Insertar en Devoluciones (ahora incluye IdPrestamo)
        cursor.execute("""
            INSERT INTO Devoluciones (OrdenCompra, CodigoProducto, CantidadDevuelta, FechaDevolucion, Cliente, IdPrestamo)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (orden, codigo_producto, cantidad_devolver, fecha, cliente, id_prestamo))

        conn.commit()
        messagebox.showinfo("Éxito", "Devolución registrada correctamente.")

        # Limpiar campos
        entry_orden_devolucion.delete(0, tk.END)
        entry_cantidad_devolucion.delete(0, tk.END)
        entry_fecha_devolucion.set_date(date.today())
        entry_id_prestamo.delete(0, tk.END)

    except Exception as e:
        messagebox.showerror("Error", f"Ocurrió un error: {str(e)}")





btn_agregar_devolucion = tk.Button(tab_prestamos, text="➕ Agregar Devolución", state="disabled", command=agregar_devolucion,bg="lightgreen")
btn_agregar_devolucion.grid(row=10, column=0, columnspan=2, pady=10)

def cargar_datos_devolucion():
    try:
        id_devolucion = entry_id_devolucion.get().strip()
        if not id_devolucion:
            messagebox.showerror("Error", "Debe ingresar un IdDevolucion válido.")
            return

        cursor.execute("""
            SELECT OrdenCompra, CantidadDevuelta, FechaDevolucion, IdPrestamo
            FROM Devoluciones WHERE IdDevolucion = ?
        """, (id_devolucion,))
        devolucion = cursor.fetchone()

        if not devolucion:
            messagebox.showerror("Error", "No se encontró la devolución con ese ID.")
            return

        orden, cantidad, fecha, id_prestamo = devolucion
        var_devolucion_activa.set(True)
        habilitar_devolucion(True)

        entry_orden_devolucion.delete(0, tk.END)
        entry_orden_devolucion.insert(0, orden)
        entry_cantidad_devolucion.delete(0, tk.END)
        entry_cantidad_devolucion.insert(0, str(cantidad))
        entry_fecha_devolucion.set_date(fecha)

        messagebox.showinfo("Éxito", f"Datos cargados correctamente.\nIdPréstamo asociado: {id_prestamo}")

    except Exception as e:
        messagebox.showerror("Error", f"Error al cargar devolución: {str(e)}")


btn_cargar_devolucion = tk.Button(frame_botones_devolucion,bg="lightgreen", text="🔄 Cargar Datos",  command=cargar_datos_devolucion)
btn_cargar_devolucion.pack(side=tk.LEFT, padx=5)


def modificar_devolucion():
    try:
        id_devolucion = entry_id_devolucion.get().strip()
        if not id_devolucion:
            messagebox.showerror("Error", "Debe ingresar el IdDevolucion para modificar.")
            return

        orden = entry_orden_devolucion.get().strip()
        nueva_cantidad = int(entry_cantidad_devolucion.get().strip())
        nueva_fecha = entry_fecha_devolucion.get_date().strftime("%Y-%m-%d")

        if nueva_cantidad <= 0:
            messagebox.showerror("Error", "La cantidad devuelta debe ser mayor que cero.")
            return

        # Obtener datos de la devolución original
        cursor.execute("""
            SELECT CodigoProducto, CantidadDevuelta, IdPrestamo, Cliente
            FROM Devoluciones
            WHERE IdDevolucion = ?
        """, (id_devolucion,))
        devolucion = cursor.fetchone()
        if not devolucion:
            messagebox.showerror("Error", "No se encontró la devolución.")
            return

        codigo_producto, cantidad_anterior, id_prestamo, cliente = devolucion
        diferencia = nueva_cantidad - cantidad_anterior

        # Obtener datos actuales del producto
        cursor.execute("""
            SELECT CantidadInventario, CantidadPrestada FROM Productos
            WHERE CodigoProducto = ?
        """, (codigo_producto,))
        inventario, prestado = cursor.fetchone()

        # Obtener préstamo original
        cursor.execute("""
            SELECT CantidadPrestada, CantidadDevuelta
            FROM Prestamos
            WHERE IdPrestamo = ?
        """, (id_prestamo,))
        prestamo = cursor.fetchone()
        if not prestamo:
            messagebox.showerror("Error", "No se encontró el préstamo relacionado.")
            return

        cantidad_pendiente, cantidad_devuelta = prestamo

        if diferencia > 0 and diferencia > cantidad_pendiente:
            messagebox.showerror("Error", "No hay suficiente cantidad pendiente para aumentar esta devolución.")
            return

        # Recalcular
        nuevo_inventario = inventario + diferencia
        nuevo_prestado = prestado - diferencia
        nuevo_pendiente = cantidad_pendiente - diferencia
        nuevo_devuelto = cantidad_devuelta + diferencia

        # Actualizar Productos
        cursor.execute("""
            UPDATE Productos
            SET CantidadInventario = ?, CantidadPrestada = ?
            WHERE CodigoProducto = ?
        """, (nuevo_inventario, nuevo_prestado, codigo_producto))

        # Actualizar Prestamo
        cursor.execute("""
            UPDATE Prestamos
            SET CantidadPrestada = ?, CantidadDevuelta = ?
            WHERE IdPrestamo = ?
        """, (nuevo_pendiente, nuevo_devuelto, id_prestamo))

        # Actualizar Devolución
        cursor.execute("""
            UPDATE Devoluciones
            SET OrdenCompra = ?, CantidadDevuelta = ?, FechaDevolucion = ?, Cliente = ?
            WHERE IdDevolucion = ?
        """, (orden, nueva_cantidad, nueva_fecha, cliente, id_devolucion))

        conn.commit()
        messagebox.showinfo("Éxito", f"Devolución #{id_devolucion} modificada correctamente.")

        # Limpiar campos
        entry_id_devolucion.delete(0, tk.END)
        entry_orden_devolucion.delete(0, tk.END)
        entry_cantidad_devolucion.delete(0, tk.END)
        entry_fecha_devolucion.set_date(date.today())


    except Exception as e:
        messagebox.showerror("Error", f"Error al modificar devolución: {str(e)}")





btn_modificar_devolucion = tk.Button(frame_botones_devolucion,bg="skyblue" ,text="✏️ Modificar Devolución",  command=modificar_devolucion)
btn_modificar_devolucion.pack(side=tk.LEFT, padx=5)


def eliminar_devolucion():
    try:
        id_devolucion = entry_id_devolucion.get().strip()
        if not id_devolucion:
            messagebox.showerror("Error", "Debe ingresar el IdDevolucion para eliminar.")
            return

        # === Obtener datos de la devolución, incluyendo IdPrestamo ===
        cursor.execute("""
            SELECT CodigoProducto, CantidadDevuelta, Cliente, IdPrestamo
            FROM Devoluciones
            WHERE IdDevolucion = ?
        """, (id_devolucion,))
        devolucion = cursor.fetchone()

        if not devolucion:
            messagebox.showerror("Error", "No se encontró la devolución.")
            return

        codigo_producto, cantidad_devuelta, cliente, id_prestamo = devolucion
        cantidad_devuelta = int(cantidad_devuelta)

        if not id_prestamo:
            messagebox.showerror("Error", "La devolución no está asociada a ningún préstamo. No se puede continuar.")
            return

        # === Obtener datos del producto ===
        cursor.execute("""
            SELECT CantidadInventario, CantidadPrestada
            FROM Productos
            WHERE CodigoProducto = ?
        """, (codigo_producto,))
        producto = cursor.fetchone()
        if not producto:
            messagebox.showerror("Error", "No se encontró el producto.")
            return

        inventario_actual, prestado_actual = producto

        # === Obtener datos del préstamo original ===
        cursor.execute("""
            SELECT CantidadPrestada, CantidadDevuelta
            FROM Prestamos
            WHERE IdPrestamo = ?
        """, (id_prestamo,))
        prestamo = cursor.fetchone()
        if not prestamo:
            messagebox.showerror("Error", "No se encontró el préstamo relacionado.")
            return

        cantidad_pendiente, cantidad_devuelta_total = prestamo

        # === Validar que se pueda revertir la devolución ===
        if cantidad_devuelta > cantidad_devuelta_total:
            messagebox.showerror("Error", "La cantidad devuelta supera la registrada en el préstamo.")
            return

        # === Revertir valores ===
        nuevo_inventario = max(0, inventario_actual - cantidad_devuelta)
        nuevo_prestado = prestado_actual + cantidad_devuelta
        nuevo_pendiente = cantidad_pendiente + cantidad_devuelta
        nuevo_devuelto = max(0, cantidad_devuelta_total - cantidad_devuelta)

        # === Actualizar Productos ===
        cursor.execute("""
            UPDATE Productos
            SET CantidadInventario = ?, CantidadPrestada = ?
            WHERE CodigoProducto = ?
        """, (nuevo_inventario, nuevo_prestado, codigo_producto))

        # === Actualizar Prestamos ===
        cursor.execute("""
            UPDATE Prestamos
            SET CantidadPrestada = ?, CantidadDevuelta = ?
            WHERE IdPrestamo = ?
        """, (nuevo_pendiente, nuevo_devuelto, id_prestamo))

        # === Eliminar devolución ===
        cursor.execute("""
            DELETE FROM Devoluciones
            WHERE IdDevolucion = ?
        """, (id_devolucion,))

        conn.commit()

        messagebox.showinfo("Éxito", f"Devolución #{id_devolucion} eliminada correctamente.\nValores revertidos al préstamo #{id_prestamo}.")

        # === Limpiar campos ===
        entry_id_devolucion.delete(0, tk.END)
        entry_orden_devolucion.delete(0, tk.END)
        entry_cantidad_devolucion.delete(0, tk.END)
        entry_fecha_devolucion.set_date(date.today())

    except Exception as e:
        messagebox.showerror("Error", f"Error al eliminar la devolución: {str(e)}")




btn_eliminar_devolucion = tk.Button(frame_botones_devolucion,bg="tomato", text="🗑️ Eliminar Devolución",  command=eliminar_devolucion)
btn_eliminar_devolucion.pack(side=tk.LEFT, padx=5)



# ---------------------- PESTAÑA Excel pdf ----------------------
def exportar_excel(datos, columnas):
    import pandas as pd
    from tkinter import filedialog
    from openpyxl.utils import get_column_letter
    from openpyxl import load_workbook
    from openpyxl.styles import numbers

    # Convertir strings numéricos/fechas a tipos reales
    def convertir_valor(valor):
        if isinstance(valor, str):
            valor = valor.strip()
            # Probar si es fecha en formato YYYY-MM-DD
            try:
                if "-" in valor and len(valor) == 10:
                    return pd.to_datetime(valor).date()
            except:
                pass
            # Probar si es numérico
            try:
                if "." in valor:
                    return float(valor)
                else:
                    return int(valor)
            except ValueError:
                return valor
        return valor

    # Convertir datos
    datos_convertidos = [[convertir_valor(celda) for celda in fila] for fila in datos]

    # Crear DataFrame
    df = pd.DataFrame(datos_convertidos, columns=columnas)

    # Guardar archivo
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        df.to_excel(file_path, index=False)

        # Formato profesional
        wb = load_workbook(file_path)
        ws = wb.active

        for col_idx, col_cells in enumerate(ws.columns, start=1):
            max_length = 0
            for cell in col_cells:
                if cell.value is not None:
                    # Ajustar ancho de columna
                    max_length = max(max_length, len(str(cell.value)))

                    # Aplicar formato si es numérico o fecha
                    if isinstance(cell.value, int):
                        cell.number_format = numbers.FORMAT_NUMBER
                    elif isinstance(cell.value, float):
                        cell.number_format = numbers.FORMAT_NUMBER_00  # dos decimales
                    elif isinstance(cell.value, (pd.Timestamp, pd._libs.tslibs.timestamps.Timestamp, type(pd.to_datetime("2020-01-01").date()))):
                        cell.number_format = numbers.FORMAT_DATE_YYYYMMDD2

            # Autoajustar ancho
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max_length + 2

        wb.save(file_path)

def exportar_excel2(datos, columnas):
    import pandas as pd
    from tkinter import filedialog
    from openpyxl.utils import get_column_letter
    from openpyxl import load_workbook
    from openpyxl.styles import numbers

    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if not file_path:
        return

    # Crear DataFrame sin conversiones extra
    df = pd.DataFrame(datos, columns=columnas)

    # Forzar columnas numéricas si se detectan como object
    numericas = ["CantidadComprada", "CostoSinIVA", "PorcentajeIVA", "IVA", "CostoConIVA"]
    for col in numericas:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')

    # Forzar fechas si aplica
    if "FechaCompra" in df.columns:
        df["FechaCompra"] = pd.to_datetime(df["FechaCompra"], errors="coerce")

    df.to_excel(file_path, index=False)

    # Ajuste de formatos en Excel
    wb = load_workbook(file_path)
    ws = wb.active

    tipo_columna = {
        "CantidadComprada": "int",
        "CostoSinIVA": "float",
        "PorcentajeIVA": "percent",
        "IVA": "float",
        "CostoConIVA": "float",
        "FechaCompra": "date"
    }

    for col_idx, col_cells in enumerate(ws.columns, start=1):
        col_name = ws.cell(row=1, column=col_idx).value
        max_length = len(str(col_name)) if col_name else 0

        for cell in col_cells:
            if cell.row == 1:
                continue
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
                tipo = tipo_columna.get(col_name)
                if tipo == "int":
                    cell.number_format = numbers.FORMAT_NUMBER
                elif tipo == "float":
                    cell.number_format = numbers.FORMAT_NUMBER_00
                elif tipo == "percent":
                    cell.number_format = "0.00%"
                elif tipo == "date":
                    cell.number_format = numbers.FORMAT_DATE_YYYYMMDD2

        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(file_path)





def exportar_pdf(datos, columnas, titulo_reporte="Reporte"):
    from reportlab.lib.pagesizes import landscape, letter
    from reportlab.platypus import Table, TableStyle, SimpleDocTemplate, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from tkinter import filedialog
    from datetime import datetime

    file_path = filedialog.asksaveasfilename(
        title="Guardar como...",
        defaultextension=".pdf",
        filetypes=[("PDF files", "*.pdf")],
        initialfile=f"{titulo_reporte.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf"
    )
    if not file_path:
        return

    doc = SimpleDocTemplate(
        file_path,
        pagesize=landscape(letter),
        leftMargin=1 * cm,
        rightMargin=1 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm
    )

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="ReportTitle", alignment=TA_CENTER, fontSize=14, leading=18, fontName="Helvetica-Bold"))
    styles.add(ParagraphStyle(name="HeaderCell", alignment=TA_CENTER, fontSize=7.5, leading=9, textColor=colors.white, splitLongWords=False))
    styles.add(ParagraphStyle(name="Cell", fontSize=7, leading=9, wordWrap='CJK'))

    contenido = [
        Paragraph(f"📊 {titulo_reporte}", styles["ReportTitle"]),
        Paragraph(f"Fecha de generación: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles["Normal"]),
        Spacer(1, 0.3 * cm)
    ]

    total_columnas = len(columnas)
    ancho_total = 27 * cm
    ancho_min = 2.5 * cm
    ancho_max = 6.5 * cm

    # 📐 Calcular número máximo de columnas que caben por página
    max_columnas_por_pagina = int(ancho_total // ancho_min)
    if total_columnas <= max_columnas_por_pagina:
        bloques = [(0, total_columnas)]
    else:
        bloques = [
            (i, min(i + max_columnas_por_pagina, total_columnas))
            for i in range(0, total_columnas, max_columnas_por_pagina)
        ]

    for bloque_index, (inicio, fin) in enumerate(bloques):
        columnas_bloque = columnas[inicio:fin]
        datos_bloque = [
            [Paragraph(str(col), styles["HeaderCell"]) for col in columnas_bloque]
        ]
        for fila in datos:
            fila_bloque = [Paragraph(str(fila[i]), styles["Cell"]) for i in range(inicio, fin)]
            datos_bloque.append(fila_bloque)

        ancho_col = ancho_total / len(columnas_bloque)
        col_widths = [ancho_col] * len(columnas_bloque)

        if len(bloques) > 1:
            contenido.append(Spacer(1, 0.3 * cm))
            contenido.append(Paragraph(f"📄 Bloque {bloque_index + 1} de {len(bloques)} (Columnas {inicio + 1}–{fin})", styles["Normal"]))

        tabla = Table(datos_bloque, repeatRows=1, colWidths=col_widths, splitByRow=True)
        tabla.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#003366")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey]),
        ]))
        contenido.append(tabla)

        if bloque_index < len(bloques) - 1:
            contenido.append(PageBreak())

    def pie(canvas, doc):
        canvas.saveState()
        canvas.setFont('Helvetica', 8)
        canvas.drawString(1 * cm, 1 * cm, f"Página {doc.page}")
        canvas.drawRightString(landscape(letter)[0] - 1 * cm, 1 * cm, f"Sistema de Gestión - {titulo_reporte}")
        canvas.restoreState()

    doc.build(contenido, onFirstPage=pie, onLaterPages=pie)


from tkinter import ttk, messagebox
import tkinter as tk
from tkcalendar import DateEntry
from datetime import datetime

def mostrar_compras():
    try:
        ventana = tk.Toplevel(root)
        ventana.title("Lista de Compras")
        ventana.geometry("1920x1080")

        # === FILTROS ===
        frame_filtros = ttk.Frame(ventana)
        frame_filtros.pack(pady=10)

        ttk.Label(frame_filtros, text="🔎 Código Producto:").grid(row=0, column=0, padx=5)
        entry_codigo = ttk.Entry(frame_filtros, width=20)
        entry_codigo.grid(row=0, column=1, padx=5)

        ttk.Label(frame_filtros, text="🏭 Proveedor:").grid(row=0, column=2, padx=5)
        entry_proveedor = ttk.Entry(frame_filtros, width=20)
        entry_proveedor.grid(row=0, column=3, padx=5)

        ttk.Label(frame_filtros, text="📅 Fecha Inicio:").grid(row=0, column=4, padx=5)
        entry_fecha_inicio = DateEntry(frame_filtros, width=15, date_pattern="yyyy-mm-dd")
        entry_fecha_inicio.set_date(datetime(2025, 1, 1).date())
        entry_fecha_inicio.grid(row=0, column=5, padx=5)

        ttk.Label(frame_filtros, text="📅 Fecha Fin:").grid(row=0, column=6, padx=5)
        entry_fecha_fin = DateEntry(frame_filtros, width=15, date_pattern="yyyy-mm-dd")
        entry_fecha_fin.set_date(datetime.today().date())
        entry_fecha_fin.grid(row=0, column=7, padx=5)

        # === TREEVIEW ===
        frame_tree = ttk.Frame(ventana)
        frame_tree.pack(fill=tk.BOTH, expand=True)

        scrollbar_x = ttk.Scrollbar(frame_tree, orient=tk.HORIZONTAL)
        scrollbar_y = ttk.Scrollbar(frame_tree, orient=tk.VERTICAL)

        columnas = [
            "IdCompra", "FechaCompra", "CodigoProducto", "NombreProducto",
            "CantidadComprada", "CostoSinIVA", "PorcentajeIVA", "IVA",
            "CostoConIVA", "Proveedor"
        ]

        tree = ttk.Treeview(
            frame_tree, columns=columnas, show="headings",
            xscrollcommand=scrollbar_x.set, yscrollcommand=scrollbar_y.set
        )
        scrollbar_x.config(command=tree.xview)
        scrollbar_y.config(command=tree.yview)
        scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)
        scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)

        for col in columnas:
            tree.heading(col, text=col)
            tree.column(col, width=450, anchor="center")
        tree.pack(fill=tk.BOTH, expand=True)

        datos_exportacion = []

        def ejecutar_consulta():
            try:
                for item in tree.get_children():
                    tree.delete(item)
                datos_exportacion.clear()

                codigo = entry_codigo.get().strip()
                proveedor = entry_proveedor.get().strip()
                fecha_inicio = entry_fecha_inicio.get_date()
                fecha_fin = entry_fecha_fin.get_date()

                query = """
                    SELECT IdCompra, FechaCompra, CodigoProducto, NombreProducto, 
                           CantidadComprada, CostoSinIVA, PorcentajeIVA, IVA, 
                           CostoConIVA, Proveedor 
                    FROM Compras 
                    WHERE FechaCompra BETWEEN ? AND ?
                """
                params = [fecha_inicio, fecha_fin]

                if codigo:
                    query += " AND CodigoProducto = ?"
                    params.append(codigo)
                if proveedor:
                    query += " AND Proveedor = ?"
                    params.append(proveedor)

                query += " ORDER BY IdCompra"

                cursor.execute(query, params)
                compras = cursor.fetchall()

                for compra in compras:
                    (
                        id_compra, fecha, cod, nombre, cantidad,
                        costo_sin_iva, porc_iva, valor_iva, costo_con_iva, prov
                    ) = compra

                    fecha_fmt = fecha.strftime("%Y-%m-%d") if fecha else ""
                    fila = (
                        id_compra, fecha_fmt, cod, nombre, cantidad,
                        costo_sin_iva, porc_iva, valor_iva, costo_con_iva, prov
                    )
                    tree.insert("", tk.END, values=fila)
                    datos_exportacion.append(list(fila))

                if not compras:
                    messagebox.showinfo("Sin resultados", "No se encontraron compras con los filtros aplicados.")

            except Exception as err:
                messagebox.showerror("Error", f"No se pudo realizar la consulta: {err}")

        # Botón consultar
        btn_consultar = ttk.Button(
            frame_filtros, text="🔍 Consultar", command=ejecutar_consulta
        )
        btn_consultar.grid(row=0, column=8, padx=10)

        # Botones exportar
        frame_botones_exportar = ttk.Frame(ventana)
        frame_botones_exportar.pack(pady=10)

        btn_excel = ttk.Button(
            frame_botones_exportar,
            text="📥 Exportar a Excel",
            command=lambda: exportar_excel2(datos_exportacion, columnas)
        )
        btn_excel.pack(side="left", padx=10)

        btn_pdf = ttk.Button(
            frame_botones_exportar,
            text="📄 Exportar a PDF",
            command=lambda: exportar_pdf(datos_exportacion, columnas, "Reporte de Compras")
        )
        btn_pdf.pack(side="left", padx=10)

        ejecutar_consulta()

    except Exception as e:
        messagebox.showerror("Error", f"No se pudo cargar la vista de compras: {e}")




# Crear un frame horizontal para los botones
frame_botones_listas = ttk.Frame(tab_reportes)
frame_botones_listas.pack(pady=15)

btn_mostrar_inventario = tk.Button(
    frame_botones_listas, text="📦 Mostrar Compras", command=mostrar_compras, bg="#5CD6C0",fg="black",font=("Arial", 10, "bold"),relief="raised",padx=10,pady=3,activebackground="#47B7A7",cursor="hand2"
)
btn_mostrar_inventario.grid(row=0, column=0, padx=10)


# Función para mostrar productos con desplazamiento horizontal y vertical
def mostrar_inventario():
    try:
        ventana_productos = tk.Toplevel(root)
        ventana_productos.title("Lista de Productos")
        ventana_productos.geometry("1920x1080")

        # === FILTRO POR CÓDIGO DE PRODUCTO ===
        frame_filtros = ttk.Frame(ventana_productos)
        frame_filtros.pack(pady=10)

        ttk.Label(frame_filtros, text="🔍 Código Producto:").pack(side="left", padx=5)
        entry_codigo = ttk.Entry(frame_filtros, width=30)
        entry_codigo.pack(side="left", padx=5)

        frame_tree = ttk.Frame(ventana_productos)
        frame_tree.pack(fill=tk.BOTH, expand=True)

        scrollbar_x = ttk.Scrollbar(frame_tree, orient=tk.HORIZONTAL)
        scrollbar_y = ttk.Scrollbar(frame_tree, orient=tk.VERTICAL)

        columnas_tree = (
            "IdProducto", "CodigoProducto", "NombreProducto", "Categoria", "CostoConIVA", "PVPSinIVA",
            "CalculoIVA", "PrecioVentaConIVA", "CantidadInicial", "CantidadVendida",
            "CantidadPrestada", "CantidadInventario", "IVA"
        )

        tree = ttk.Treeview(
            frame_tree,
            columns=columnas_tree,
            show="headings",
            xscrollcommand=scrollbar_x.set,
            yscrollcommand=scrollbar_y.set
        )

        scrollbar_x.config(command=tree.xview)
        scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)
        scrollbar_y.config(command=tree.yview)
        scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)

        style = ttk.Style()
        style.configure("Treeview.Heading", font=("Arial", 13, "bold"))
        style.configure("Treeview", font=("Arial", 10))

        encabezados = [
            ("IdProducto", "IdProducto", 120),
            ("CodigoProducto", "CodigoProducto", 200),
            ("NombreProducto", "Nombre del Producto", 750),
            ("Categoria", "Categoría", 200),
            ("CostoConIVA", "Costo con IVA", 150),
            ("PVPSinIVA", "PVP sin IVA", 150),
            ("CalculoIVA", "Cálculo IVA", 150),
            ("PrecioVentaConIVA", "Precio Venta con IVA", 200),
            ("CantidadInicial", "Cantidad Inicial", 200),
            ("CantidadVendida", "Cantidad Vendida", 200),
            ("CantidadPrestada", "Cantidad Prestada", 200),
            ("CantidadInventario", "Cantidad en Inventario", 200),
            ("IVA", "IVA", 150)
        ]

        for col, text, width in encabezados:
            tree.heading(col, text=text)
            tree.column(col, width=width, anchor="center")

        tree.pack(fill=tk.BOTH, expand=True)

        datos_exportacion = []

        def ejecutar_filtro_inventario():
            try:
                tree.delete(*tree.get_children())
                datos_exportacion.clear()

                codigo = entry_codigo.get().strip()

                if codigo:
                    cursor.execute("SELECT * FROM Productos WHERE CodigoProducto = ? ORDER BY CodigoProducto", (codigo,))
                else:
                    cursor.execute("SELECT * FROM Productos ORDER BY CodigoProducto")

                productos = cursor.fetchall()

                for producto in productos:
                    fila = [str(x) if x is not None else "" for x in producto]
                    tree.insert("", tk.END, values=fila)
                    datos_exportacion.append(fila)

            except Exception as e:
                messagebox.showerror("Error", f"No se pudo aplicar el filtro: {e}")

        # Botón Buscar
        btn_buscar = ttk.Button(frame_filtros, text="🔍 Buscar", command=ejecutar_filtro_inventario)
        btn_buscar.pack(side="left", padx=10)

        # Ejecutar consulta inicial
        ejecutar_filtro_inventario()

        columnas = [col for col, _, _ in encabezados]

        frame_botones_exportar = ttk.Frame(ventana_productos)
        frame_botones_exportar.pack(pady=10)

        btn_excel = ttk.Button(
            frame_botones_exportar,
            text="📥 Exportar a Excel",
            command=lambda: exportar_excel(datos_exportacion, columnas)
        )
        btn_excel.pack(side="left", padx=10)

        btn_pdf = ttk.Button(
            frame_botones_exportar,
            text="📄 Exportar a PDF", style="Blue.TButton",
            command=lambda: exportar_pdf(datos_exportacion, columnas, "Reporte de Inventario")
        )
        btn_pdf.pack(side="left", padx=10)

    except Exception as e:
        messagebox.showerror("Error", f"No se pudo cargar la lista de productos: {e}")





btn_mostrar_inventario = tk.Button(
    frame_botones_listas, text="🏷️ Mostrar Inventario", command=mostrar_inventario,bg="#5CD6C0",fg="black",font=("Arial", 10, "bold"),relief="raised",padx=10,pady=3,activebackground="#47B7A7",cursor="hand2"
)
btn_mostrar_inventario.grid(row=0, column=1, padx=10)


from tkcalendar import DateEntry
from datetime import datetime

def mostrar_prestamos():
    try:
        ventana = tk.Toplevel(root)
        ventana.title("Lista de Préstamos")
        ventana.geometry("1920x1080")

        # === FILTROS ===
        frame_filtros = ttk.Frame(ventana)
        frame_filtros.pack(pady=10)

        ttk.Label(frame_filtros, text="🔎 Código Producto:").grid(row=0, column=0, padx=5)
        entry_codigo = ttk.Entry(frame_filtros, width=20)
        entry_codigo.grid(row=0, column=1, padx=5)

        ttk.Label(frame_filtros, text="👤 Cliente:").grid(row=0, column=2, padx=5)
        entry_cliente = ttk.Entry(frame_filtros, width=20)
        entry_cliente.grid(row=0, column=3, padx=5)

        ttk.Label(frame_filtros, text="📅 Fecha Inicio:").grid(row=0, column=4, padx=5)
        entry_fecha_inicio = DateEntry(frame_filtros, width=15, date_pattern="yyyy-mm-dd")
        entry_fecha_inicio.set_date(datetime(2025, 1, 1).date())
        entry_fecha_inicio.grid(row=0, column=5, padx=5)

        ttk.Label(frame_filtros, text="📅 Fecha Fin:").grid(row=0, column=6, padx=5)
        entry_fecha_fin = DateEntry(frame_filtros, width=15, date_pattern="yyyy-mm-dd")
        entry_fecha_fin.set_date(datetime.today().date())
        entry_fecha_fin.grid(row=0, column=7, padx=5)

        # === TREEVIEW ===
        frame_tree = ttk.Frame(ventana)
        frame_tree.pack(fill=tk.BOTH, expand=True)

        scrollbar_x = ttk.Scrollbar(frame_tree, orient=tk.HORIZONTAL)
        scrollbar_y = ttk.Scrollbar(frame_tree, orient=tk.VERTICAL)

        columnas_tree = [
            "IdPrestamo", "CodigoProducto", "NombreProducto", "CantidadPrestadaTotal",
            "CantidadPrestada", "CantidadDevuelta", "FechaPrestamo", "Cliente"
        ]

        tree = ttk.Treeview(
            frame_tree,
            columns=columnas_tree,
            show="headings",
            xscrollcommand=scrollbar_x.set,
            yscrollcommand=scrollbar_y.set
        )

        scrollbar_x.config(command=tree.xview)
        scrollbar_y.config(command=tree.yview)
        scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)
        scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)

        style = ttk.Style()
        style.configure("Treeview.Heading", font=("Arial", 13, "bold"))
        style.configure("Treeview", font=("Arial", 10))

        encabezados = [
            ("IdPrestamo", "ID Préstamo", 300),
            ("CodigoProducto", "Código Producto", 300),
            ("NombreProducto", "Nombre Producto", 600),
            ("CantidadPrestadaTotal", "Cantidad Prestada Total", 300),
            ("CantidadPrestada", "Cantidad Prestada", 300),
            ("CantidadDevuelta", "Cantidad Devuelta", 300),
            ("FechaPrestamo", "Fecha de Préstamo", 300),
            ("Cliente", "Cliente", 300)
        ]

        for col, text, width in encabezados:
            tree.heading(col, text=text)
            tree.column(col, width=width, anchor="center")
        tree.pack(fill=tk.BOTH, expand=True)

        datos_exportacion = []
        resumen_pendientes = []

        def ejecutar_consulta():
            for item in tree.get_children():
                tree.delete(item)
            datos_exportacion.clear()
            resumen_pendientes.clear()
            total_pendientes = 0

            codigo = entry_codigo.get().strip()
            cliente = entry_cliente.get().strip()
            fecha_inicio = entry_fecha_inicio.get_date()
            fecha_fin = entry_fecha_fin.get_date()

            query = """
                SELECT IdPrestamo, CodigoProducto, NombreProducto, CantidadPrestadaTotal,
                       CantidadPrestada, CantidadDevuelta, FechaPrestamo, Cliente
                FROM Prestamos
                WHERE FechaPrestamo BETWEEN ? AND ?
            """
            params = [fecha_inicio, fecha_fin]

            if codigo:
                query += " AND CodigoProducto = ?"
                params.append(codigo)
            if cliente:
                query += " AND Cliente = ?"
                params.append(cliente)

            query += " ORDER BY IdPrestamo"

            cursor.execute(query, params)
            prestamos = cursor.fetchall()

            for prestamo in prestamos:
                id_, cod, nombre, cant_total, cant_prestada, cant_devuelta, fecha, cli = prestamo
                fecha_str = fecha.strftime("%Y-%m-%d") if fecha else "N/A"
                fila = [id_, cod, nombre, cant_total, cant_prestada, cant_devuelta, fecha_str, cli]
                tree.insert("", tk.END, values=fila)
                datos_exportacion.append(fila)

                if cant_prestada and cant_prestada > 0:
                    resumen_pendientes.append((id_, cod, cant_prestada, cli))
                    total_pendientes += cant_prestada

            # Actualizar TextArea
            text_area.config(state="normal")
            text_area.delete("1.0", tk.END)
            text_area.insert(tk.END, "📦 Productos pendientes por devolución:\n\n", "titulo")
            for idp, cod, cant, cli in resumen_pendientes:
                text_area.insert(tk.END, f"• IdPréstamo: {idp} | Código: {cod} | Cantidad Prestada: {cant} | Cliente: {cli}\n", "texto")
            text_area.insert(tk.END, f"\n📊 Total general aún pendiente por devolver: {total_pendientes} producto(s)", "titulo")
            text_area.config(state="disabled")

            if not prestamos:
                messagebox.showinfo("Sin resultados", "No se encontraron préstamos con los filtros aplicados.")

        # === TEXTAREA DE RESUMEN DE PENDIENTES ===
        frame_text = ttk.Frame(ventana)
        frame_text.pack(fill="both", expand=True, padx=10, pady=10)

        scroll_text_y = ttk.Scrollbar(frame_text, orient="vertical")
        text_area = tk.Text(
            frame_text, height=1, font=("Segoe UI", 11), wrap="word",
            yscrollcommand=scroll_text_y.set, bg="#f9f9f9", relief="flat"
        )
        scroll_text_y.config(command=text_area.yview)
        scroll_text_y.pack(side="right", fill="y")
        text_area.pack(side="left", fill="both", expand=True)

        text_area.tag_configure("titulo", foreground="#007F5F", font=("Segoe UI", 12, "bold"))
        text_area.tag_configure("texto", foreground="#333333", font=("Segoe UI", 11))

        # === BOTÓN CONSULTAR ===
        btn_consultar = ttk.Button(frame_filtros, text="🔍 Consultar", command=ejecutar_consulta)
        btn_consultar.grid(row=0, column=8, padx=10)

        # === BOTONES EXPORTAR ===
        frame_botones_exportar = ttk.Frame(ventana)
        frame_botones_exportar.pack(pady=10)

        btn_excel = ttk.Button(
            frame_botones_exportar,
            text="📥 Exportar a Excel",
            command=lambda: exportar_excel(datos_exportacion, columnas_tree)
        )
        btn_excel.pack(side="left", padx=10)

        btn_pdf = ttk.Button(
            frame_botones_exportar,
            text="📄 Exportar a PDF", style="Blue.TButton",
            command=lambda: exportar_pdf(datos_exportacion, columnas_tree, "Reporte de Préstamos")
        )
        btn_pdf.pack(side="left", padx=10)

        ejecutar_consulta()

    except Exception as e:
        messagebox.showerror("Error", f"No se pudo cargar la lista de préstamos: {e}")




from datetime import datetime, date

def mostrar_devoluciones():
    try:
        # Crear ventana
        ventana = tk.Toplevel(root)
        ventana.title("Lista de Devoluciones")
        ventana.geometry("1920x1080")

        # Filtros
        frame_filtros = ttk.Frame(ventana)
        frame_filtros.pack(pady=5)

        ttk.Label(frame_filtros, text="🔍 Código Producto:").grid(row=0, column=0, padx=5)
        entry_codigo = ttk.Entry(frame_filtros)
        entry_codigo.grid(row=0, column=1, padx=5)

        ttk.Label(frame_filtros, text="👤 Cliente:").grid(row=0, column=2, padx=5)
        entry_cliente = ttk.Entry(frame_filtros)
        entry_cliente.grid(row=0, column=3, padx=5)

        ttk.Label(frame_filtros, text="📅 Fecha Inicio:").grid(row=0, column=4, padx=5)
        fecha_inicio = DateEntry(frame_filtros, date_pattern='yyyy-mm-dd')
        fecha_inicio.set_date(date(2025, 1, 1))  
        fecha_inicio.grid(row=0, column=5, padx=5)

        ttk.Label(frame_filtros, text="📅 Fecha Fin:").grid(row=0, column=6, padx=5)
        fecha_fin = DateEntry(frame_filtros, date_pattern='yyyy-mm-dd')
        fecha_fin.set_date(date.today())  # hasta hoy
        fecha_fin.grid(row=0, column=7, padx=5)

        def aplicar_filtros():
            codigo = entry_codigo.get().strip()
            cliente = entry_cliente.get().strip()
            fecha_ini = fecha_inicio.get_date()
            fecha_fin_val = fecha_fin.get_date()

            query = """
                SELECT IdDevolucion, OrdenCompra, CodigoProducto, CantidadDevuelta,
                       FechaDevolucion, Cliente, IdPrestamo
                FROM Devoluciones
                WHERE FechaDevolucion BETWEEN ? AND ?
            """
            params = [fecha_ini, fecha_fin_val]

            if codigo:
                query += " AND CodigoProducto = ?"
                params.append(codigo)

            if cliente:
                query += " AND Cliente = ?"
                params.append(cliente)

            query += " ORDER BY IdDevolucion"

            cursor.execute(query, params)
            devoluciones = cursor.fetchall()

            for row in tree.get_children():
                tree.delete(row)

            datos_exportacion.clear()

            if not devoluciones:
                messagebox.showinfo("Sin resultados", "No se encontraron devoluciones con los filtros aplicados.")
                return

            for devolucion in devoluciones:
                id_, orden, cod, cantidad, fecha, cli, idp = devolucion
                fecha_str = fecha.strftime("%Y-%m-%d") if fecha else "N/A"
                fila = [id_, orden, cod, cantidad, fecha_str, cli, idp]
                tree.insert("", tk.END, values=fila)
                datos_exportacion.append(fila)

        btn_filtrar = ttk.Button(frame_filtros, text="🔍 Consultar", command=aplicar_filtros)
        btn_filtrar.grid(row=0, column=8, padx=10)

        # Tabla Treeview
        frame_tree = ttk.Frame(ventana)
        frame_tree.pack(fill=tk.BOTH, expand=True)

        scrollbar_x = ttk.Scrollbar(frame_tree, orient=tk.HORIZONTAL)
        scrollbar_y = ttk.Scrollbar(frame_tree, orient=tk.VERTICAL)

        columnas_tree = [
            "IdDevolucion", "OrdenCompra", "CodigoProducto",
            "CantidadDevuelta", "FechaDevolucion", "Cliente", "IdPrestamo"
        ]

        tree = ttk.Treeview(
            frame_tree, columns=columnas_tree, show="headings",
            xscrollcommand=scrollbar_x.set, yscrollcommand=scrollbar_y.set
        )

        scrollbar_x.config(command=tree.xview)
        scrollbar_y.config(command=tree.yview)
        scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)
        scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)

        style = ttk.Style()
        style.configure("Treeview.Heading", font=("Arial", 13, "bold"))
        style.configure("Treeview", font=("Arial", 10))

        encabezados = [
            ("IdDevolucion", "ID Devolución", 200),
            ("OrdenCompra", "Orden de Compra", 200),
            ("CodigoProducto", "Código Producto", 200),
            ("CantidadDevuelta", "Cantidad Devuelta", 200),
            ("FechaDevolucion", "Fecha de Devolución", 200),
            ("Cliente", "Cliente", 200),
            ("IdPrestamo", "ID Préstamo", 200)
        ]

        for col, text, width in encabezados:
            tree.heading(col, text=text)
            tree.column(col, width=width, anchor="center")

        tree.pack(fill=tk.BOTH, expand=True)

        datos_exportacion = []

        # Botones exportar
        frame_botones = ttk.Frame(ventana)
        frame_botones.pack(pady=10)

        btn_excel = ttk.Button(frame_botones, text="📥 Exportar a Excel",
                               command=lambda: exportar_excel(datos_exportacion, columnas_tree))
        btn_excel.pack(side="left", padx=10)

        btn_pdf = ttk.Button(frame_botones, text="📄 Exportar a PDF", style="Blue.TButton",
                             command=lambda: exportar_pdf(datos_exportacion, columnas_tree, "Reporte de Devoluciones"))
        btn_pdf.pack(side="left", padx=10)

        # Carga inicial
        aplicar_filtros()

    except Exception as e:
        messagebox.showerror("Error", f"No se pudo cargar la lista de devoluciones: {e}")


btn_mostrar_prestamos = tk.Button(
    frame_botones_listas,
    text="📤 Mostrar Préstamos",
    command=mostrar_prestamos,
    bg="#5CD6C0",        # Color aqua
    fg="black",
    font=("Arial", 10, "bold"),
    relief="raised",
    padx=10,
    pady=3,
    activebackground="#47B7A7",  # Más oscuro al presionar
    cursor="hand2"
)
btn_mostrar_prestamos.grid(row=0, column=2, padx=10)

btn_mostrar_devoluciones = tk.Button(
    frame_botones_listas,
    text="📥 Mostrar Devoluciones",
    command=mostrar_devoluciones,
    bg="#5CD6C0",        # Mismo color aqua
    fg="black",
    font=("Arial", 10, "bold"),
    relief="raised",
    padx=10,
    pady=3,
    activebackground="#47B7A7",
    cursor="hand2"
)
btn_mostrar_devoluciones.grid(row=0, column=3, padx=10)


def mostrar_productos_stock():
    try:
        ventana_stock = tk.Toplevel(root)
        ventana_stock.title("Productos con Stock")
        ventana_stock.geometry("1920x1080")

        # === FILTRO POR CÓDIGO DE PRODUCTO ===
        frame_filtros = ttk.Frame(ventana_stock)
        frame_filtros.pack(pady=10)

        ttk.Label(frame_filtros, text="🔍 Código Producto:").pack(side="left", padx=5)
        entry_codigo = ttk.Entry(frame_filtros, width=30)
        entry_codigo.pack(side="left", padx=5)

        frame_tree = ttk.Frame(ventana_stock)
        frame_tree.pack(fill=tk.BOTH, expand=True)

        scrollbar_x = ttk.Scrollbar(frame_tree, orient=tk.HORIZONTAL)
        scrollbar_y = ttk.Scrollbar(frame_tree, orient=tk.VERTICAL)

        columnas_tree = (
            "IdProducto", "CodigoProducto", "NombreProducto", "Categoria", "CostoConIVA", "PVPSinIVA",
            "CalculoIVA", "PrecioVentaConIVA", "CantidadInicial", "CantidadVendida",
            "CantidadPrestada", "CantidadInventario", "IVA"
        )

        tree = ttk.Treeview(
            frame_tree,
            columns=columnas_tree,
            show="headings",
            xscrollcommand=scrollbar_x.set,
            yscrollcommand=scrollbar_y.set
        )

        scrollbar_x.config(command=tree.xview)
        scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)
        scrollbar_y.config(command=tree.yview)
        scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)

        style = ttk.Style()
        style.configure("Treeview.Heading", font=("Arial", 13, "bold"))
        style.configure("Treeview", font=("Arial", 10))

        encabezados = [
            ("IdProducto", "IdProducto", 250),
            ("CodigoProducto", "CodigoProducto", 250),
            ("NombreProducto", "Nombre del Producto", 450),
            ("Categoria", "Categoría", 250),
            ("CostoConIVA", "Costo con IVA", 250),
            ("PVPSinIVA", "PVP sin IVA", 250),
            ("CalculoIVA", "Cálculo IVA", 250),
            ("PrecioVentaConIVA", "Precio Venta con IVA", 250),
            ("CantidadInicial", "Cantidad Inicial", 250),
            ("CantidadVendida", "Cantidad Vendida", 250),
            ("CantidadPrestada", "Cantidad Prestada", 250),
            ("CantidadInventario", "Cantidad en Inventario", 250),
            ("IVA", "IVA", 250)
        ]

        for col, text, width in encabezados:
            tree.heading(col, text=text)
            tree.column(col, width=width, anchor="center")

        tree.pack(fill=tk.BOTH, expand=True)

        datos_exportacion = []

        def ejecutar_filtro_stock():
            try:
                tree.delete(*tree.get_children())
                datos_exportacion.clear()

                codigo = entry_codigo.get().strip()

                if codigo:
                    cursor.execute("SELECT * FROM Productos WHERE CantidadInventario > 0 AND CodigoProducto = ? ORDER BY CodigoProducto", (codigo,))
                else:
                    cursor.execute("SELECT * FROM Productos WHERE CantidadInventario > 0 ORDER BY CodigoProducto")

                productos = cursor.fetchall()

                for producto in productos:
                    fila = [str(x) if x is not None else "" for x in producto]
                    tree.insert("", tk.END, values=fila)
                    datos_exportacion.append(fila)

            except Exception as e:
                messagebox.showerror("Error", f"No se pudo aplicar el filtro: {e}")

        # Botón buscar
        btn_buscar = ttk.Button(frame_filtros, text="🔍 Buscar", command=ejecutar_filtro_stock)
        btn_buscar.pack(side="left", padx=10)

        # Ejecutar consulta inicial
        ejecutar_filtro_stock()

        columnas_exportar = [col for col, _, _ in encabezados]

        frame_botones_exportar = ttk.Frame(ventana_stock)
        frame_botones_exportar.pack(pady=10)

        btn_excel = ttk.Button(
            frame_botones_exportar,
            text="📥 Exportar a Excel",
            command=lambda: exportar_excel(datos_exportacion, columnas_exportar)
        )
        btn_excel.pack(side="left", padx=10)

        btn_pdf = ttk.Button(
            frame_botones_exportar,
            text="📄 Exportar a PDF", style="Blue.TButton",
            command=lambda: exportar_pdf(datos_exportacion, columnas_exportar, "Reporte de Productos con Stock")
        )
        btn_pdf.pack(side="left", padx=10)

    except Exception as e:
        messagebox.showerror("Error", f"No se pudo cargar la lista de productos con stock: {e}")



# Botón para mostrar productos con stock (puedes añadirlo a la pestaña correspondiente)
btn_mostrar_stock = tk.Button(tab_reportes, text="📦 Productos con Stock", command=mostrar_productos_stock, bg="#5CD6C0",fg="black",font=("Arial", 10, "bold"),relief="raised",padx=10,pady=3,activebackground="#47B7A7",cursor="hand2")
btn_mostrar_stock.pack(pady=18)


   
def mostrar_ordencompra():
    try:
        cursor.execute("SELECT * FROM OrdenCompra ORDER BY IDOrdenCompra")
        ordencompra = cursor.fetchall()

        ventana_ordencompra = tk.Toplevel(root)
        ventana_ordencompra.title("Lista de Ordenes de Compras")
        ventana_ordencompra.geometry("1920x1080")

        frame_tree = ttk.Frame(ventana_ordencompra)
        frame_tree.pack(fill=tk.BOTH, expand=True)

        scrollbar_x = ttk.Scrollbar(frame_tree, orient=tk.HORIZONTAL)
        scrollbar_y = ttk.Scrollbar(frame_tree, orient=tk.VERTICAL)

        columnas_tree = [
            "IdOrdenCompra", "NumOrdenCompra", "FechaOrdenCompra", "NombreCliente", "Telefono", "Ciudad",
            "NombreConsultor", "PorcentajeComisionConsultor", "ComisionPorPagarConsultor",
            "NombrePadreEmpresarial", "PorcentajePadreEmpresarial", "ComisionPorPagarPadreEmpresarial",
            "PorcentajeIVA", "CodigoProducto", "NombreProducto", "CantidadVendida",
            "PorcentajeDescuento", "PrecioVentaConIVA", "PVPSinIVA",
            "ValorDescuento", "BaseRetencion", "ValorBaseRetencion",
            "ValorCliente",  "ValorXCobrarConIVA"

        ]

        tree = ttk.Treeview(
            frame_tree, columns=columnas_tree, show="headings",
            xscrollcommand=scrollbar_x.set, yscrollcommand=scrollbar_y.set
        )

        scrollbar_x.config(command=tree.xview)
        scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)
        scrollbar_y.config(command=tree.yview)
        scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)

        style = ttk.Style()
        style.configure("Treeview.Heading", font=("Arial", 13, "bold"))
        style.configure("Treeview", font=("Arial", 10))

        encabezados = [
            ("IdOrdenCompra", "IdOrdenCompra", 250),
            ("NumOrdenCompra", "NumOrdenCompra", 250),
            ("FechaOrdenCompra", "FechaOrdenCompra", 250),
            ("NombreCliente", "NombreCliente", 250),
            ("Telefono", "Telefono", 250),
            ("Ciudad", "Ciudad", 250),
            ("NombreConsultor", "NombreConsultor", 250),
            ("PorcentajeComisionConsultor", "PorcentajeComisionConsultor", 250),
            ("ComisionPorPagarConsultor", "ComisionPorPagarConsultor", 250),
            ("NombrePadreEmpresarial", "NombrePadreEmpresarial", 250),
            ("PorcentajePadreEmpresarial", "PorcentajePadreEmpresarial", 250),
            ("ComisionPorPagarPadreEmpresarial", "ComisionPorPagarPadreEmpresarial", 250),
            ("PorcentajeIVA", "PorcentajeIVA", 250),
            ("CodigoProducto", "CodigoProducto", 250),
            ("NombreProducto", "NombreProducto", 250),
            ("CantidadVendida", "CantidadVendida", 250),
            ("PorcentajeDescuento", "PorcentajeDescuento", 250),
            ("PrecioVentaConIVA", "PrecioVentaConIVA", 250),
            ("PVPSinIVA", "PVPSinIVA", 250),
            ("ValorDescuento", "ValorDescuento", 250),
            ("BaseRetencion", "BaseRetencion", 250),
            ("ValorBaseRetencion", "ValorBaseRetencion", 250),
            ("ValorCliente", "ValorCliente", 250),
            ("ValorXCobrarConIVA", "ValorXCobrarConIVA", 250)
        ]

        for col, text, width in encabezados:
            tree.heading(col, text=text)
            tree.column(col, width=width, anchor="center")

        tree.pack(fill=tk.BOTH, expand=True)

        # Insertar datos y preparar para exportación
        datos_exportacion = []
        for ordencompra in ordencompra:
            fila = [str(x) if x is not None else "" for x in ordencompra]
            tree.insert("", tk.END, values=fila)
            datos_exportacion.append(fila)

        columnas_exportar = [col for col, _, _ in encabezados]

        frame_botones_exportar = ttk.Frame(ventana_ordencompra)
        frame_botones_exportar.pack(pady=10)

        btn_excel = ttk.Button(
            frame_botones_exportar,
            text="📥 Exportar a Excel",
            command=lambda: exportar_excel(datos_exportacion, columnas_exportar)
        )
        btn_excel.pack(side="left", padx=10)


    except Exception as e:
        messagebox.showerror("Error", f"No se pudo cargar la lista de órdenes de compra: {str(e)}")



btn_mostrar_ventas = tk.Button(tab_reportes, text="📈 Mostrar Órdenes de compras", command=mostrar_ordencompra, bg="#5CD6C0",fg="black",font=("Arial", 10, "bold"),relief="raised",padx=10,pady=3,activebackground="#47B7A7",cursor="hand2")
btn_mostrar_ventas.pack(pady=18)

def mostrar_cuentasporcobrar():
    try:
        ventana = tk.Toplevel(root)
        ventana.title("Lista de Cuentas por Cobrar")
        ventana.geometry("1920x1080")

        # === FILTROS ===
        frame_filtros = ttk.Frame(ventana)
        frame_filtros.pack(pady=10)

        ttk.Label(frame_filtros, text="🔢 N° Orden Compra:").grid(row=0, column=0, padx=5)
        entry_orden = ttk.Entry(frame_filtros, width=25)
        entry_orden.grid(row=0, column=1, padx=5)

        ttk.Label(frame_filtros, text="👤 Nombre Cliente:").grid(row=0, column=2, padx=5)
        entry_cliente = ttk.Entry(frame_filtros, width=25)
        entry_cliente.grid(row=0, column=3, padx=5)

        def ejecutar_filtro():
            try:
                num_orden = entry_orden.get().strip()
                nombre_cliente = entry_cliente.get().strip().lower()

                query = "SELECT * FROM CuentasPorCobrar"
                condiciones = []
                params = []

                if num_orden:
                    condiciones.append("NumOrdenCompra = ?")
                    params.append(num_orden)
                if nombre_cliente:
                    condiciones.append("LOWER(NombreCliente) = ?")
                    params.append(nombre_cliente)

                if condiciones:
                    query += " WHERE " + " AND ".join(condiciones)

                query += " ORDER BY IdCuenta"

                cursor.execute(query, params)
                cuentasporcobrar = cursor.fetchall()

                if not cuentasporcobrar:
                    messagebox.showwarning("Advertencia", "No se encontraron cuentas con esos filtros.")
                    return

                tree.delete(*tree.get_children())
                datos_exportacion.clear()
                text_area.config(state="normal")
                text_area.delete("1.0", tk.END)

                total_saldo = 0
                for cuenta in cuentasporcobrar:
                    fila = [str(x) if x is not None else "" for x in cuenta]
                    tree.insert("", tk.END, values=fila)
                    datos_exportacion.append(fila)

                    saldo = cuenta[44] or 0
                    if saldo <= 0:
                        continue

                    num_orden = cuenta[1] or "Sin número"
                    nombre_cliente = cuenta[2] or "Desconocido"
                    factura = cuenta[11] or "-"
                    num_factura = str(int(float(cuenta[12]))) if cuenta[12] else "-"
                    lote2 = str(int(float(cuenta[17]))) if cuenta[17] else ""
                    lote3 = str(int(float(cuenta[30]))) if cuenta[30] else ""

                    text_area.insert(tk.END, f"🔢 Orden: {num_orden} - 👤 Cliente: {nombre_cliente}\n", "orden")
                    text_area.insert(tk.END, f"🧾 Factura: {factura} - {num_factura}\n", "texto")
                    if lote2:
                        text_area.insert(tk.END, f"📦 Lote2: {lote2}\n", "texto")
                    if lote3:
                        text_area.insert(tk.END, f"📦 Lote3: {lote3}\n", "texto")
                    text_area.insert(tk.END, f"💰 Saldo: ${round(saldo, 2)}\n", "texto")
                    text_area.insert(tk.END, "—" * 60 + "\n", "texto")
                    total_saldo += saldo

                resumen_final = f"\n📊 Total acumulado por cobrar: ${round(total_saldo, 2)}"
                text_area.insert(tk.END, resumen_final, "total")
                text_area.config(state="disabled")

            except Exception as ex:
                messagebox.showerror("Error", f"Error al aplicar filtros: {ex}")

        btn_filtrar = ttk.Button(frame_filtros, text="🔍 Consultar", command=ejecutar_filtro)
        btn_filtrar.grid(row=0, column=4, padx=10)

        # === TREEVIEW ===
        frame_tree = ttk.Frame(ventana)
        frame_tree.pack(fill=tk.BOTH, expand=True)

        scrollbar_x = ttk.Scrollbar(frame_tree, orient=tk.HORIZONTAL)
        scrollbar_y = ttk.Scrollbar(frame_tree, orient=tk.VERTICAL)

        columnas_tree = [
            "IdCuenta", "NumOrdenCompra", "NombreCliente", "TipoPagoEfecTrans", "AbonoEfectivoTransferencia1", "FechaPagadoEfectivo1",
            "AbonoEfectivoTransferencia2", "FechaPagadoEfectivo2", "AbonoEfectivoTransferencia3", "FechaPagadoEfectivo3",
            "TotalEfectivo", "Factura", "NumeroFactura", "IVAPagoEfectivoFactura", "TipoPago2", "ValorPagadoTarjeta2",
            "Banco2", "Lote2", "FechaPagado2", "PorcentajeComisionBanco2", "ComisionTCfactura2", "PorcentajeIRF2",
            "IRF2", "PorcentajeRetIVA2", "RetIVAPagoTarjetaCredito2", "TotalComisionBanco2", "ValorNetoTC2",
            "TipoPago3", "ValorPagadoTarjeta3", "Banco3", "Lote3", "FechaPagado3", "PorcentajeComisionBanco3",
            "ComisionTCfactura3", "PorcentajeIRF3", "IRF3", "PorcentajeRetIVA3", "RetIVAPagoTarjetaCredito3",
            "TotalComisionBanco3", "ValorNetoTC3", "ComisionBancoTotales", "TotalesValorNetoTC",
            "ValorXCobrarConIVATotal", "BaseRetencionTotal", "SaldoXCobrarCliente", "CostoConIVA",
            "UtilidadDescontadoIVASRI", "PorcentajeGanancia"
        ]

        tree = ttk.Treeview(
            frame_tree, columns=columnas_tree, show="headings",
            xscrollcommand=scrollbar_x.set, yscrollcommand=scrollbar_y.set
        )
        scrollbar_x.config(command=tree.xview)
        scrollbar_y.config(command=tree.yview)
        scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)
        scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)

        style = ttk.Style()
        style.configure("Treeview.Heading", font=("Arial", 13, "bold"))
        style.configure("Treeview", font=("Arial", 10))

        for col in columnas_tree:
            tree.heading(col, text=col)
            tree.column(col, width=280, anchor="center")
        tree.pack(fill=tk.BOTH, expand=True)

        datos_exportacion = []

        # === TEXT AREA ===
        frame_text = ttk.Frame(ventana)
        frame_text.pack(fill="both", expand=True, padx=10, pady=10)

        scroll_text_y = ttk.Scrollbar(frame_text, orient="vertical")
        text_area = tk.Text(
            frame_text, height=1, font=("Segoe UI", 11), wrap="word",
            yscrollcommand=scroll_text_y.set, bg="#f0f0f0", relief="flat"
        )
        scroll_text_y.config(command=text_area.yview)
        scroll_text_y.pack(side="right", fill="y")
        text_area.pack(side="left", fill="both", expand=True)

        text_area.tag_configure("orden", foreground="#007F5F", font=("Segoe UI", 11, "bold"))
        text_area.tag_configure("texto", foreground="#333333", font=("Segoe UI", 11))
        text_area.tag_configure("total", foreground="#004085", font=("Segoe UI", 11, "bold"))

        # === BOTÓN EXPORTACIÓN ===
        frame_botones_exportar = ttk.Frame(ventana)
        frame_botones_exportar.pack(pady=10)
        btn_excel = ttk.Button(
            frame_botones_exportar,
            text="📥 Exportar a Excel",
            command=lambda: exportar_excel(datos_exportacion, columnas_tree)
        )
        btn_excel.pack(side="left", padx=10)

        # Cargar resultados sin filtros al inicio
        ejecutar_filtro()

    except Exception as e:
        messagebox.showerror("Error", f"No se pudo cargar la lista de cuentas por cobrar: {str(e)}")



# Crear un contenedor centrado (intermedio)
frame_contenedor_ordenes = ttk.Frame(tab_reportes)
frame_contenedor_ordenes.pack(fill="x")  # ocupa el ancho completo para centrar con padx

# Sub-contenedor que movemos con padding
frame_busqueda_cxc = ttk.Frame(frame_contenedor_ordenes)
frame_busqueda_cxc.pack(pady=5, anchor="center")





btn_mostrar_ventas = tk.Button(
    frame_busqueda_cxc,
    text="💳 Mostrar Cuentas por Cobrar Pagadas",
    command=mostrar_cuentasporcobrar,
    bg="#5CD6C0", fg="black",
    font=("Arial", 10, "bold"),
    relief="raised", padx=10, pady=3,
    activebackground="#47B7A7",
    cursor="hand2"
)
btn_mostrar_ventas.pack(side="left", padx=20)


# --- Lógica para buscar y mostrar ---
def consultar_cuentas_por_cobrar():
    try:
        ventana_resultados = tk.Toplevel(root)
        ventana_resultados.title("Consultar Cuentas por Cobrar (Incluye órdenes aún no registradas)")
        ventana_resultados.geometry("1920x1080")

        # === FILTROS ===
        frame_filtros = ttk.Frame(ventana_resultados)
        frame_filtros.pack(pady=10)

        ttk.Label(frame_filtros, text="🔎 Filtro Estado: ").grid(row=0, column=0, padx=5)
        combo_filtro_estado = ttk.Combobox(frame_filtros, values=["Mostrar todas", "Solo cobradas (Saldo = 0)", "Por cobrar (Saldo > 0)"], state="readonly", width=30)
        combo_filtro_estado.grid(row=0, column=1, padx=5)
        combo_filtro_estado.current(0)

        ttk.Label(frame_filtros, text="👤 Nombre Cliente: ").grid(row=0, column=2, padx=5)
        entry_nombre_cliente = ttk.Entry(frame_filtros, width=30)
        entry_nombre_cliente.grid(row=0, column=3, padx=5)

        ttk.Label(frame_filtros, text="📄 N° Orden Compra:").grid(row=0, column=4, padx=5)
        entry_num_orden = ttk.Entry(frame_filtros, width=20)
        entry_num_orden.grid(row=0, column=5, padx=5)

        # === TREEVIEW ===
        frame_tree = ttk.Frame(ventana_resultados)
        frame_tree.pack(fill=tk.BOTH, expand=True)

        scrollbar_x = ttk.Scrollbar(frame_tree, orient=tk.HORIZONTAL)
        scrollbar_y = ttk.Scrollbar(frame_tree, orient=tk.VERTICAL)

        columnas = ["NumOrdenCompra", "NombreCliente", "Telefono", "Ciudad", "ValorXCobrarConIVA", "SaldoXCobrarCliente"]

        tree = ttk.Treeview(
            frame_tree, columns=columnas, show="headings",
            xscrollcommand=scrollbar_x.set, yscrollcommand=scrollbar_y.set
        )
        scrollbar_x.config(command=tree.xview)
        scrollbar_y.config(command=tree.yview)
        scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)
        scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)

        for col in columnas:
            tree.heading(col, text=col, anchor='center')
            tree.column(col, width=250, anchor='center')

        tree.pack(fill=tk.BOTH, expand=True)

        datos_exportacion = []

        def ejecutar_consulta():
            try:
                filtro_estado = combo_filtro_estado.get()
                nombre_cliente = entry_nombre_cliente.get().strip()
                num_orden = entry_num_orden.get().strip()

                query = """
                    SELECT 
                        cpc.NumOrdenCompra,
                        cpc.NombreCliente,
                        oc.Telefono,
                        oc.Ciudad,
                        oc.ValorXCobrarConIVA,
                        cpc.SaldoXCobrarCliente
                    FROM 
                        CuentasPorCobrar cpc
                    INNER JOIN 
                        (SELECT NumOrdenCompra, MAX(NombreCliente) AS NombreCliente, MAX(Telefono) AS Telefono, MAX(Ciudad) AS Ciudad, SUM(ValorXCobrarConIVA) AS ValorXCobrarConIVA
                         FROM OrdenCompra
                         GROUP BY NumOrdenCompra) oc
                    ON cpc.NumOrdenCompra = oc.NumOrdenCompra

                    UNION ALL

                    SELECT
                        oc.NumOrdenCompra,
                        oc.NombreCliente,
                        oc.Telefono,
                        oc.Ciudad,
                        oc.ValorXCobrarConIVA,
                        oc.ValorXCobrarConIVA
                    FROM 
                        (SELECT 
                            NumOrdenCompra, 
                            MAX(NombreCliente) AS NombreCliente, 
                            MAX(Telefono) AS Telefono,
                            MAX(Ciudad) AS Ciudad,
                            SUM(ValorXCobrarConIVA) AS ValorXCobrarConIVA
                        FROM OrdenCompra
                        GROUP BY NumOrdenCompra) oc
                    LEFT JOIN 
                        CuentasPorCobrar cpc
                    ON oc.NumOrdenCompra = cpc.NumOrdenCompra
                    WHERE cpc.NumOrdenCompra IS NULL
                """

                filtros = []
                params = []

                if filtro_estado == "Solo cobradas (Saldo = 0)":
                    filtros.append("SaldoXCobrarCliente = 0")
                elif filtro_estado == "Por cobrar (Saldo > 0)":
                    filtros.append("SaldoXCobrarCliente > 0")

                if nombre_cliente:
                    filtros.append("LOWER(NombreCliente) = ?")
                    params.append(nombre_cliente.lower())

                if num_orden:
                    filtros.append("NumOrdenCompra = ?")
                    params.append(num_orden)

                if filtros:
                    query = f"SELECT * FROM ({query}) AS subquery WHERE " + " AND ".join(filtros)

                query += " ORDER BY NumOrdenCompra"

                cursor.execute(query, params)
                resultados = cursor.fetchall()

                # Limpiar tree
                for item in tree.get_children():
                    tree.delete(item)
                datos_exportacion.clear()

                for fila in resultados:
                    fila_formateada = [str(x) if x is not None else "" for x in fila]
                    tree.insert("", tk.END, values=fila_formateada)
                    datos_exportacion.append(fila_formateada)

                if not resultados:
                    messagebox.showinfo("Sin resultados", "No se encontraron cuentas con los filtros seleccionados.")

            except Exception as e:
                messagebox.showerror("Error", f"Ocurrió un error al consultar: {e}")

        # Botón Consultar
        btn_consultar = tk.Button(
            frame_filtros,
            text="🔍 Consultar",
            bg="lightgreen",
            command=ejecutar_consulta
        )
        btn_consultar.grid(row=0, column=6, padx=10)

        # Botón Exportar a Excel
        frame_botones = ttk.Frame(ventana_resultados)
        frame_botones.pack(pady=10)

        btn_exportar_excel = ttk.Button(
            frame_botones,
            text="📥 Exportar a Excel",
            command=lambda: exportar_excel(datos_exportacion, columnas)
        )
        btn_exportar_excel.pack()

        # Ejecutar la primera consulta automáticamente
        ejecutar_consulta()

    except Exception as e:
        messagebox.showerror("Error", f"Ocurrió un error: {e}")



# Botón Consultar
btn_consultar_cuentas_cxc = tk.Button(
    frame_busqueda_cxc,
    text="🔍 Consultar Cuentas por Cobrar",
    command=consultar_cuentas_por_cobrar,
    bg="#5CD6C0", fg="black",
    font=("Arial", 10, "bold"),
    relief="raised", padx=10, pady=3,
    activebackground="#47B7A7",
    cursor="hand2"
)
btn_consultar_cuentas_cxc.pack(side="left", padx=20)


# Crear un marco para la búsqueda de productos
frame_buscar_productos = ttk.Frame(tab_reportes)
frame_buscar_productos.pack(pady=20)

ttk.Label(frame_buscar_productos, text="🔍 Buscar Producto:").pack(side="left", padx=5)
entry_buscar_producto = ttk.Entry(frame_buscar_productos, width=30)
entry_buscar_producto.pack(side="left", padx=5)

# Botón para buscar productos
btn_buscar_producto = tk.Button(frame_buscar_productos, text="🔍 Buscar", bg="lightgreen", command=lambda: buscar_producto(entry_buscar_producto.get().strip()))
btn_buscar_producto.pack(side="left", padx=5)

def buscar_producto(nombre_producto):
    """Busca productos por nombre y muestra los resultados en una nueva ventana con scroll horizontal y vertical."""
    if not nombre_producto:
        messagebox.showwarning("Atención", "Por favor, ingrese un nombre de producto para buscar.")
        return

    cursor.execute("""
        SELECT CodigoProducto, NombreProducto 
        FROM Productos 
        WHERE NombreProducto LIKE ?
    """, ('%' + nombre_producto + '%',))
    
    rows = cursor.fetchall()

    if not rows:
        messagebox.showinfo("Resultado", "No se encontraron productos con ese nombre.")
        return

    ventana_resultados = tk.Toplevel(root)
    ventana_resultados.title("Resultados de Búsqueda")
    ventana_resultados.geometry("800x500")

    # Frame que contiene el Text y los Scrollbars
    frame_resultados = tk.Frame(ventana_resultados)
    frame_resultados.pack(fill="both", expand=True, padx=10, pady=10)

    scroll_y = tk.Scrollbar(frame_resultados, orient="vertical")
    scroll_x = tk.Scrollbar(frame_resultados, orient="horizontal")
    text_resultados = tk.Text(frame_resultados, wrap="none", font=("Arial", 11),
                              yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)

    scroll_y.config(command=text_resultados.yview)
    scroll_x.config(command=text_resultados.xview)

    scroll_y.pack(side="right", fill="y")
    scroll_x.pack(side="bottom", fill="x")
    text_resultados.pack(side="left", fill="both", expand=True)

    for row in rows:
        codigo_producto, nombre_producto = row
        text_resultados.insert(tk.END, f"📌 Código: {codigo_producto}\n🔹 Producto: {nombre_producto}\n\n")

    text_resultados.config(state="disabled")



def buscar_ordencompra():
    try:
        fecha_inicio = entry_fecha_inicio.get_date().strftime('%Y-%m-%d')
        fecha_fin = entry_fecha_fin.get_date().strftime('%Y-%m-%d')
        num_orden = entry_num_orden_filtro.get().strip()
        codigo_producto = entry_codigo_producto.get().strip()
        cliente = entry_cliente.get().strip()
        consultor = entry_consultor.get().strip()
        ciudad = entry_ciudad.get().strip()
        padre_empresarial = entry_padre_empresarial.get().strip()

        query = "SELECT * FROM OrdenCompra WHERE FechaOrdenCompra BETWEEN ? AND ?"
        params = [fecha_inicio, fecha_fin]

        if codigo_producto:
            query += " AND CodigoProducto LIKE ?"
            params.append(codigo_producto)
        if cliente:
            query += " AND NombreCliente LIKE ?"
            params.append(cliente)
        if consultor:
            query += " AND NombreConsultor LIKE ?"
            params.append(consultor)
        if padre_empresarial:
            query += " AND NombrePadreEmpresarial LIKE ?"
            params.append(padre_empresarial)
        if ciudad:
            query += " AND Ciudad LIKE ?"
            params.append(ciudad)
        if num_orden:
            query += " AND NumOrdenCompra LIKE ?"
            params.append(num_orden)

        cursor.execute(query, params)
        rows = cursor.fetchall()

        if not rows:
            messagebox.showinfo("Resultado", "No se encontraron órdenes de compra con los filtros aplicados.")
            return

        column_names = [desc[0] for desc in cursor.description]
        idx_valor_cobrar = column_names.index("ValorXCobrarConIVA")
        column_names.insert(idx_valor_cobrar + 1, "SaldoPorCobrar")

        # Índices para totales
        indices_a_sumar = {
            "CantidadVendida": column_names.index("CantidadVendida"),
            "ValorXCobrarConIVA": column_names.index("ValorXCobrarConIVA"),
            "ComisionPorPagarConsultor": column_names.index("ComisionPorPagarConsultor"),
            "ComisionPorPagarPadreEmpresarial": column_names.index("ComisionPorPagarPadreEmpresarial"),
        }

        total_cantidad_vendida = 0
        total_ValorXCobrarConIVA = 0
        total_comision_consultor = 0
        total_comision_padre_empresarial = 0
        total_saldo_por_cobrar = 0

        # === Agrupar índices por NumOrdenCompra
        orden_fila_indices = {}
        for i, row in enumerate(rows):
            orden = row[column_names.index("NumOrdenCompra")]
            if orden not in orden_fila_indices:
                orden_fila_indices[orden] = []
            orden_fila_indices[orden].append(i)

        ventana_ventas = tk.Toplevel(root)
        ventana_ventas.title("Órdenes de Compras Filtradas")
        ventana_ventas.geometry("1920x1080")

        frame_tree = ttk.Frame(ventana_ventas)
        frame_tree.pack(fill=tk.BOTH, expand=True)

        scrollbar_x = ttk.Scrollbar(frame_tree, orient=tk.HORIZONTAL)
        scrollbar_y = ttk.Scrollbar(frame_tree, orient=tk.VERTICAL)

        tree = ttk.Treeview(
            frame_tree, columns=column_names, show="headings",
            xscrollcommand=scrollbar_x.set, yscrollcommand=scrollbar_y.set
        )
        scrollbar_x.config(command=tree.xview)
        scrollbar_y.config(command=tree.yview)
        scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)
        scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)

        for col in column_names:
            tree.heading(col, text=col, anchor='center')
            tree.column(col, width=250, anchor='center')

        tree.pack(fill=tk.BOTH, expand=True)

        datos_exportacion = []

        for i, row in enumerate(rows):
            num_orden_actual = row[column_names.index("NumOrdenCompra")]

            # Ver si esta es la última fila de ese NumOrdenCompra
            es_ultima_fila = (i == orden_fila_indices[num_orden_actual][-1])

            if es_ultima_fila:
                cursor.execute("SELECT SUM(SaldoXCobrarCliente) FROM CuentasPorCobrar WHERE NumOrdenCompra = ?", (num_orden_actual,))
                saldo = cursor.fetchone()[0]
                if saldo is None:
                    cursor.execute("SELECT SUM(ValorXCobrarConIVA) FROM OrdenCompra WHERE NumOrdenCompra = ?", (num_orden_actual,))
                    saldo = cursor.fetchone()[0] or 0
                total_saldo_por_cobrar += saldo
                saldo_str = f"{saldo:.2f}"
            else:
                saldo_str = "0.00"

            formatted_row = [str(value) if value is not None else '' for value in row]
            formatted_row.insert(idx_valor_cobrar + 1, saldo_str)

            tree.insert("", tk.END, values=formatted_row)
            datos_exportacion.append(formatted_row)

            total_cantidad_vendida += float(row[indices_a_sumar["CantidadVendida"]] or 0)
            total_ValorXCobrarConIVA += float(row[indices_a_sumar["ValorXCobrarConIVA"]] or 0)
            total_comision_consultor += float(row[indices_a_sumar["ComisionPorPagarConsultor"]] or 0)
            total_comision_padre_empresarial += float(row[indices_a_sumar["ComisionPorPagarPadreEmpresarial"]] or 0)

        # Texto totales
        total_frame = ttk.Frame(ventana_ventas)
        total_frame.pack(fill='x', padx=10, pady=10)

        total_text = f"""
        🔹 Totales del Rango de Fecha Seleccionado 🔹
        --------------------------------------------
        🏷️ Cantidad Vendida: {total_cantidad_vendida:,.4f}
        💲 Valor por Cobrar con IVA: {total_ValorXCobrarConIVA:,.4f}
        💼 Saldo por Cobrar Total: {total_saldo_por_cobrar:,.4f}
        💰 Comisión Consultor: {total_comision_consultor:,.4f}
        📌 Comisión Padre Empresarial: {total_comision_padre_empresarial:,.4f}
        """

        text_area_totales = tk.Text(total_frame, height=10, wrap="word", font=("Arial", 12, "bold"))
        text_area_totales.insert(tk.END, total_text)
        text_area_totales.config(state="disabled")
        text_area_totales.pack(fill='both', expand=True)

        # Botón exportar
        frame_botones_exportar = ttk.Frame(ventana_ventas)
        frame_botones_exportar.pack(pady=10)

        btn_excel = ttk.Button(
            frame_botones_exportar,
            text="📥 Exportar a Excel",
            command=lambda: exportar_excel(datos_exportacion, column_names)
        )
        btn_excel.pack(side="left", padx=10)

    except Exception as e:
        messagebox.showerror("Error", f"Ocurrió un error al buscar ventas: {e}")



def buscar_compras():
    """Muestra las compras filtradas por fecha y código de producto, con exportación a Excel y PDF."""
    try:
        fecha_inicio = entry_fecha_inicio.get_date().strftime('%Y-%m-%d')
        fecha_fin = entry_fecha_fin.get_date().strftime('%Y-%m-%d')
        codigo_producto = entry_buscar_codigo.get().strip()

        query = """
            SELECT IdCompra, FechaCompra, CodigoProducto, NombreProducto, CantidadComprada, 
                   CostoSinIVA, PorcentajeIVA, IVA, CostoConIVA, Proveedor
            FROM Compras
            WHERE FechaCompra BETWEEN ? AND ?
        """
        params = [fecha_inicio, fecha_fin]

        if codigo_producto:
            query += " AND CodigoProducto LIKE ?"
            params.append(f"%{codigo_producto}%")

        cursor.execute(query, params)
        compras = cursor.fetchall()

        if not compras:
            messagebox.showinfo("Resultado", "No se encontraron compras con los filtros aplicados.")
            return

        ventana_compras = tk.Toplevel(root)
        ventana_compras.title("Compras Filtradas")
        ventana_compras.geometry("1920x1080")

        frame_tree = ttk.Frame(ventana_compras)
        frame_tree.pack(fill=tk.BOTH, expand=True)

        scrollbar_x = ttk.Scrollbar(frame_tree, orient=tk.HORIZONTAL)
        scrollbar_y = ttk.Scrollbar(frame_tree, orient=tk.VERTICAL)

        columnas = [
            "IdCompra", "FechaCompra", "CodigoProducto", "NombreProducto",
            "CantidadComprada", "CostoSinIVA", "PorcentajeIVA", "IVA",
            "CostoConIVA", "Proveedor"
        ]

        tree = ttk.Treeview(
            frame_tree, columns=columnas, show="headings",
            xscrollcommand=scrollbar_x.set, yscrollcommand=scrollbar_y.set
        )

        scrollbar_x.config(command=tree.xview)
        scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)
        scrollbar_y.config(command=tree.yview)
        scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)

        style = ttk.Style()
        style.configure("Treeview.Heading", font=("Arial", 13, "bold"))
        style.configure("Treeview", font=("Arial", 10))

        encabezados = [
            ("IdCompra", "ID Compra", 100),
            ("FechaCompra", "Fecha de Compra", 150),
            ("CodigoProducto", "Código Producto", 150),
            ("NombreProducto", "Nombre Producto", 300),
            ("CantidadComprada", "Cantidad Comprada", 150),
            ("CostoSinIVA", "Costo Sin IVA", 150),
            ("PorcentajeIVA", "Porcentaje IVA", 100),
            ("IVA", "Valor IVA", 150),
            ("CostoConIVA", "Costo Con IVA", 150),
            ("Proveedor", "Proveedor", 200)
        ]

        for col, text, width in encabezados:
            tree.heading(col, text=text)
            tree.column(col, width=width, anchor="center")

        tree.pack(fill=tk.BOTH, expand=True)

        total_costo_sin_iva = 0
        total_valor_iva = 0
        total_costo_con_iva = 0

        datos_exportacion = []

        for compra in compras:
            (
                id_compra, fecha, codigo, nombre, cantidad,
                costo_sin_iva, iva, valor_iva, costo_con_iva, proveedor
            ) = compra

            fecha = fecha.strftime("%Y-%m-%d") if fecha else "N/A"
            costo_sin_iva = float(costo_sin_iva or 0)
            valor_iva = float(valor_iva or 0)
            costo_con_iva = float(costo_con_iva or 0)

            total_costo_sin_iva += costo_sin_iva
            total_valor_iva += valor_iva
            total_costo_con_iva += costo_con_iva

            fila = [
                id_compra, fecha, codigo, nombre, cantidad,
                f"{costo_sin_iva:,.2f}",
                f"{iva:.2%}" if iva is not None else "0.00%",
                f"{valor_iva:,.2f}",
                f"{costo_con_iva:,.2f}",
                proveedor if proveedor else "N/A"
            ]

            tree.insert("", tk.END, values=fila)
            datos_exportacion.append([
                id_compra, fecha, codigo, nombre, cantidad,
                costo_sin_iva, iva, valor_iva, costo_con_iva, proveedor
            ])

        total_frame = ttk.Frame(ventana_compras)
        total_frame.pack(fill='x', padx=10, pady=10)

        total_text = f"""
        🔹 Totales del Rango de Fecha Seleccionado 🔹
        --------------------------------------------
        📊 Total Costo Sin IVA: ${total_costo_sin_iva:,.2f}
        💰 Total Valor IVA: ${total_valor_iva:,.2f}
        📦 Total Costo Con IVA: ${total_costo_con_iva:,.2f}
        """

        text_area_totales = tk.Text(total_frame, height=13, wrap="word", font=("Arial", 12, "bold"))
        text_area_totales.insert(tk.END, total_text)
        text_area_totales.config(state="disabled")
        text_area_totales.pack(fill='both', expand=True)

        # Botones de exportación
        frame_botones_exportar = ttk.Frame(ventana_compras)
        frame_botones_exportar.pack(pady=10)

        columnas_exportar = [col for col, _, _ in encabezados]

        btn_excel = ttk.Button(
            frame_botones_exportar,
            text="📥 Exportar a Excel",
            command=lambda: exportar_excel(datos_exportacion, columnas_exportar)
        )
        btn_excel.pack(side="left", padx=10)

        btn_pdf = ttk.Button(
            frame_botones_exportar,
            text="📄 Exportar a PDF",style="Blue.TButton",
            command=lambda: exportar_pdf(datos_exportacion, columnas_exportar, "Compras Filtradas")
        )
        btn_pdf.pack(side="left", padx=10)

    except Exception as e:
        messagebox.showerror("Error", f"Ocurrió un error al buscar compras: {e}")

def buscar_compras2():
    """Muestra las compras filtradas por fecha y proveedor, con exportación a Excel y PDF."""
    try:
        fecha_inicio = entry_fecha_inicio.get_date().strftime('%Y-%m-%d')
        fecha_fin = entry_fecha_fin.get_date().strftime('%Y-%m-%d')
        proveedor = entry_buscar_proveedor.get().strip()

        query = """
            SELECT IdCompra, FechaCompra, CodigoProducto, NombreProducto, CantidadComprada, 
                   CostoSinIVA, PorcentajeIVA, IVA, CostoConIVA, Proveedor
            FROM Compras
            WHERE FechaCompra BETWEEN ? AND ?
        """
        params = [fecha_inicio, fecha_fin]

        if proveedor:
            query += " AND LOWER(Proveedor) LIKE ?"
            params.append(f"%{proveedor.lower()}%")

        cursor.execute(query, params)
        compras = cursor.fetchall()

        if not compras:
            messagebox.showinfo("Resultado", "No se encontraron compras con los filtros aplicados.")
            return

        ventana = tk.Toplevel(root)
        ventana.title("Compras Filtradas por Proveedor")
        ventana.geometry("1920x1080")

        frame_tree = ttk.Frame(ventana)
        frame_tree.pack(fill=tk.BOTH, expand=True)

        scrollbar_x = ttk.Scrollbar(frame_tree, orient=tk.HORIZONTAL)
        scrollbar_y = ttk.Scrollbar(frame_tree, orient=tk.VERTICAL)

        columnas = [
            "IdCompra", "FechaCompra", "CodigoProducto", "NombreProducto",
            "CantidadComprada", "CostoSinIVA", "PorcentajeIVA", "IVA",
            "CostoConIVA", "Proveedor"
        ]

        tree = ttk.Treeview(
            frame_tree, columns=columnas, show="headings",
            xscrollcommand=scrollbar_x.set, yscrollcommand=scrollbar_y.set
        )

        scrollbar_x.config(command=tree.xview)
        scrollbar_x.pack(side=tk.BOTTOM, fill=tk.X)
        scrollbar_y.config(command=tree.yview)
        scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)

        style = ttk.Style()
        style.configure("Treeview.Heading", font=("Arial", 13, "bold"))
        style.configure("Treeview", font=("Arial", 10))

        encabezados = [
            ("IdCompra", "ID Compra", 100),
            ("FechaCompra", "Fecha de Compra", 150),
            ("CodigoProducto", "Código Producto", 150),
            ("NombreProducto", "Nombre Producto", 300),
            ("CantidadComprada", "Cantidad Comprada", 150),
            ("CostoSinIVA", "Costo Sin IVA", 150),
            ("PorcentajeIVA", "Porcentaje IVA", 100),
            ("IVA", "Valor IVA", 150),
            ("CostoConIVA", "Costo Con IVA", 150),
            ("Proveedor", "Proveedor", 200)
        ]

        for col, text, width in encabezados:
            tree.heading(col, text=text)
            tree.column(col, width=width, anchor="center")

        tree.pack(fill=tk.BOTH, expand=True)

        total_costo_sin_iva = 0
        total_valor_iva = 0
        total_costo_con_iva = 0

        datos_exportacion = []

        for compra in compras:
            (
                id_compra, fecha, codigo, nombre, cantidad,
                costo_sin_iva, iva, valor_iva, costo_con_iva, proveedor
            ) = compra

            fecha = fecha.strftime("%Y-%m-%d") if fecha else "N/A"
            costo_sin_iva = float(costo_sin_iva or 0)
            valor_iva = float(valor_iva or 0)
            costo_con_iva = float(costo_con_iva or 0)

            total_costo_sin_iva += costo_sin_iva
            total_valor_iva += valor_iva
            total_costo_con_iva += costo_con_iva

            tree.insert("", tk.END, values=(
                id_compra, fecha, codigo, nombre, cantidad,
                f"{costo_sin_iva:,.2f}",
                f"{iva:.2%}" if iva is not None else "0.00%",
                f"{valor_iva:,.2f}",
                f"{costo_con_iva:,.2f}",
                proveedor if proveedor else "N/A"
            ))

            datos_exportacion.append([
                id_compra, fecha, codigo, nombre, cantidad,
                costo_sin_iva, iva, valor_iva, costo_con_iva, proveedor
            ])

        total_frame = ttk.Frame(ventana)
        total_frame.pack(fill='x', padx=10, pady=10)

        total_text = f"""
        🔹 Totales del Rango de Fecha Seleccionado 🔹
        --------------------------------------------
        📊 Total Costo Sin IVA: ${total_costo_sin_iva:,.2f}
        💰 Total Valor IVA: ${total_valor_iva:,.2f}
        📦 Total Costo Con IVA: ${total_costo_con_iva:,.2f}
        """

        text_area_totales = tk.Text(total_frame, height=13, wrap="word", font=("Arial", 12, "bold"))
        text_area_totales.insert(tk.END, total_text)
        text_area_totales.config(state="disabled")
        text_area_totales.pack(fill='both', expand=True)

        # Botones de exportación
        frame_botones_exportar = ttk.Frame(ventana)
        frame_botones_exportar.pack(pady=10)

        columnas_exportar = [col for col, _, _ in encabezados]

        btn_excel = ttk.Button(
            frame_botones_exportar,
            text="📥 Exportar a Excel",
            command=lambda: exportar_excel(datos_exportacion, columnas_exportar)
        )
        btn_excel.pack(side="left", padx=10)

        btn_pdf = ttk.Button(
            frame_botones_exportar,
            text="📄 Exportar a PDF",style="Blue.TButton",
            command=lambda: exportar_pdf(datos_exportacion, columnas_exportar, "Compras por Proveedor")
        )
        btn_pdf.pack(side="left", padx=10)

    except Exception as e:
        messagebox.showerror("Error", f"Ocurrió un error al buscar compras por proveedor: {e}")

# Frame de la pestaña "Listas"
frame_listas = ttk.Frame(tab_reportes)
frame_listas.pack(padx=10, pady=(1, 2))

# FILTRO POR FECHAS — CENTRADO Y SIN ESPACIO ENTRE ETIQUETA Y CAMPO
frame_fechas_listas = ttk.Frame(frame_listas)
frame_fechas_listas.grid(row=0, column=0, columnspan=4, pady=(10, 5), sticky="n")  # ⬅ centrado en columna

# Etiqueta y entrada de Fecha Inicio
ttk.Label(frame_fechas_listas, text="⏳ Fecha Inicio (YYYY-MM-DD):").pack(side="left", padx=(10, 5))
entry_fecha_inicio = DateEntry(frame_fechas_listas, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='yyyy-MM-dd')
entry_fecha_inicio.pack(side="left", padx=(0, 20))

# Etiqueta y entrada de Fecha Fin
ttk.Label(frame_fechas_listas, text="🚧 Fecha Fin (YYYY-MM-DD):").pack(side="left", padx=(10, 5))
entry_fecha_fin = DateEntry(frame_fechas_listas, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='yyyy-MM-dd')
entry_fecha_fin.pack(side="left", padx=(0, 10))


# - BUSCAR COMPRAS POR CÓDIGO 
ttk.Label(frame_listas, text="🛒 Buscar Compras por Código de Producto:").grid(row=2, column=0, padx=5, pady=5, sticky="w")
entry_buscar_codigo = ttk.Entry(frame_listas, width=30)
entry_buscar_codigo.grid(row=2, column=1, padx=5, pady=5)
btn_buscar_compras = tk.Button(frame_listas, text="🔍 Buscar", command=buscar_compras, bg="lightgreen")
btn_buscar_compras.grid(row=2, column=2, padx=5, pady=5)

# - BUSCAR COMPRAS POR CÓDIGO 
ttk.Label(frame_listas, text="🛒 Buscar Compras por Proveedor:").grid(row=3, column=0, padx=5, pady=(1,2), sticky="w")
entry_buscar_proveedor = ttk.Entry(frame_listas, width=30)
entry_buscar_proveedor.grid(row=3, column=1, padx=5, pady=(1,2))
btn_buscar_compras2 = tk.Button(frame_listas, text="🔍 Buscar", command=buscar_compras2,bg="lightgreen")
btn_buscar_compras2.grid(row=3, column=2, padx=5, pady=(1,2))



# FILTROS ADICIONALES
ttk.Label(frame_listas, text="🏷️ Código Producto:").grid(row=4, column=0, padx=5, pady=5, sticky="w")
entry_codigo_producto = ttk.Entry(frame_listas, width=20)
entry_codigo_producto.grid(row=4, column=1, padx=5, pady=5)

ttk.Label(frame_listas, text="👤 Consultor:").grid(row=4, column=2, padx=5, pady=5, sticky="w")
entry_consultor = ttk.Entry(frame_listas, width=20)
entry_consultor.grid(row=4, column=3, padx=5, pady=5)

ttk.Label(frame_listas, text="👤 Cliente:").grid(row=5, column=0, padx=5, pady=5, sticky="w")
entry_cliente = ttk.Entry(frame_listas, width=20)
entry_cliente.grid(row=5, column=1, padx=5, pady=5)

ttk.Label(frame_listas, text="👤 Padre Empresarial:").grid(row=5, column=2, padx=5, pady=5, sticky="w")
entry_padre_empresarial = ttk.Entry(frame_listas, width=20)
entry_padre_empresarial.grid(row=5, column=3, padx=5, pady=5)

ttk.Label(frame_listas, text="🧾 Número Orden de Compra:").grid(row=6, column=0, padx=5, pady=5, sticky="w")
entry_num_orden_filtro = ttk.Entry(frame_listas, width=20)
entry_num_orden_filtro.grid(row=6, column=1, padx=5, pady=5)


ttk.Label(frame_listas, text="🌆 Ciudad:").grid(row=6, column=2, padx=5, pady=5, sticky="w")
entry_ciudad = ttk.Entry(frame_listas, width=20)
entry_ciudad.grid(row=6, column=3, padx=5, pady=5)



#  BOTÓN BUSCAR Ordenes de Compras 
btn_buscar_ordencompra = tk.Button(frame_listas, text="🔍 Buscar Ordenes de Compras", command=buscar_ordencompra,bg="lightgreen")
btn_buscar_ordencompra.grid(row=8, column=1, padx=5, pady=10)










root.mainloop()



